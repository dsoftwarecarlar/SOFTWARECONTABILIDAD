from __future__ import annotations

import argparse
import json
import re
import sys
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from bootstrap import bootstrap_vendor

bootstrap_vendor()

import openpyxl
import xlrd

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    category=UserWarning,
)

INVALID_XML_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def sanitize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = INVALID_XML_CONTROL_CHARS.sub(" ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_decimal_like(value: Any) -> float:
    normalized = re.sub(r"[^\d,.\-]", "", sanitize_text(value))
    if normalized == "":
        return 0.0

    has_comma = "," in normalized
    has_dot = "." in normalized

    if has_comma and has_dot:
        if normalized.rfind(".") > normalized.rfind(","):
            normalized = normalized.replace(",", "")
        else:
            normalized = normalized.replace(".", "").replace(",", ".")
    elif has_comma:
        comma_count = normalized.count(",")
        if comma_count > 1:
            normalized = normalized.replace(",", "")
        else:
            int_part, _, frac_part = normalized.partition(",")
            if len(frac_part) == 3:
                normalized = f"{int_part}{frac_part}"
            else:
                normalized = f"{int_part}.{frac_part}"
    elif has_dot:
        dot_count = normalized.count(".")
        if dot_count > 1:
            last_dot = normalized.rfind(".")
            int_part = normalized[:last_dot].replace(".", "")
            frac_part = normalized[last_dot + 1 :]
            if len(frac_part) == 3:
                normalized = f"{int_part}{frac_part}"
            else:
                normalized = f"{int_part}.{frac_part}"

    try:
        return float(normalized)
    except ValueError:
        return 0.0


def excel_date_serial(value: date | datetime) -> float:
    current = value.date() if isinstance(value, datetime) else value
    return float((current - date(1899, 12, 30)).days)


def parse_date_text(value: Any) -> float | None:
    text = sanitize_text(value)
    if text == "":
        return None

    match = re.fullmatch(r"(\d{1,2})[\/.\-](\d{1,2})[\/.\-](\d{2,4})", text)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        raw_year = match.group(3)
        year = int(f"20{raw_year}" if len(raw_year) == 2 else raw_year)
        try:
            return excel_date_serial(date(year, month, day))
        except ValueError:
            return None

    for parser in ("%Y-%m-%d", "%Y/%m/%d", "%d %b %Y", "%d %B %Y"):
        try:
            return excel_date_serial(datetime.strptime(text, parser).date())
        except ValueError:
            continue

    try:
        return excel_date_serial(datetime.fromisoformat(text).date())
    except ValueError:
        return None


def parse_report_date_to_excel_serial(value: Any) -> float | None:
    text = sanitize_text(value).upper()
    match = re.fullmatch(r"(\d{2})-([A-Z]{3})-(\d{2})", text)
    if match is None:
        return None

    month_map = {
        "JAN": 1,
        "ENE": 1,
        "FEB": 2,
        "MAR": 3,
        "APR": 4,
        "ABR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AUG": 8,
        "AGO": 8,
        "SEP": 9,
        "SET": 9,
        "OCT": 10,
        "NOV": 11,
        "DEC": 12,
        "DIC": 12,
    }

    month = month_map.get(match.group(2))
    if month is None:
        return None

    try:
        return excel_date_serial(date(2000 + int(match.group(3)), month, int(match.group(1))))
    except ValueError:
        return None


def trim_document(value: Any) -> str:
    text = sanitize_text(value)
    if text == "":
        return ""

    digits = re.sub(r"[^\d]", "", text)
    if digits == "":
        return text

    trimmed = digits.lstrip("0")
    return trimmed or "0"


def sanitize_preserving_internal_spaces(value: Any) -> str:
    text = "" if value is None else str(value)
    return text.replace("\ufeff", "").strip()


def normalize_brand_key(value: Any) -> str:
    text = sanitize_text(value).upper()
    return {
        "CHANGAN": "changan",
        "PEUGEOT": "peug",
        "SUZUKI": "szk",
        "TOYOTA": "tyt",
        "MATRIZ": "tyt",
    }.get(text, "")


def get_template_key(agency: Any, order: Any) -> str | None:
    agency_text = sanitize_text(agency).upper()
    order_text = sanitize_text(order).upper()

    if agency_text == "CHANGAN":
        return "changan"
    if agency_text == "PEUGEOT":
        return "peug"
    if agency_text == "MATRIZ":
        return "tyt"
    if agency_text == "SUZUKI AMBATO":
        return "szk"
    if agency_text == "SUZUKI RIOBAMBA":
        return "changan" if re.fullmatch(r"D\d+", order_text) else "szk"
    return None


def simple_zero_padded(value: float, number_format: str) -> str | None:
    if not float(value).is_integer():
        return None

    candidate = sanitize_text(number_format)
    candidate = re.sub(r'".*?"', "", candidate)
    candidate = re.sub(r"\[[^\]]+\]", "", candidate)
    candidate = candidate.replace("\\", "")
    if re.fullmatch(r"0+", candidate or ""):
        return f"{int(value):0{len(candidate)}d}"
    return None


def normalize_excel_format(number_format: str) -> str:
    normalized = sanitize_text(number_format)
    normalized = re.sub(r'".*?"', "", normalized)
    normalized = re.sub(r"\[[^\]]+\]", "", normalized)
    normalized = normalized.replace("\\", "")
    normalized = normalized.replace("_-", "").replace("_ ", "")
    normalized = normalized.replace("*", "")
    return normalized


def format_numeric_with_excel_mask(value: float, number_format: str) -> str | None:
    normalized = normalize_excel_format(number_format)
    match = re.fullmatch(r"[#0,]+(?:\.([#0]+))?", normalized)
    if match is None:
        return None

    decimals = len(match.group(1) or "")
    use_grouping = "," in normalized.split(".", 1)[0]
    format_spec = f",.{decimals}f" if use_grouping else f".{decimals}f"
    return format(value, format_spec)


def format_date_with_excel_mask(value: date | datetime, number_format: str) -> str:
    current = value if isinstance(value, datetime) else datetime(value.year, value.month, value.day)
    normalized = normalize_excel_format(number_format).lower()
    month_abbrev = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    if "dd-mmm-yy" in normalized:
        return f"{current.day:02d}-{month_abbrev[current.month - 1]}-{current.year % 100:02d}"
    if "dd/mm/yyyy" in normalized or "dd/mm/yy" in normalized:
        year = f"{current.year:04d}" if "yyyy" in normalized else f"{current.year % 100:02d}"
        return f"{current.day:02d}/{current.month:02d}/{year}"
    return current.strftime("%d/%m/%Y")


@dataclass
class SheetReader:
    name: str
    max_row: int
    max_col: int

    def text(self, row: int, column: int) -> str:
        raise NotImplementedError

    def number(self, row: int, column: int) -> float:
        raise NotImplementedError

    def date_value(self, row: int, column: int) -> float | None:
        raise NotImplementedError

    def row_values(self, row: int) -> list[str]:
        return [self.text(row, column) for column in range(1, self.max_col + 1)]


class OpenPyxlSheetReader(SheetReader):
    def __init__(self, sheet: Any) -> None:
        super().__init__(name=str(sheet.title), max_row=int(sheet.max_row or 0), max_col=int(sheet.max_column or 0))
        self.sheet = sheet

    def _cell(self, row: int, column: int) -> Any:
        return self.sheet.cell(row=row, column=column)

    def text(self, row: int, column: int) -> str:
        cell = self._cell(row, column)
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, (datetime, date)):
            return format_date_with_excel_mask(value, str(cell.number_format or ""))
        if isinstance(value, (int, float)):
            padded = simple_zero_padded(float(value), str(cell.number_format or ""))
            if padded is not None:
                return padded
            formatted = format_numeric_with_excel_mask(float(value), str(cell.number_format or ""))
            if formatted is not None:
                return formatted
            if float(value).is_integer():
                return str(int(value))
        return sanitize_text(value)

    def number(self, row: int, column: int) -> float:
        cell = self._cell(row, column)
        value = cell.value
        if value is None or value == "":
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        return parse_decimal_like(value)

    def date_value(self, row: int, column: int) -> float | None:
        cell = self._cell(row, column)
        value = cell.value
        if value is None or value == "":
            return None
        if isinstance(value, (datetime, date)):
            return excel_date_serial(value)
        if isinstance(value, (int, float)):
            return float(value)
        return parse_date_text(value)


class XlrdSheetReader(SheetReader):
    def __init__(self, book: xlrd.book.Book, sheet: xlrd.sheet.Sheet) -> None:
        super().__init__(name=str(sheet.name), max_row=int(sheet.nrows or 0), max_col=int(sheet.ncols or 0))
        self.book = book
        self.sheet = sheet

    def _cell(self, row: int, column: int) -> xlrd.sheet.Cell:
        return self.sheet.cell(row - 1, column - 1)

    def _format_string(self, row: int, column: int) -> str:
        try:
            cell = self._cell(row, column)
            xf = self.book.xf_list[cell.xf_index]
            format_obj = self.book.format_map.get(xf.format_key)
            return sanitize_text(format_obj.format_str if format_obj is not None else "")
        except Exception:
            return ""

    def text(self, row: int, column: int) -> str:
        cell = self._cell(row, column)
        value = cell.value
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(value, self.book.datemode).strftime("%d/%m/%Y")
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            padded = simple_zero_padded(float(value), self._format_string(row, column))
            if padded is not None:
                return padded
            if float(value).is_integer():
                return str(int(value))
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return "TRUE" if bool(value) else "FALSE"
        return sanitize_text(value)

    def number(self, row: int, column: int) -> float:
        cell = self._cell(row, column)
        value = cell.value
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return 0.0
        if cell.ctype in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_DATE):
            return float(value)
        return parse_decimal_like(value)

    def date_value(self, row: int, column: int) -> float | None:
        cell = self._cell(row, column)
        value = cell.value
        if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
            return None
        if cell.ctype == xlrd.XL_CELL_DATE:
            return float(value)
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            return float(value)
        return parse_date_text(value)


def open_sheet_reader(input_path: Path) -> SheetReader:
    extension = input_path.suffix.lower()
    if extension in {".xlsx", ".xlsm"}:
        workbook = openpyxl.load_workbook(input_path, read_only=False, data_only=False)
        sheet_name = workbook.sheetnames[0] if workbook.sheetnames else None
        if sheet_name is None:
            raise RuntimeError("El archivo fuente no contiene hojas.")
        return OpenPyxlSheetReader(workbook[sheet_name])

    if extension == ".xls":
        workbook = xlrd.open_workbook(str(input_path), on_demand=True)
        if workbook.nsheets <= 0:
            raise RuntimeError("El archivo fuente no contiene hojas.")
        return XlrdSheetReader(workbook, workbook.sheet_by_index(0))

    raise RuntimeError(f"Extension no soportada para lectura: {input_path.suffix}")


def matches_header_checks(sheet: SheetReader, checks: list[tuple[int, int, str]]) -> bool:
    for row, column, needle in checks:
        actual = sheet.text(row, column).upper()
        if needle.upper() not in actual:
            return False
    return True


def validate_headers(sheet: SheetReader) -> str:
    modern_checks = [
        (1, 1, "AGENCIA"),
        (1, 2, "CENTRO"),
        (1, 3, "No. ORDEN"),
        (1, 8, "TIPO DOC"),
        (1, 9, "CEDULA"),
        (1, 10, "FACTURADO A"),
        (1, 12, "DOCUMENTO"),
        (1, 15, "F. FACT"),
        (1, 18, "F. NOTA"),
        (1, 36, "ANULADA"),
    ]
    if matches_header_checks(sheet, modern_checks):
        return "modern"

    legacy_checks = [
        (1, 1, "AGENCIA"),
        (1, 2, "CENTRO"),
        (1, 3, "No. ORDEN"),
        (1, 7, "FACTURA"),
        (1, 8, "F. FACT"),
        (1, 9, "F. NOTA"),
        (1, 11, "TOTAL MANO OBRA"),
        (1, 20, "C. COSTO"),
    ]
    if matches_header_checks(sheet, legacy_checks):
        return "legacy"

    row, column, needle = modern_checks[0]
    actual = sheet.text(row, column).upper()
    for candidate_row, candidate_column, candidate_needle in modern_checks:
        candidate_actual = sheet.text(candidate_row, candidate_column).upper()
        if candidate_needle.upper() not in candidate_actual:
            row, column, needle, actual = candidate_row, candidate_column, candidate_needle, candidate_actual
            break
    raise RuntimeError(
        f"El archivo fuente no coincide con la estructura esperada en fila {row} columna {column}. "
        f"Esperado contiene '{needle}' y llego '{actual}'."
    )


def read_source_rows(input_path: Path) -> tuple[str, list[dict[str, Any]]]:
    sheet = open_sheet_reader(input_path)
    layout = validate_headers(sheet)
    if layout == "modern":
        return read_modern_source_rows(sheet)
    if layout == "legacy":
        return read_legacy_source_rows(sheet)
    raise RuntimeError(f"Layout de fuente no soportado: {layout}")


def read_modern_source_rows(sheet: SheetReader) -> tuple[str, list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []

    for row in range(2, sheet.max_row + 1):
        agency = sheet.text(row, 1)
        if agency == "":
            continue

        order = sheet.text(row, 3)
        series = sheet.text(row, 14) or sheet.text(row, 13)
        template_key = get_template_key(agency, order)
        if template_key is None:
            continue

        doc_type = sheet.text(row, 8).upper()
        if doc_type not in {"FA", "FC", "DC", "DE"}:
            continue

        anulada = sheet.text(row, 36).upper()
        if anulada in {"SI", "S", "YES", "Y", "ANULADA"}:
            continue

        document_raw = sheet.text(row, 12)
        if document_raw == "":
            continue

        affected_raw = sheet.text(row, 37)

        rows.append(
            {
                "RowIndex": row,
                "TemplateKey": template_key,
                "Agency": agency,
                "AgencyRaw": sheet.text(row, 1),
                "Center": sheet.text(row, 2),
                "CenterRaw": sheet.text(row, 2),
                "Order": order,
                "OrderRaw": order,
                "Advisor": sheet.text(row, 5),
                "AdvisorRaw": sheet.text(row, 5),
                "Line": sheet.text(row, 7),
                "LineRaw": sheet.text(row, 7),
                "DocType": doc_type,
                "Cedula": sheet.text(row, 9),
                "CedulaRaw": sheet.text(row, 9),
                "Customer": sheet.text(row, 10),
                "CustomerRaw": sheet.text(row, 10),
                "DocumentRaw": document_raw,
                "DocumentTrim": trim_document(document_raw),
                "Series": series,
                "SeriesRaw": series,
                "FormaPago": sheet.text(row, 16),
                "Authorization": sheet.text(row, 17),
                "DateFactValue": sheet.date_value(row, 15),
                "DateNoteValue": sheet.date_value(row, 18),
                "NoteCredit": sheet.number(row, 19),
                "TotalManoObra": sheet.number(row, 20),
                "TotalSubcontratos": sheet.number(row, 21),
                "TotalInsumos": sheet.number(row, 22),
                "TotalServicio": sheet.number(row, 23),
                "TotalAccesorios": sheet.number(row, 24),
                "TotalRepuestos": sheet.number(row, 25),
                "Interes": sheet.number(row, 26),
                "Iva": sheet.number(row, 27),
                "Total": sheet.number(row, 28),
                "Costo": sheet.number(row, 29),
                "CostoLubricantes": sheet.number(row, 30),
                "CostoAccesorios": sheet.number(row, 31),
                "CostoRepuestos": sheet.number(row, 32),
                "CostoPintura": sheet.number(row, 33),
                "CostoSubconNc": sheet.number(row, 34),
                "GarExt": sheet.text(row, 35),
                "GarExtRaw": sheet.text(row, 35),
                "Anulada": anulada,
                "AffectedDocumentTrim": trim_document(affected_raw),
                "AffectedDocumentRaw": affected_raw,
                "MotivoNc": sheet.text(row, 38),
                "ObservacionNc": sheet.text(row, 39),
            }
        )

    return sheet.name, rows


def legacy_order_base(value: Any) -> str:
    text = sanitize_text(value)
    return re.sub(r"[A-Z]+$", "", text.upper())


def infer_legacy_doc_type(sheet: SheetReader, row: int) -> str:
    note_marker = sheet.number(row, 10)
    note_date = sheet.date_value(row, 9)
    fact_date = sheet.date_value(row, 8)
    total_value = sheet.number(row, 19)
    if note_date is not None and fact_date in (None, ""):
        return "DC"
    if note_marker < 0 or total_value < 0:
        return "DC"
    return "FA"


def infer_legacy_cost_total(sheet: SheetReader, row: int) -> float:
    return float(sum(abs(sheet.number(row, column)) for column in range(20, 26)))


def attach_legacy_affected_documents(rows: list[dict[str, Any]]) -> None:
    invoice_rows_by_order_base: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        if sanitize_text(row.get("DocType")).upper() not in {"FA", "FC"}:
            continue
        order_base = legacy_order_base(row.get("Order"))
        if order_base == "":
            continue
        invoice_rows_by_order_base.setdefault(order_base, []).append(row)

    for row in rows:
        if sanitize_text(row.get("DocType")).upper() not in {"DC", "DE"}:
            continue
        if sanitize_text(row.get("AffectedDocumentTrim")) != "":
            continue

        order_base = legacy_order_base(row.get("Order"))
        if order_base == "":
            continue

        candidates = invoice_rows_by_order_base.get(order_base, [])
        if not candidates:
            continue

        note_order = sanitize_text(row.get("Order")).upper()
        note_customer = sanitize_text(row.get("Customer")).upper()
        note_cedula = sanitize_text(row.get("Cedula"))
        note_date = row.get("DateNoteValue") or row.get("DateFactValue")
        note_row_index = int(row.get("RowIndex", 0) or 0)

        def candidate_score(candidate: dict[str, Any]) -> tuple[int, int, int, int, int]:
            candidate_order = sanitize_text(candidate.get("Order")).upper()
            candidate_customer = sanitize_text(candidate.get("Customer")).upper()
            candidate_cedula = sanitize_text(candidate.get("Cedula"))
            candidate_date = candidate.get("DateFactValue") or candidate.get("DateNoteValue")
            candidate_row_index = int(candidate.get("RowIndex", 0) or 0)
            exact_order_penalty = 0 if candidate_order == note_order else 1
            cedula_penalty = 0 if note_cedula != "" and candidate_cedula == note_cedula else 1
            customer_penalty = 0 if note_customer != "" and candidate_customer == note_customer else 1
            future_penalty = 0 if note_date in (None, "") or candidate_date in (None, "") or float(candidate_date) <= float(note_date) else 1
            distance_penalty = abs(candidate_row_index - note_row_index)
            return (exact_order_penalty, cedula_penalty, customer_penalty, future_penalty, distance_penalty)

        selected = sorted(candidates, key=candidate_score)[0]
        row["AffectedDocumentRaw"] = selected.get("DocumentRaw", "")
        row["AffectedDocumentTrim"] = selected.get("DocumentTrim", "")


def read_legacy_source_rows(sheet: SheetReader) -> tuple[str, list[dict[str, Any]]]:
    rows: list[dict[str, Any]] = []

    for row in range(2, sheet.max_row + 1):
        agency = sheet.text(row, 1)
        if agency == "":
            continue

        order = sheet.text(row, 3)
        template_key = get_template_key(agency, order)
        if template_key is None:
            continue

        doc_type = infer_legacy_doc_type(sheet, row)
        if doc_type not in {"FA", "FC", "DC", "DE"}:
            continue

        document_raw = sheet.text(row, 7)
        if document_raw == "":
            continue

        date_fact_value = sheet.date_value(row, 8)
        date_note_value = sheet.date_value(row, 9)
        total_cost = infer_legacy_cost_total(sheet, row)

        rows.append(
            {
                "RowIndex": row,
                "TemplateKey": template_key,
                "Agency": agency,
                "AgencyRaw": sheet.text(row, 1),
                "Center": sheet.text(row, 2),
                "CenterRaw": sheet.text(row, 2),
                "Order": order,
                "OrderRaw": order,
                "Advisor": "",
                "AdvisorRaw": "",
                "Line": sheet.text(row, 4),
                "LineRaw": sheet.text(row, 4),
                "DocType": doc_type,
                "Cedula": sheet.text(row, 5),
                "CedulaRaw": sheet.text(row, 5),
                "Customer": sheet.text(row, 6),
                "CustomerRaw": sheet.text(row, 6),
                "DocumentRaw": document_raw,
                "DocumentTrim": trim_document(document_raw),
                "Series": "",
                "SeriesRaw": "",
                "FormaPago": "",
                "Authorization": "",
                "DateFactValue": date_fact_value,
                "DateNoteValue": date_note_value,
                "NoteCredit": 0.0,
                "TotalManoObra": sheet.number(row, 11),
                "TotalSubcontratos": sheet.number(row, 12),
                "TotalInsumos": sheet.number(row, 13),
                "TotalServicio": sheet.number(row, 14),
                "TotalAccesorios": sheet.number(row, 15),
                "TotalRepuestos": sheet.number(row, 16),
                "Interes": sheet.number(row, 17),
                "Iva": sheet.number(row, 18),
                "Total": sheet.number(row, 19),
                "Costo": total_cost,
                "CostoLubricantes": 0.0,
                "CostoAccesorios": 0.0,
                "CostoRepuestos": 0.0,
                "CostoPintura": 0.0,
                "CostoSubconNc": 0.0,
                "GarExt": "",
                "GarExtRaw": "",
                "Anulada": "",
                "AffectedDocumentTrim": "",
                "AffectedDocumentRaw": "",
                "MotivoNc": "",
                "ObservacionNc": "",
            }
        )

    attach_legacy_affected_documents(rows)
    return sheet.name, rows


def read_px_rows(input_path: Path, brand_key: str) -> tuple[str, list[list[str]]]:
    sheet = open_sheet_reader(input_path)
    collected: list[list[str]] = []
    capture = False

    for row_index in range(1, sheet.max_row + 1):
        row = sheet.row_values(row_index)
        marker = sanitize_text(row[1] if len(row) > 1 else "").upper()
        if marker == "MARCA:":
            current_brand = normalize_brand_key(row[3] if len(row) > 3 else "")
            capture = current_brand == brand_key
            continue

        if not capture:
            continue

        collected.append(row)

    return sheet.name, collected


def read_mayor_rows(input_path: Path) -> list[dict[str, Any]]:
    lines = [
        line.replace("\ufeff", "").rstrip()
        for line in input_path.read_text(encoding="utf-8", errors="ignore").splitlines()
    ]
    rows: list[dict[str, Any]] = []
    for line in lines:
        if sanitize_text(line) == "":
            continue

        columns = line.split("\t")
        if len(columns) < 30:
            continue

        account = sanitize_preserving_internal_spaces(columns[6])
        date_text = sanitize_preserving_internal_spaces(columns[22]).upper()
        if re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", account) is None:
            continue
        if re.fullmatch(r"\d{2}-[A-Z]{3}-\d{2}", date_text) is None:
            continue

        rows.append(
            {
                "account": account,
                "name": sanitize_preserving_internal_spaces(columns[7]),
                "ext": sanitize_preserving_internal_spaces(columns[21]) or "N",
                "date_text": date_text,
                "date_value": parse_report_date_to_excel_serial(date_text),
                "origin": sanitize_preserving_internal_spaces(columns[23]).upper(),
                "seat": sanitize_preserving_internal_spaces(columns[24]),
                "reference": sanitize_preserving_internal_spaces(columns[25]),
                "detail": sanitize_preserving_internal_spaces(columns[26]),
                "debit": parse_decimal_like(columns[27]),
                "credit": parse_decimal_like(columns[28]),
                "balance": parse_decimal_like(columns[29]),
            }
        )

    return rows


def write_output(output_path: Path, payload: dict[str, Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(f"{json.dumps(payload, indent=2, ensure_ascii=True)}\n", encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Lectores Python para Servicios por Marca")
    subparsers = parser.add_subparsers(dest="command", required=True)

    source_parser = subparsers.add_parser("source")
    source_parser.add_argument("--input", required=True)
    source_parser.add_argument("--output-json", required=True)

    px_parser = subparsers.add_parser("px")
    px_parser.add_argument("--input", required=True)
    px_parser.add_argument("--brand", required=True)
    px_parser.add_argument("--output-json", required=True)

    mayor_parser = subparsers.add_parser("mayor")
    mayor_parser.add_argument("--input", required=True)
    mayor_parser.add_argument("--output-json", required=True)

    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        input_path = Path(args.input).resolve()
        output_path = Path(getattr(args, "output_json")).resolve()
        if not input_path.is_file():
            raise RuntimeError(f"No existe archivo de entrada: {input_path}")

        if args.command == "source":
            sheet_name, rows = read_source_rows(input_path)
            write_output(output_path, {"sheet_name": sheet_name, "rows": rows})
            print(f"INFO|source_read|rows={len(rows)}")
            return 0

        if args.command == "px":
            brand_key = sanitize_text(getattr(args, "brand", "")).lower()
            if brand_key == "":
                raise RuntimeError("Falta --brand para filtrar la seccion correspondiente.")

            sheet_name, rows = read_px_rows(input_path, brand_key)
            write_output(output_path, {"sheet_name": sheet_name, "rows": rows})
            print(f"INFO|px_read|rows={len(rows)}|brand={brand_key}")
            return 0

        if args.command == "mayor":
            rows = read_mayor_rows(input_path)
            write_output(output_path, {"rows": rows})
            print(f"INFO|mayor_read|rows={len(rows)}|file={input_path.name}")
            return 0

        raise RuntimeError(f"Comando no soportado: {args.command}")
    except Exception as exc:  # noqa: BLE001
        print(str(exc), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
