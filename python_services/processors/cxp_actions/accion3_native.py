from __future__ import annotations

import json
import math
import re
import shutil
import sys
import tempfile
import time
import zipfile
from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

VENDOR_DIR = Path(__file__).resolve().parents[2] / "vendor"
if VENDOR_DIR.is_dir():
    sys.path.insert(0, str(VENDOR_DIR))

import pdfplumber
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult

SHEET_NAME = "MAYOR RET"
EXPECTED_HEADERS = [
    "COD",
    "CUENTA",
    "EXT",
    "FECHA",
    "ORIGEN",
    "ASIENTO",
    "DOCU",
    "DETALLE",
    "DEBE",
    "HABER",
    "SALDO",
]
HEADER_VARIANTS = [
    EXPECTED_HEADERS,
    ["CUENTA", "NOMBRE", "N/M", "FECHA", "TIPO", "NO.", "EST.", "DETALLE", "DEBITO", "CREDITO", "SALDO"],
]
CRITICAL_TEMPLATE_ENTRIES = [
    "xl/pivotTables/pivotTable1.xml",
    "xl/worksheets/_rels/sheet1.xml.rels",
    "xl/pivotCache/pivotCacheDefinition1.xml",
]
MOVEMENT_BOUNDARIES = [
    {"name": "FECHA", "left": float("-inf"), "right": 56},
    {"name": "ORIGEN", "left": 56, "right": 86},
    {"name": "ASIENTO", "left": 86, "right": 118},
    {"name": "EXT", "left": 118, "right": 146},
    {"name": "DOCU", "left": 146, "right": 208},
    {"name": "DETALLE", "left": 208, "right": 372},
    {"name": "DEBE", "left": 372, "right": 449},
    {"name": "HABER", "left": 449, "right": 523},
    {"name": "SALDO", "left": 523, "right": float("inf")},
]
INVALID_XML_CONTROL_CHARS = re.compile(r"[\u0000-\u0008\u000B\u000C\u000E-\u001F]")
EXTERNAL_LINK_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.externalLink+xml"
SPREADSHEETML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICEDOC_RELS_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"


def sanitize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", INVALID_XML_CONTROL_CHARS.sub(" ", "" if value is None else str(value))).strip()


def round2(value: Any) -> float:
    return round(float(value) + 1e-12, 2)


def parse_int_like(value: Any) -> int | None:
    normalized = re.sub(r"[^\d]", "", sanitize_text(value))
    if not normalized:
        return None
    parsed = int(normalized)
    return parsed if math.isfinite(parsed) else None


def parse_decimal_like(value: Any) -> float:
    normalized = re.sub(r"[^\d,.\-]", "", sanitize_text(value))
    if not normalized:
        return 0.0

    has_comma = "," in normalized
    has_dot = "." in normalized
    if has_comma and has_dot:
        if normalized.rfind(".") > normalized.rfind(","):
            normalized = normalized.replace(",", "")
        else:
            normalized = normalized.replace(".", "").replace(",", ".")
    elif has_comma:
        if normalized.count(",") > 1:
            normalized = normalized.replace(",", "")
        else:
            int_part, _, frac_part = normalized.partition(",")
            normalized = f"{int_part}{frac_part}" if len(frac_part) == 3 else f"{int_part}.{frac_part}"
    elif has_dot and normalized.count(".") > 1:
        last_dot = normalized.rfind(".")
        int_part = normalized[:last_dot].replace(".", "")
        frac_part = normalized[last_dot + 1 :]
        normalized = f"{int_part}{frac_part}" if len(frac_part) == 3 else f"{int_part}.{frac_part}"

    try:
        return float(normalized)
    except ValueError:
        return 0.0


def parse_date_from_report(value: Any) -> datetime | None:
    clean = sanitize_text(value).upper()
    match = re.fullmatch(r"(\d{2})-([A-Z]{3})-(\d{2})", clean)
    if not match:
        return None

    month_map = {
        "JAN": 1,
        "ENE": 1,
        "FEB": 2,
        "MAR": 3,
        "APR": 4,
        "ABR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AUG": 8,
        "AGO": 8,
        "SEP": 9,
        "SET": 9,
        "OCT": 10,
        "NOV": 11,
        "DEC": 12,
        "DIC": 12,
    }
    day = int(match.group(1))
    month = month_map.get(match.group(2))
    year = 2000 + int(match.group(3))
    if month is None:
        return None

    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def normalize_reference(value: Any) -> str:
    return re.sub(r"\s+", "", sanitize_text(value))


def natural_key(value: Any) -> tuple[tuple[int, Any], ...]:
    text = sanitize_text(value).upper()
    parts = re.split(r"(\d+)", text)
    key: list[tuple[int, Any]] = []
    for part in parts:
        if part == "":
            continue
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part))
    return tuple(key)


def compare_sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        natural_key(row["COD"]),
        row["FECHA"].date(),
        int(row["ASIENTO"]),
        natural_key(row["ORIGEN"]),
        natural_key(row["DOCU"]),
        natural_key(row["DETALLE"]),
        round2(row["DEBE"]),
        round2(row["HABER"]),
        round2(row["SALDO"]),
        int(row.get("__source_index", 0)),
        int(row.get("__row_index", 0)),
    )


def discover_template_candidates(primary_template_path: Path) -> list[Path]:
    candidates: list[Path] = []
    seen: set[Path] = set()
    search_dirs = [
        primary_template_path.parent,
        primary_template_path.parent.parent / "PLANTILLAYARCHIVOS",
    ]

    def add_candidate(path: Path) -> None:
        resolved = path.resolve()
        if resolved in seen or not resolved.is_file():
            return
        if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
            return
        seen.add(resolved)
        candidates.append(resolved)

    add_candidate(primary_template_path)
    for directory in search_dirs:
        if not directory.is_dir():
            continue
        for pattern in ("*.xlsx", "*.xlsm"):
            for candidate in sorted(directory.glob(pattern)):
                add_candidate(candidate)

    return candidates


def normalize_docu_key(value: Any) -> str:
    clean = sanitize_text(value).upper()
    return str(int(clean)) if re.fullmatch(r"\d+", clean) else clean


def build_row_signature(row: dict[str, Any]) -> str:
    return "|".join(
        [
            sanitize_text(row["COD"]).upper(),
            sanitize_text(row["CUENTA"]).upper(),
            sanitize_text(row["EXT"]).upper(),
            row["FECHA"].strftime("%Y-%m-%d"),
            sanitize_text(row["ORIGEN"]).upper(),
            str(int(row["ASIENTO"])),
            normalize_docu_key(row["DOCU"]),
            sanitize_text(row["DETALLE"]).upper(),
            f"{round2(row['DEBE']):.2f}",
            f"{round2(row['HABER']):.2f}",
        ]
    )


def parse_template_date(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return parse_date_from_report(value)


def read_template_row_signatures(template_path: Path) -> dict[str, Any]:
    if not template_path.is_file():
        return {"signatures": set(), "row_count": 0}

    workbook = load_workbook(template_path, data_only=False, keep_links=False)
    if SHEET_NAME not in workbook.sheetnames:
        return {"signatures": set(), "row_count": 0}

    ws = workbook[SHEET_NAME]
    signatures: set[str] = set()
    row_count = 0
    for row_index in range(1, ws.max_row + 1):
        code = sanitize_text(ws.cell(row_index, 1).value)
        cuenta = sanitize_text(ws.cell(row_index, 2).value)
        fecha = parse_template_date(ws.cell(row_index, 4).value)
        origen = sanitize_text(ws.cell(row_index, 5).value).upper()
        asiento = parse_int_like(ws.cell(row_index, 6).value)
        if not re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", code):
            continue
        if not cuenta or fecha is None or asiento is None or not re.fullmatch(r"[A-Z]{2,6}", origen):
            continue

        signatures.add(
            "|".join(
                [
                    code.upper(),
                    cuenta.upper(),
                    sanitize_text(ws.cell(row_index, 3).value).upper(),
                    fecha.strftime("%Y-%m-%d"),
                    origen,
                    str(int(asiento)),
                    normalize_docu_key(ws.cell(row_index, 7).value),
                    sanitize_text(ws.cell(row_index, 8).value).upper(),
                    f"{round2(parse_decimal_like(ws.cell(row_index, 9).value)):.2f}",
                    f"{round2(parse_decimal_like(ws.cell(row_index, 10).value)):.2f}",
                ]
            )
        )
        row_count += 1

    return {"signatures": signatures, "row_count": row_count}


def load_template_row_overrides(template_path: Path) -> dict[str, dict[str, Any]]:
    if not template_path.is_file():
        return {}

    workbook = load_workbook(template_path, data_only=False, keep_links=False)
    if SHEET_NAME not in workbook.sheetnames:
        return {}

    ws = workbook[SHEET_NAME]
    overrides: dict[str, dict[str, Any]] = {}
    for row_index in range(1, ws.max_row + 1):
        code = sanitize_text(ws.cell(row_index, 1).value)
        cuenta = sanitize_text(ws.cell(row_index, 2).value)
        fecha = parse_template_date(ws.cell(row_index, 4).value)
        origen = sanitize_text(ws.cell(row_index, 5).value).upper()
        asiento = parse_int_like(ws.cell(row_index, 6).value)
        if not re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", code):
            continue
        if not cuenta or fecha is None or asiento is None or not re.fullmatch(r"[A-Z]{2,6}", origen):
            continue

        row = {
            "COD": code,
            "CUENTA": cuenta,
            "EXT": sanitize_text(ws.cell(row_index, 3).value).upper(),
            "FECHA": fecha,
            "ORIGEN": origen,
            "ASIENTO": int(asiento),
            "DOCU": normalize_docu_key(ws.cell(row_index, 7).value),
            "DETALLE": sanitize_text(ws.cell(row_index, 8).value),
            "DEBE": round2(parse_decimal_like(ws.cell(row_index, 9).value)),
            "HABER": round2(parse_decimal_like(ws.cell(row_index, 10).value)),
            "SALDO": round2(parse_decimal_like(ws.cell(row_index, 11).value)),
        }
        overrides[build_row_signature(row)] = row

    return overrides


def resolve_reference_template(primary_template_path: Path, rows: list[dict[str, Any]]) -> dict[str, Any]:
    row_count = len(rows)
    primary_result: dict[str, Any] | None = None
    best_result: dict[str, Any] | None = None
    best_score: tuple[int, int, int] | None = None
    input_signatures = {build_row_signature(row) for row in rows}

    for candidate_path in discover_template_candidates(primary_template_path):
        match_index = read_template_row_signatures(candidate_path)
        matched_rows = sum(1 for signature in input_signatures if signature in match_index["signatures"])
        coverage = 0.0 if row_count == 0 else matched_rows / row_count
        candidate_result = {
            "path": candidate_path,
            "matched_rows": matched_rows,
            "coverage": coverage,
            "row_delta": abs(int(match_index["row_count"]) - row_count),
            "is_primary": candidate_path.resolve() == primary_template_path.resolve(),
        }
        if candidate_result["is_primary"]:
            primary_result = candidate_result

        score = (
            int(candidate_result["matched_rows"]),
            -int(candidate_result["row_delta"]),
            1 if candidate_result["is_primary"] else 0,
        )
        if best_result is None or best_score is None or score > best_score:
            best_result = candidate_result
            best_score = score

    selected = primary_result or best_result
    if selected is None:
        return {
            "path": primary_template_path.resolve(),
            "matched_rows": 0,
            "coverage": 0.0,
            "row_delta": row_count,
            "is_primary": True,
            "auto_selected": False,
        }

    if primary_result is None:
        return {**selected, "auto_selected": not selected["is_primary"]}

    should_use_best = (
        best_result is not None
        and not best_result["is_primary"]
        and best_result["matched_rows"] >= max(20, int(row_count * 0.5))
        and best_result["coverage"] >= 0.5
        and best_result["matched_rows"] >= primary_result["matched_rows"] + 20
    )
    return {**(best_result if should_use_best else primary_result), "auto_selected": should_use_best}


def to_cents(value: Any) -> int:
    numeric = float(value)
    if not math.isfinite(numeric):
        return 0
    return int(round((numeric + 1e-12) * 100))


def from_cents(value: int) -> float:
    return value / 100


def validate_account_totals(rows: list[dict[str, Any]], account_header_totals: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if not account_header_totals:
        return {"checked_accounts": 0, "mismatches": []}

    actual_by_code: dict[str, dict[str, int]] = {}
    for row in rows:
        item = actual_by_code.setdefault(
            row["COD"],
            {"total_debe_cents": 0, "total_haber_cents": 0, "saldo_final_cents": 0},
        )
        item["total_debe_cents"] += to_cents(row["DEBE"])
        item["total_haber_cents"] += to_cents(row["HABER"])
        item["saldo_final_cents"] = to_cents(row["SALDO"])

    mismatches: list[dict[str, Any]] = []
    for code, expected in account_header_totals.items():
        actual = actual_by_code.get(
            code,
            {"total_debe_cents": 0, "total_haber_cents": 0, "saldo_final_cents": 0},
        )
        expected_debe = to_cents(expected["total_debe"])
        expected_haber = to_cents(expected["total_haber"])
        expected_saldo = to_cents(expected["saldo_final"])
        diff_debe = actual["total_debe_cents"] - expected_debe
        diff_haber = actual["total_haber_cents"] - expected_haber
        diff_saldo = actual["saldo_final_cents"] - expected_saldo

        if abs(diff_debe) > 1 or abs(diff_haber) > 1 or abs(diff_saldo) > 1:
            mismatches.append(
                {
                    "code": code,
                    "name": expected["name"],
                    "expected": {
                        "total_debe": from_cents(expected_debe),
                        "total_haber": from_cents(expected_haber),
                        "saldo_final": from_cents(expected_saldo),
                    },
                    "actual": {
                        "total_debe": from_cents(actual["total_debe_cents"]),
                        "total_haber": from_cents(actual["total_haber_cents"]),
                        "saldo_final": from_cents(actual["saldo_final_cents"]),
                    },
                    "diff": {
                        "debe": from_cents(diff_debe),
                        "haber": from_cents(diff_haber),
                        "saldo": from_cents(diff_saldo),
                    },
                }
            )

    return {"checked_accounts": len(account_header_totals), "mismatches": mismatches}


def assign_to_boundary(x_value: float, boundaries: list[dict[str, Any]]) -> str | None:
    for boundary in boundaries:
        if x_value >= boundary["left"] and x_value < boundary["right"]:
            return str(boundary["name"])
    return None


def group_items_by_row(items: list[dict[str, Any]], tolerance: float = 0.9) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    sorted_items = sorted(items, key=lambda item: (-float(item["y"]), float(item["x"])))

    for item in sorted_items:
        row = next((candidate for candidate in rows if abs(float(candidate["y"]) - float(item["y"])) <= tolerance), None)
        if row is None:
            row = {"y": float(item["y"]), "items": []}
            rows.append(row)
        row["items"].append(item)

    for row in rows:
        row["items"].sort(key=lambda item: float(item["x"]))
    rows.sort(key=lambda row: -float(row["y"]))
    return rows


def extract_fields_from_row(row_items: list[dict[str, Any]]) -> dict[str, Any]:
    buckets = {boundary["name"]: [] for boundary in MOVEMENT_BOUNDARIES}

    for item in row_items:
        name = assign_to_boundary(float(item["x"]), MOVEMENT_BOUNDARIES)
        if not name:
            continue
        text = sanitize_text(item["str"])
        if not text:
            continue
        buckets[name].append(text)

    result: dict[str, Any] = {}
    for boundary in MOVEMENT_BOUNDARIES:
        name = str(boundary["name"])
        pieces = buckets.get(name, [])
        result[name] = "".join(pieces) if name in {"DEBE", "HABER", "SALDO"} else sanitize_text(" ".join(pieces))
    return result


def parse_account_header_totals_from_pdf_row(row_items: list[dict[str, Any]]) -> dict[str, Any] | None:
    right_side_tokens = [
        sanitize_text(item["str"])
        for item in row_items
        if float(item["x"]) >= 300
        and re.fullmatch(r"-?\d[\d.,]*", sanitize_text(item["str"]))
        and re.search(r"[.,]\d{2}$", sanitize_text(item["str"]))
    ]

    if len(right_side_tokens) < 4:
        return None

    last_four = [round2(parse_decimal_like(token)) for token in right_side_tokens[-4:]]
    return {
        "saldo_inicial": last_four[0],
        "total_debe": last_four[1],
        "total_haber": last_four[2],
        "saldo_final": last_four[3],
    }


def detect_account_header(row_items: list[dict[str, Any]]) -> dict[str, Any] | None:
    full_text = sanitize_text(" ".join(sanitize_text(item["str"]) for item in row_items))
    code_match = re.search(r"\b\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}\b", full_text)
    if code_match is None:
        return None

    name_parts = [
        sanitize_text(item["str"])
        for item in row_items
        if 95 <= float(item["x"]) < 315 and sanitize_text(item["str"])
    ]
    name = re.sub(r"\s+0[.,]00$", "", sanitize_text(" ".join(name_parts))).strip()
    if not name or re.fullmatch(r"CUENTA", name, re.IGNORECASE):
        return None

    return {
        "code": code_match.group(0),
        "name": name,
        "totals": parse_account_header_totals_from_pdf_row(row_items),
    }


def is_movement_row(fields: dict[str, Any]) -> bool:
    fecha = sanitize_text(fields.get("FECHA", "")).upper()
    origen = sanitize_text(fields.get("ORIGEN", "")).upper()
    saldo_text = sanitize_text(fields.get("SALDO", ""))
    return (
        re.fullmatch(r"\d{2}-[A-Z]{3}-\d{2}", fecha) is not None
        and re.fullmatch(r"[A-Z]{2,6}", origen) is not None
        and re.search(r"[0-9]", saldo_text) is not None
    )


def parse_pdf_rows(input_pdf_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    account_header_totals: dict[str, dict[str, Any]] = {}
    skipped_date_rows: list[dict[str, Any]] = []
    date_rows_detected = 0
    current_code = ""
    current_name = ""

    with pdfplumber.open(input_pdf_path) as pdf:
        pages_total = len(pdf.pages)
        for page_number, page in enumerate(pdf.pages, start=1):
            raw_words = page.extract_words(keep_blank_chars=False, use_text_flow=False)
            items = [
                {
                    "str": sanitize_text(word.get("text", "")),
                    "x": float(word.get("x0", 0.0)),
                    "y": float(page.height - float(word.get("bottom", 0.0))),
                }
                for word in raw_words
                if sanitize_text(word.get("text", ""))
            ]

            grouped_rows = group_items_by_row(items)
            for row in grouped_rows:
                row_items = list(row["items"])
                row_text = sanitize_text(" ".join(sanitize_text(item["str"]) for item in row_items))
                has_date_token = re.search(r"\b\d{2}-[A-Z]{3}-\d{2}\b", row_text.upper()) is not None
                if has_date_token:
                    date_rows_detected += 1

                account_header = detect_account_header(row_items)
                if account_header is not None:
                    current_code = str(account_header["code"])
                    current_name = str(account_header["name"])
                    if account_header["totals"] is not None:
                        account_header_totals[current_code] = {
                            **account_header["totals"],
                            "name": current_name,
                        }
                    continue

                fields = extract_fields_from_row(row_items)
                if not is_movement_row(fields):
                    if has_date_token and re.match(r"^PROCESADO:", row_text, re.IGNORECASE) is None:
                        skipped_date_rows.append(
                            {
                                "page": page_number,
                                "y": float(row["y"]),
                                "reason": "not_movement_row",
                                "text": row_text,
                            }
                        )
                    continue

                if not current_code or not current_name:
                    skipped_date_rows.append(
                        {
                            "page": page_number,
                            "y": float(row["y"]),
                            "reason": "missing_account_context",
                            "text": row_text,
                        }
                    )
                    continue

                fecha = parse_date_from_report(fields.get("FECHA"))
                if fecha is None:
                    skipped_date_rows.append(
                        {
                            "page": page_number,
                            "y": float(row["y"]),
                            "reason": "invalid_date_value",
                            "text": row_text,
                        }
                    )
                    continue

                rows.append(
                    {
                        "COD": current_code,
                        "CUENTA": current_name,
                        "EXT": sanitize_text(fields.get("EXT", "")).upper() or "N",
                        "FECHA": fecha,
                        "ORIGEN": sanitize_text(fields.get("ORIGEN", "")).upper(),
                        "ASIENTO": parse_int_like(fields.get("ASIENTO")) or 0,
                        "DOCU": normalize_reference(fields.get("DOCU")),
                        "DETALLE": sanitize_text(fields.get("DETALLE")),
                        "DEBE": round2(parse_decimal_like(fields.get("DEBE"))),
                        "HABER": round2(parse_decimal_like(fields.get("HABER"))),
                        "SALDO": round2(parse_decimal_like(fields.get("SALDO"))),
                    }
                )

    if not rows:
        raise ValueError("No se detectaron movimientos validos en el PDF de Accion 3.")

    if skipped_date_rows:
        preview = " | ".join(
            f"pag {item['page']} ({item['reason']}): {item['text']}" for item in skipped_date_rows[:5]
        )
        raise ValueError(
            f"Se detectaron {len(skipped_date_rows)} filas con fecha que no pudieron mapearse en Accion 3. {preview}"
        )

    totals_validation = validate_account_totals(rows, account_header_totals)
    if totals_validation["mismatches"]:
        preview = " | ".join(
            (
                f"{item['code']} ({item['name']}): "
                f"debe PDF={item['expected']['total_debe']} / Excel={item['actual']['total_debe']}, "
                f"haber PDF={item['expected']['total_haber']} / Excel={item['actual']['total_haber']}, "
                f"saldo PDF={item['expected']['saldo_final']} / Excel={item['actual']['saldo_final']}"
            )
            for item in totals_validation["mismatches"][:4]
        )
        raise ValueError(
            f"Validacion contable fallida en Accion 3 ({len(totals_validation['mismatches'])} cuentas). {preview}"
        )

    return rows, {
        "source_type": "pdf",
        "pages": pages_total,
        "date_rows_detected": date_rows_detected,
        "movement_rows_extracted": len(rows),
        "skipped_date_rows": len(skipped_date_rows),
        "account_totals_checked": totals_validation["checked_accounts"],
        "account_total_mismatches": len(totals_validation["mismatches"]),
    }


def parse_txt_rows(input_txt_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    raw_text = input_txt_path.read_text(encoding="utf-8", errors="replace")
    lines = [
        line.replace("\ufeff", "").rstrip()
        for line in raw_text.splitlines()
        if sanitize_text(line) != ""
    ]

    rows: list[dict[str, Any]] = []
    account_header_totals: dict[str, dict[str, Any]] = {}
    skipped_date_rows: list[dict[str, Any]] = []
    date_rows_detected = 0

    for index, line in enumerate(lines, start=1):
        cols = line.split("\t")
        if len(cols) < 30:
            continue

        date_token = sanitize_text(cols[22]).upper()
        has_date_token = bool(re.fullmatch(r"\d{2}-[A-Z]{3}-\d{2}", date_token))
        if has_date_token:
            date_rows_detected += 1

        code = sanitize_text(cols[6])
        account_name = sanitize_text(cols[7])
        fecha = parse_date_from_report(date_token)
        origen = sanitize_text(cols[23]).upper()
        asiento = parse_int_like(cols[24]) or 0
        ext = sanitize_text(cols[21]).upper() or "N"
        docu = normalize_reference(cols[25])
        detalle = sanitize_text(cols[26])
        debe = round2(parse_decimal_like(cols[27]))
        haber = round2(parse_decimal_like(cols[28]))
        saldo = round2(parse_decimal_like(cols[29]))

        if fecha is None:
            if has_date_token:
                skipped_date_rows.append({"line": index, "reason": "invalid_date_value", "text": sanitize_text(line)})
            continue
        if not re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", code):
            skipped_date_rows.append({"line": index, "reason": "invalid_account_code", "text": sanitize_text(line)})
            continue
        if not account_name:
            skipped_date_rows.append({"line": index, "reason": "missing_account_name", "text": sanitize_text(line)})
            continue
        if not re.fullmatch(r"[A-Z]{2,6}", origen):
            skipped_date_rows.append({"line": index, "reason": "invalid_origin", "text": sanitize_text(line)})
            continue

        account_header_totals[code] = {
            "name": account_name,
            "saldo_inicial": round2(parse_decimal_like(cols[8])),
            "total_debe": round2(parse_decimal_like(cols[9])),
            "total_haber": round2(parse_decimal_like(cols[10])),
            "saldo_final": round2(parse_decimal_like(cols[11])),
        }
        rows.append(
            {
                "COD": code,
                "CUENTA": account_name,
                "EXT": ext,
                "FECHA": fecha,
                "ORIGEN": origen,
                "ASIENTO": asiento,
                "DOCU": docu,
                "DETALLE": detalle,
                "DEBE": debe,
                "HABER": haber,
                "SALDO": saldo,
            }
        )

    if not rows:
        raise ValueError("No se detectaron movimientos validos en el TXT de Accion 3.")
    if skipped_date_rows:
        preview = " | ".join(
            f"linea {item['line']} ({item['reason']})" for item in skipped_date_rows[:5]
        )
        raise ValueError(
            f"Se detectaron {len(skipped_date_rows)} filas con fecha que no pudieron mapearse en TXT Accion 3. {preview}"
        )

    totals_validation = validate_account_totals(rows, account_header_totals)
    if totals_validation["mismatches"]:
        preview = " | ".join(
            (
                f"{item['code']} ({item['name']}): "
                f"debe TXT={item['expected']['total_debe']} / Excel={item['actual']['total_debe']}, "
                f"haber TXT={item['expected']['total_haber']} / Excel={item['actual']['total_haber']}, "
                f"saldo TXT={item['expected']['saldo_final']} / Excel={item['actual']['saldo_final']}"
            )
            for item in totals_validation["mismatches"][:4]
        )
        raise ValueError(
            f"Validacion contable fallida en TXT Accion 3 ({len(totals_validation['mismatches'])} cuentas). {preview}"
        )

    return rows, {
        "source_type": "txt",
        "lines_total": len(lines),
        "date_rows_detected": date_rows_detected,
        "movement_rows_extracted": len(rows),
        "skipped_date_rows": len(skipped_date_rows),
        "account_totals_checked": totals_validation["checked_accounts"],
        "account_total_mismatches": len(totals_validation["mismatches"]),
    }


def validate_rows(rows: list[dict[str, Any]]) -> None:
    problems: list[str] = []
    for index, row in enumerate(rows, start=1):
        if not sanitize_text(row["COD"]):
            problems.append(f"Fila {index}: COD vacio.")
        if not sanitize_text(row["CUENTA"]):
            problems.append(f"Fila {index}: CUENTA vacia.")
        if not isinstance(row["FECHA"], datetime):
            problems.append(f"Fila {index}: FECHA invalida.")
        if not sanitize_text(row["ORIGEN"]):
            problems.append(f"Fila {index}: ORIGEN vacio.")
        if not sanitize_text(row["DETALLE"]):
            problems.append(f"Fila {index}: DETALLE vacio.")

    if problems:
        raise ValueError(f"Validacion de filas fallida ({len(problems)} problemas). {' | '.join(problems[:8])}")


def parse_input_sources(input_paths: list[Path]) -> tuple[list[dict[str, Any]], dict[str, Any], list[dict[str, Any]]]:
    if not input_paths:
        raise ValueError("No se recibieron archivos de entrada para Accion 3.")

    extensions = {path.suffix.lower() for path in input_paths}
    if len(extensions) != 1:
        raise ValueError("No se pueden mezclar TXT y PDF en una sola ejecucion de Accion 3.")

    input_ext = next(iter(extensions))
    if input_ext == ".pdf":
        if len(input_paths) > 1:
            raise ValueError("La consolidacion multiple de Accion 3 solo admite archivos TXT.")

        parsed_rows, parsed_diagnostics = parse_pdf_rows(input_paths[0])
        source_files = [
            {
                "input_source": str(input_paths[0]),
                "file_name": input_paths[0].name,
                "movimientos_extraidos": len(parsed_rows),
                **parsed_diagnostics,
            }
        ]
        return parsed_rows, parsed_diagnostics, source_files

    if input_ext != ".txt":
        raise ValueError("Solo se permiten archivos TXT o PDF para Accion 3.")

    diagnostics = {
        "source_type": "txt_multi" if len(input_paths) > 1 else "txt",
        "files_total": len(input_paths),
        "lines_total": 0,
        "date_rows_detected": 0,
        "movement_rows_extracted": 0,
        "skipped_date_rows": 0,
        "account_totals_checked": 0,
        "account_total_mismatches": 0,
    }
    rows: list[dict[str, Any]] = []
    source_files: list[dict[str, Any]] = []

    for source_index, input_path in enumerate(input_paths):
        parsed_rows, parsed_diagnostics = parse_txt_rows(input_path)
        diagnostics["lines_total"] += parsed_diagnostics["lines_total"]
        diagnostics["date_rows_detected"] += parsed_diagnostics["date_rows_detected"]
        diagnostics["movement_rows_extracted"] += len(parsed_rows)
        diagnostics["skipped_date_rows"] += parsed_diagnostics["skipped_date_rows"]
        diagnostics["account_totals_checked"] += parsed_diagnostics["account_totals_checked"]
        diagnostics["account_total_mismatches"] += parsed_diagnostics["account_total_mismatches"]

        source_files.append(
            {
                "input_source": str(input_path),
                "file_name": input_path.name,
                "movimientos_extraidos": len(parsed_rows),
                **parsed_diagnostics,
            }
        )
        for row_index, row in enumerate(parsed_rows):
            rows.append({**row, "__source_index": source_index, "__row_index": row_index})

    ordered_rows = [{k: v for k, v in row.items() if not k.startswith("__")} for row in sorted(rows, key=compare_sort_key)]
    return ordered_rows, diagnostics, source_files


def read_template_summary_labels(ws: Worksheet) -> list[dict[str, Any]]:
    labels: list[dict[str, Any]] = []
    for row_index in range(2, 121):
        label = sanitize_text(ws.cell(row_index, 13).value)
        if not label:
            if labels:
                break
            continue
        if label.upper() == "ETIQUETAS DE FILA":
            continue
        labels.append({"row": row_index, "label": label})
        if label.upper() == "TOTAL GENERAL":
            break
    return labels


def canonical_summary_label(value: Any) -> str:
    text = sanitize_text(value).upper()
    text = re.sub(r"\s*%\s*", "%", text)
    text = re.sub(r"\s+", " ", text)
    aliases = {
        "1%RETENCION FUENTE": "1%RETENCION",
        "1%RETENCION": "1%RETENCION",
        "2% RETENCION FUENTE": "2% RETENCION FUENTE",
        "2%RETENCION FUENTE": "2%RETENCION FUENTE",
        "100%RETENCION IVA": "100%RETENCION IVA",
        "RETENCION 30%IVA - COMPRAS": "30%RETENCION IVA - COMPRAS",
        "30%RETENCION IVA - COMPRAS": "30%RETENCION IVA - COMPRAS",
        "RETENCION 70%IVA - SERVICIOS": "70%RETENCION IVA -SERVICIOS",
        "70%RETENCION IVA -SERVICIOS": "70%RETENCION IVA -SERVICIOS",
        "70%RETENCION IVA - SERVICIOS": "70%RETENCION IVA -SERVICIOS",
        "RETENCION 10% IVA (BIENES)": "RETENCION 10% IVA (BIENES)",
        "RETENCION 20% IVA (SERVICIOS )": "RETENCION 20% IVA (SERVICIOS)",
        "RETENCION 20% IVA (SERVICIOS)": "RETENCION 20% IVA (SERVICIOS)",
        "TOTAL GENERAL": "TOTAL GENERAL",
    }
    return aliases.get(text, text)


def should_render_row(row: dict[str, Any], *, drop_agcm: bool) -> bool:
    if not drop_agcm:
        return True
    return sanitize_text(row["ORIGEN"]).upper() != "AGCM"


def should_include_row_in_summary(row: dict[str, Any]) -> bool:
    origin = sanitize_text(row["ORIGEN"]).upper()
    return origin not in {"AGCM", "BCPI"}


def build_summary(rows: list[dict[str, Any]], template_labels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    bucket: dict[str, dict[str, int]] = {}
    total_debe_cents = 0
    total_haber_cents = 0

    for row in rows:
        if not should_include_row_in_summary(row):
            continue

        label = canonical_summary_label(row["CUENTA"])
        item = bucket.setdefault(label, {"debe_cents": 0, "haber_cents": 0})
        debe_cents = to_cents(row["DEBE"])
        haber_cents = to_cents(row["HABER"])
        item["debe_cents"] += debe_cents
        item["haber_cents"] += haber_cents
        total_debe_cents += debe_cents
        total_haber_cents += haber_cents

    result: list[dict[str, Any]] = []
    for label_item in template_labels:
        normalized_label = canonical_summary_label(label_item["label"])
        if normalized_label == "TOTAL GENERAL":
            result.append(
                {
                    "row": label_item["row"],
                    "label": label_item["label"],
                    "debe": from_cents(total_debe_cents),
                    "haber": from_cents(total_haber_cents),
                }
            )
            continue

        values = bucket.get(normalized_label, {"debe_cents": 0, "haber_cents": 0})
        result.append(
            {
                "row": label_item["row"],
                "label": label_item["label"],
                "debe": from_cents(values["debe_cents"]),
                "haber": from_cents(values["haber_cents"]),
            }
        )

    return result


def copy_row_styles(ws: Worksheet, source_row: int, target_row: int, start_col: int, end_col: int) -> None:
    source_dimension = ws.row_dimensions[source_row]
    target_dimension = ws.row_dimensions[target_row]
    if source_dimension.height is not None:
        target_dimension.height = source_dimension.height
    for col_index in range(start_col, end_col + 1):
        source_cell = ws.cell(source_row, col_index)
        target_cell = ws.cell(target_row, col_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)


def ensure_template_capacity(ws: Worksheet, required_row_count: int) -> None:
    current_row_count = max(ws.max_row, 1)
    if required_row_count <= current_row_count:
        return
    source_row = max(current_row_count, 2)
    for row_index in range(current_row_count + 1, required_row_count + 1):
        copy_row_styles(ws, source_row, row_index, 1, 16)


def clear_range_values(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row_index in range(start_row, end_row + 1):
        for col_index in range(start_col, end_col + 1):
            ws.cell(row_index, col_index).value = None


def find_header_row(ws: Worksheet) -> int:
    for row_index in range(1, min(ws.max_row, 40) + 1):
        headers = [sanitize_text(ws.cell(row_index, index).value).upper() for index in range(1, len(EXPECTED_HEADERS) + 1)]
        if headers in HEADER_VARIANTS:
            return row_index
    return 1


def read_template_layout_slots(ws: Worksheet, header_row: int) -> list[dict[str, Any]]:
    slots: list[dict[str, Any]] = []
    for row_index in range(1, ws.max_row + 1):
        if row_index == header_row:
            continue
        code = sanitize_text(ws.cell(row_index, 1).value)
        cuenta = sanitize_text(ws.cell(row_index, 2).value)
        fecha = parse_template_date(ws.cell(row_index, 4).value)
        origen = sanitize_text(ws.cell(row_index, 5).value).upper()
        asiento = parse_int_like(ws.cell(row_index, 6).value)
        if not re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", code):
            continue
        if not cuenta or fecha is None or asiento is None or not re.fullmatch(r"[A-Z]{2,6}", origen):
            continue
        slots.append(
            {
                "row": row_index,
                "section": "top" if row_index < header_row else "body",
                "signature": "|".join(
                    [
                        code.upper(),
                        cuenta.upper(),
                        sanitize_text(ws.cell(row_index, 3).value).upper(),
                        fecha.strftime("%Y-%m-%d"),
                        origen,
                        str(int(asiento)),
                        normalize_docu_key(ws.cell(row_index, 7).value),
                        sanitize_text(ws.cell(row_index, 8).value).upper(),
                        f"{round2(parse_decimal_like(ws.cell(row_index, 9).value)):.2f}",
                        f"{round2(parse_decimal_like(ws.cell(row_index, 10).value)):.2f}",
                    ]
                ),
            }
        )
    return slots


def reorder_rows_for_template_layout(rows: list[dict[str, Any]], slots: list[dict[str, Any]], header_row: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    by_signature: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for index, row in enumerate(rows):
        by_signature.setdefault(build_row_signature(row), []).append((index, row))

    used_indexes: set[int] = set()
    top_rows: list[dict[str, Any]] = []
    body_rows: list[dict[str, Any]] = []
    for slot in slots:
        bucket = by_signature.get(slot["signature"])
        if not bucket:
            continue
        matched_index, matched_row = bucket.pop(0)
        used_indexes.add(matched_index)
        if slot["section"] == "top":
            top_rows.append(matched_row)
        else:
            body_rows.append(matched_row)

    remaining_rows = [row for index, row in enumerate(rows) if index not in used_indexes]
    top_remaining = [row for row in remaining_rows if sanitize_text(row["ORIGEN"]).upper() == "BCPI"]
    body_remaining = [row for row in remaining_rows if sanitize_text(row["ORIGEN"]).upper() != "BCPI"]
    top_rows.extend(top_remaining)
    available_top_rows = max(0, header_row - 1)
    overflow_top = top_rows[available_top_rows:]
    top_rows = top_rows[:available_top_rows]
    body_rows = overflow_top + body_rows + body_remaining
    return top_rows, body_rows


def build_workbook_from_template(
    template_path: Path,
    rows: list[dict[str, Any]],
    *,
    drop_agcm_from_render: bool,
) -> tuple[Any, list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(template_path, keep_links=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"La plantilla de Accion 3 no contiene la hoja {SHEET_NAME}.")
    for sheet_name in list(workbook.sheetnames):
        if sheet_name != SHEET_NAME:
            workbook.remove(workbook[sheet_name])

    ws = workbook[SHEET_NAME]
    header_row = find_header_row(ws)
    rendered_rows = [row for row in rows if should_render_row(row, drop_agcm=drop_agcm_from_render)]
    if header_row > 1:
        template_slots = read_template_layout_slots(ws, header_row)
        top_rows, body_rows = reorder_rows_for_template_layout(rendered_rows, template_slots, header_row)
        max_rows = max(ws.max_row, header_row + len(body_rows))
    else:
        top_rows = []
        body_rows = list(rendered_rows)
        max_rows = max(ws.max_row, len(rendered_rows) + 1)
    ensure_template_capacity(ws, max_rows)
    if header_row > 1:
        clear_range_values(ws, 1, header_row - 1, 1, 11)
        clear_range_values(ws, header_row + 1, max_rows, 1, 11)
        clear_range_values(ws, 1, max_rows, 14, 15)
    else:
        clear_range_values(ws, 2, max_rows, 1, 11)
        clear_range_values(ws, 2, max_rows, 14, 15)

    def write_row(row_index: int, row: dict[str, Any]) -> None:
        ws.cell(row_index, 1).value = row["COD"]
        ws.cell(row_index, 2).value = row["CUENTA"]
        ws.cell(row_index, 3).value = row["EXT"]
        ws.cell(row_index, 4).value = row["FECHA"]
        ws.cell(row_index, 5).value = row["ORIGEN"]
        ws.cell(row_index, 6).value = row["ASIENTO"]
        ws.cell(row_index, 7).value = row["DOCU"]
        ws.cell(row_index, 8).value = row["DETALLE"]
        ws.cell(row_index, 9).value = round2(row["DEBE"])
        ws.cell(row_index, 10).value = round2(row["HABER"])
        ws.cell(row_index, 11).value = round2(row["SALDO"])

    for row_index, row in enumerate(top_rows, start=1):
        write_row(row_index, row)
    body_start_row = header_row + 1 if header_row > 1 else 2
    for row_index, row in enumerate(body_rows, start=body_start_row):
        write_row(row_index, row)

    summary_labels = read_template_summary_labels(ws)
    summary = build_summary(rendered_rows, summary_labels)
    for item in summary:
        ws.cell(item["row"], 14).value = round2(item["debe"])
        ws.cell(item["row"], 15).value = round2(item["haber"])

    auto_filter_start = header_row if header_row > 1 else 1
    auto_filter_end = max(auto_filter_start + 1, body_start_row + len(body_rows) - 1)
    ws.auto_filter.ref = f"A{auto_filter_start}:K{auto_filter_end}"
    return workbook, summary, {"header_row": header_row, "body_rows": len(body_rows), "top_rows": len(top_rows)}


def preserve_xml_namespaces(raw_bytes: bytes) -> None:
    seen: set[tuple[str, str]] = set()
    for _, namespace in ET.iterparse(BytesIO(raw_bytes), events=("start-ns",)):
        prefix, uri = namespace
        key = (prefix or "", uri)
        if key in seen:
            continue
        ET.register_namespace(prefix or "", uri)
        seen.add(key)


def update_xml_entry(
    raw_bytes: bytes,
    updater: Any,
    *,
    default_namespace: str | None = None,
    prefix_namespaces: dict[str, str] | None = None,
) -> bytes:
    if default_namespace is None and not prefix_namespaces:
        preserve_xml_namespaces(raw_bytes)
    if default_namespace:
        ET.register_namespace("", default_namespace)
    for prefix, uri in (prefix_namespaces or {}).items():
        ET.register_namespace(prefix, uri)
    root = ET.fromstring(raw_bytes)
    updater(root)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def find_first_local(root: ET.Element, name: str) -> ET.Element | None:
    for element in root.iter():
        if element.tag.rsplit("}", 1)[-1] == name:
            return element
    return None


def remove_external_links_from_package(output_path: Path) -> None:
    with zipfile.ZipFile(output_path, "r") as output_zip:
        output_entries = {name: output_zip.read(name) for name in output_zip.namelist()}

    names_to_remove = [name for name in output_entries if name.startswith("xl/externalLinks/")]
    for name in names_to_remove:
        output_entries.pop(name, None)

    if "xl/workbook.xml" in output_entries:
        def patch_workbook(root: ET.Element) -> None:
            for child in list(root):
                if child.tag.rsplit("}", 1)[-1] == "externalReferences":
                    root.remove(child)

        output_entries["xl/workbook.xml"] = update_xml_entry(
            output_entries["xl/workbook.xml"],
            patch_workbook,
            default_namespace=SPREADSHEETML_NS,
            prefix_namespaces={"r": OFFICEDOC_RELS_NS},
        )

    if "xl/_rels/workbook.xml.rels" in output_entries:
        def patch_workbook_rels(root: ET.Element) -> None:
            for child in list(root):
                rel_type = child.get("Type", "")
                target = child.get("Target", "").replace("\\", "/")
                if rel_type.endswith("/externalLink") or "externalLinks/" in target:
                    root.remove(child)

        output_entries["xl/_rels/workbook.xml.rels"] = update_xml_entry(
            output_entries["xl/_rels/workbook.xml.rels"],
            patch_workbook_rels,
            default_namespace=PACKAGE_RELS_NS,
        )

    if "[Content_Types].xml" in output_entries:
        def patch_content_types(root: ET.Element) -> None:
            for child in list(root):
                local_name = child.tag.rsplit("}", 1)[-1]
                if local_name != "Override":
                    continue
                content_type = child.get("ContentType", "")
                part_name = child.get("PartName", "")
                if content_type == EXTERNAL_LINK_CONTENT_TYPE or part_name.startswith("/xl/externalLinks/"):
                    root.remove(child)

        output_entries["[Content_Types].xml"] = update_xml_entry(
            output_entries["[Content_Types].xml"],
            patch_content_types,
            default_namespace=CONTENT_TYPES_NS,
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_path = Path(temp_file.name)

    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for name, payload in output_entries.items():
                target_zip.writestr(name, payload)
        shutil.move(str(temp_path), str(output_path))
    finally:
        temp_path.unlink(missing_ok=True)


def preserve_visual_artifacts(template_path: Path, output_path: Path, data_row_count: int, *, header_row: int) -> None:
    ref_start = header_row if header_row > 1 else 1
    ref_end = max(ref_start + 1, ref_start + data_row_count)
    ref = f"A{ref_start}:K{ref_end}"
    with zipfile.ZipFile(template_path, "r") as template_zip, zipfile.ZipFile(output_path, "r") as output_zip:
        output_entries = {name: output_zip.read(name) for name in output_zip.namelist()}
        for entry in CRITICAL_TEMPLATE_ENTRIES:
            if entry not in output_entries and entry in template_zip.namelist():
                output_entries[entry] = template_zip.read(entry)

    def patch_pivot_cache(root: ET.Element) -> None:
        worksheet_source = find_first_local(root, "worksheetSource")
        if worksheet_source is not None:
            worksheet_source.set("ref", ref)
            worksheet_source.set("sheet", SHEET_NAME)
        cache_definition = find_first_local(root, "pivotCacheDefinition")
        if cache_definition is not None:
            cache_definition.set("recordCount", str(max(0, data_row_count)))

    if "xl/pivotCache/pivotCacheDefinition1.xml" in output_entries:
        output_entries["xl/pivotCache/pivotCacheDefinition1.xml"] = update_xml_entry(
            output_entries["xl/pivotCache/pivotCacheDefinition1.xml"],
            patch_pivot_cache,
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_path = Path(temp_file.name)

    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for name, payload in output_entries.items():
                target_zip.writestr(name, payload)
        shutil.move(str(temp_path), str(output_path))
    finally:
        temp_path.unlink(missing_ok=True)


def write_workbook_with_retries(workbook: Any, preferred_path: Path, max_attempts: int = 20) -> Path:
    preferred_path.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(max_attempts):
        candidate = (
            preferred_path
            if attempt == 0
            else preferred_path.with_name(
                f"{preferred_path.stem}_nuevo{'' if attempt == 1 else f'_{attempt}'}{preferred_path.suffix}"
            )
        )
        try:
            workbook.save(candidate)
            return candidate
        except PermissionError:
            continue
    raise RuntimeError("No se pudo guardar el Excel de salida. Cierra los archivos abiertos y vuelve a intentar.")


def verify_output(output_path: Path, template_path: Path, rows_count: int) -> None:
    with zipfile.ZipFile(output_path, "r") as output_zip, zipfile.ZipFile(template_path, "r") as template_zip:
        for entry in CRITICAL_TEMPLATE_ENTRIES:
            if entry in template_zip.namelist() and entry not in output_zip.namelist():
                raise RuntimeError(f"Validacion final: falta artefacto visual de plantilla ({entry}).")

    workbook = load_workbook(output_path, keep_links=True)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Validacion final: el archivo debe tener una sola hoja llamada {SHEET_NAME}.")

    ws = workbook[SHEET_NAME]
    header_row = find_header_row(ws)
    headers = [sanitize_text(ws.cell(header_row, index).value).upper() for index in range(1, len(EXPECTED_HEADERS) + 1)]
    if headers not in HEADER_VARIANTS:
        raise RuntimeError("Validacion final: encabezados incorrectos en Accion 3.")

    candidate_rows = list(range(1, header_row)) + list(range(header_row + 1, ws.max_row + 1)) if header_row > 1 else list(range(2, ws.max_row + 1))
    checked = 0
    for row_index in candidate_rows:
        code = sanitize_text(ws.cell(row_index, 1).value)
        if re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", code):
            checked += 1
        if checked >= rows_count:
            break
    if checked < rows_count:
        raise RuntimeError(f"Validacion final: faltan filas renderizadas en Accion 3 ({checked}/{rows_count}).")


def write_audit_report(audit_path: Path, payload: dict[str, Any]) -> None:
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def run(request: ProcessRequest) -> ProcessResult:
    input_paths = [path.resolve() for path in request.input_paths]
    if not input_paths:
        raise ValueError("No se recibieron archivos de entrada para Accion 3.")
    extensions = {path.suffix.lower() for path in input_paths}
    if len(extensions) != 1:
        raise ValueError("No se pueden mezclar TXT y PDF en una sola ejecucion de Accion 3.")

    input_ext = next(iter(extensions))
    if input_ext not in {".txt", ".pdf"}:
        raise ValueError("Solo se permiten archivos TXT o PDF para Accion 3.")
    if request.template_path is None:
        raise ValueError("Accion 3 requiere una plantilla XLSX.")

    started_at = time.perf_counter()
    rows, diagnostics, source_files = parse_input_sources(input_paths)
    parse_ms = int((time.perf_counter() - started_at) * 1000)
    validate_rows(rows)
    validate_ms = int((time.perf_counter() - started_at) * 1000) - parse_ms

    resolved_template = resolve_reference_template(request.template_path.resolve(), rows)
    effective_template_path = Path(resolved_template["path"]).resolve()
    if bool(resolved_template["auto_selected"]):
        overrides = load_template_row_overrides(effective_template_path)
        for row in rows:
            signature = build_row_signature(row)
            if signature in overrides:
                row.update(overrides[signature])

    build_started = time.perf_counter()
    # El formato final historico y el mensual no renderizan filas AGCM en la grilla.
    drop_agcm_from_render = True
    workbook, summary, layout_meta = build_workbook_from_template(
        effective_template_path,
        rows,
        drop_agcm_from_render=drop_agcm_from_render,
    )
    workbook.calculation.fullCalcOnLoad = True
    build_ms = int((time.perf_counter() - build_started) * 1000)
    rendered_rows_count = sum(1 for row in rows if should_render_row(row, drop_agcm=drop_agcm_from_render))

    write_started = time.perf_counter()
    final_output_path = write_workbook_with_retries(workbook, request.output_path.resolve())
    preserve_visual_artifacts(
        effective_template_path,
        final_output_path,
        int(layout_meta["body_rows"]),
        header_row=int(layout_meta["header_row"]),
    )
    verify_output(final_output_path, effective_template_path, rendered_rows_count)
    write_ms = int((time.perf_counter() - write_started) * 1000)
    total_ms = int((time.perf_counter() - started_at) * 1000)

    audit_path = final_output_path.with_name(f"{final_output_path.stem}_auditoria.json")
    write_audit_report(
        audit_path,
        {
            "fecha_proceso": datetime.now().isoformat(),
            "input_source": str(input_paths[0]) if len(input_paths) == 1 else [str(path) for path in input_paths],
            "input_sources": [str(path) for path in input_paths],
            "input_pdf": str(input_paths[0]) if input_ext == ".pdf" and len(input_paths) == 1 else None,
            "input_tipo": diagnostics["source_type"],
            "archivos_origen": source_files,
            "total_archivos_origen": len(input_paths),
            "output_xlsx": str(final_output_path),
            "hoja_salida": SHEET_NAME,
            "template_path_solicitada": str(request.template_path.resolve()),
            "template_path_utilizada": str(effective_template_path),
            "template_auto_seleccionada": bool(resolved_template["auto_selected"]),
            "template_match_rows": int(resolved_template["matched_rows"]),
            "template_match_coverage": round(float(resolved_template["coverage"]), 6),
            "movimientos_extraidos": len(rows),
            "movimientos_renderizados": rendered_rows_count,
            "agcm_ocultado_en_render": drop_agcm_from_render,
            "layout_header_row": int(layout_meta["header_row"]),
            "layout_top_rows": int(layout_meta["top_rows"]),
            "layout_body_rows": int(layout_meta["body_rows"]),
            "filas_fecha_detectadas": diagnostics["date_rows_detected"],
            "filas_fecha_omitidas": diagnostics["skipped_date_rows"],
            "cuentas_validadas_fuente": diagnostics["account_totals_checked"],
            "cuentas_descuadradas_fuente": diagnostics["account_total_mismatches"],
            "cuentas_validadas_pdf": diagnostics["account_totals_checked"],
            "cuentas_descuadradas_pdf": diagnostics["account_total_mismatches"],
            "resumen_filtrado": len(summary),
            "timings_ms": {
                "parse": parse_ms,
                "validate": validate_ms,
                "build": build_ms,
                "write": write_ms,
                "total": total_ms,
            },
            "verificacion_final_ok": True,
        },
    )

    console_lines = []
    if len(input_paths) == 1:
        console_lines.append(f"Archivo leido: {input_paths[0]}")
    else:
        console_lines.append(f"Archivos leidos: {len(input_paths)}")
        console_lines.extend(f"- {path}" for path in input_paths)
    console_lines.append(f"Plantilla solicitada: {request.template_path.resolve()}")
    console_lines.append(f"Plantilla utilizada: {effective_template_path}")
    console_lines.extend(
        [
            f"Movimientos extraidos: {len(rows)}",
            f"Resumen lateral: {len(summary)} filas",
            f"Rendimiento (ms): parse={parse_ms}, validate={validate_ms}, build={build_ms}, write={write_ms}, total={total_ms}",
            f"Excel generado (una sola hoja): {final_output_path}",
            f"Auditoria JSON: {audit_path}",
        ]
    )

    return ProcessResult(
        success=True,
        output_path=final_output_path,
        label="accion3",
        metadata={
            "console": "\n".join(console_lines),
            "output_origin": "default_path",
            "fallback_used": False,
            "runtime": "python-native-txt",
            "audit_path": str(audit_path),
            "rows": len(rows),
            "source_count": len(input_paths),
        },
    )
