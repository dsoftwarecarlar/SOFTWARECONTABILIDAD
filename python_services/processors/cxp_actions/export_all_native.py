from __future__ import annotations

import json
import re
from copy import copy, deepcopy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from ..legacy_node import resolve_generated_artifact
from .accion3_native import write_workbook_with_retries

DEFAULT_OUTPUT_XLSX = "acciones_resumen.xlsx"


def normalize_extension(value: Any) -> str:
    return str(value or "").strip().lstrip(".").lower()


def normalize_timestamp(path: Path) -> float:
    stats = path.stat()
    return float(getattr(stats, "st_birthtime", 0) or stats.st_mtime or 0)


def load_action_definitions(config_path: Path) -> list[dict[str, Any]]:
    if not config_path.is_file():
        raise FileNotFoundError("No existe config/cxp/action_exports.json en el proyecto.")

    parsed = json.loads(config_path.read_text(encoding="utf-8"))
    if not isinstance(parsed, list) or not parsed:
        raise ValueError("config/cxp/action_exports.json no contiene acciones exportables validas.")

    definitions: list[dict[str, Any]] = []
    for item in parsed:
        definitions.append(
            {
                "key": str(item.get("key", "")).strip(),
                "label": str(item.get("label") or item.get("key") or "").strip(),
                "sheet_name": str(item.get("sheet_name") or item.get("key") or "").strip(),
                "bundle_extensions": [
                    normalize_extension(extension)
                    for extension in (item.get("bundle_extensions") or ["xlsx"])
                    if normalize_extension(extension)
                ],
                "file_match": item.get("file_match") if isinstance(item.get("file_match"), dict) else {},
            }
        )

    return [item for item in definitions if item["key"] and item["sheet_name"]]


def matches_action_file(action: dict[str, Any], file_name: str) -> bool:
    rule = action.get("file_match") or {}
    rule_type = str(rule.get("type", "")).strip().lower()
    value = str(rule.get("value", ""))

    if rule_type == "contains":
        return value != "" and value.lower() in file_name.lower()

    if rule_type == "regex":
        if value == "":
            return False
        flags = 0
        if "i" in str(rule.get("flags", "")).lower():
            flags |= re.IGNORECASE
        return re.search(value, file_name, flags) is not None

    return False


def collect_bundle_extensions(actions: list[dict[str, Any]]) -> set[str]:
    extensions = {extension for action in actions for extension in action["bundle_extensions"] if extension}
    return extensions or {"xlsx"}


def list_output_files(output_dir: Path, allowed_extensions: set[str]) -> list[dict[str, Any]]:
    if not output_dir.is_dir():
        return []

    entries: list[dict[str, Any]] = []
    for child in output_dir.iterdir():
        if not child.is_file():
            continue
        if normalize_extension(child.suffix) not in allowed_extensions:
            continue
        entries.append(
            {
                "name": child.name,
                "path": child.resolve(),
                "timestamp": normalize_timestamp(child),
            }
        )

    entries.sort(key=lambda item: (-item["timestamp"], item["name"]), reverse=False)
    return entries


def find_latest_action_files(output_dir: Path, actions: list[dict[str, Any]]) -> list[dict[str, Any]]:
    files = list_output_files(output_dir, collect_bundle_extensions(actions))
    latest: list[dict[str, Any]] = []
    for action in actions:
        latest_file = next((file for file in files if matches_action_file(action, file["name"])), None)
        latest.append({**action, "latest": latest_file})
    return latest


def clone_sheet_layout(source: Worksheet, target: Worksheet) -> None:
    target.sheet_properties = copy(source.sheet_properties)
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)
    target.sheet_format = copy(source.sheet_format)
    target.freeze_panes = source.freeze_panes
    target.auto_filter.ref = source.auto_filter.ref
    target.sheet_state = source.sheet_state

    for key, dimension in source.column_dimensions.items():
        target_dimension = target.column_dimensions[key]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.bestFit = dimension.bestFit
        target_dimension.outlineLevel = dimension.outlineLevel

    for row_index, dimension in source.row_dimensions.items():
        target_dimension = target.row_dimensions[row_index]
        target_dimension.height = dimension.height
        target_dimension.hidden = dimension.hidden
        target_dimension.outlineLevel = dimension.outlineLevel


def copy_sheet_cells(source: Worksheet, target: Worksheet) -> None:
    for row in source.iter_rows():
        for source_cell in row:
            target_cell = target.cell(row=source_cell.row, column=source_cell.column)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)
            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    for merge_range in source.merged_cells.ranges:
        target.merge_cells(str(merge_range))


def quote_sheet_reference(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'!"


def rewrite_formula_sheet_references(formula: str, sheet_name_map: dict[str, str]) -> str:
    updated = formula
    for source_sheet_name, target_sheet_name in sheet_name_map.items():
        replacement = quote_sheet_reference(target_sheet_name)
        quoted_external_pattern = re.compile(rf"'(?:\[[^\]]+\])?{re.escape(source_sheet_name)}'!")
        updated = quoted_external_pattern.sub(replacement, updated)

        if source_sheet_name != target_sheet_name:
            quoted_internal_pattern = re.compile(rf"'{re.escape(source_sheet_name)}'!")
            updated = quoted_internal_pattern.sub(replacement, updated)
    return updated


def rewrite_bundle_formulas(workbook: Workbook, sheet_name_map: dict[str, str]) -> int:
    rewrites = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                updated_formula = rewrite_formula_sheet_references(cell.value, sheet_name_map)
                if updated_formula != cell.value:
                    cell.value = updated_formula
                    rewrites += 1
    return rewrites


def count_external_workbook_formulas(workbook: Workbook) -> int:
    count = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "[" in cell.value:
                    count += 1
    return count


def copy_workbook_sheet(target_workbook: Workbook, source_path: Path, target_sheet_name: str) -> str:
    source_workbook = load_workbook(source_path, data_only=False, keep_links=True)
    source_sheet = source_workbook.worksheets[0] if source_workbook.worksheets else None
    if source_sheet is None:
        raise ValueError(f"El archivo {source_path.name} no contiene hojas para copiar.")

    if target_workbook.sheetnames == ["Sheet"] and target_workbook["Sheet"].max_row == 1 and target_workbook["Sheet"].max_column == 1 and target_workbook["Sheet"]["A1"].value is None:
        target_sheet = target_workbook.active
        target_sheet.title = target_sheet_name
    else:
        target_sheet = target_workbook.create_sheet(title=target_sheet_name)

    clone_sheet_layout(source_sheet, target_sheet)
    copy_sheet_cells(source_sheet, target_sheet)
    return str(source_sheet.title)


def run(request: ProcessRequest) -> ProcessResult:
    root = Path(str(request.options.get("cwd", Path.cwd()))).resolve()
    outputs_dir = root / "storage" / "outputs"
    config_path = root / "config" / "cxp" / "action_exports.json"

    actions = load_action_definitions(config_path)
    latest_files = find_latest_action_files(outputs_dir, actions)
    missing = [item for item in latest_files if item["latest"] is None]
    if missing:
        raise ValueError(f"Faltan archivos generados para: {', '.join(item['label'] for item in missing)}.")

    workbook = Workbook()
    workbook.properties.creator = "Codex"
    workbook.properties.modified = workbook.properties.created
    workbook.calculation.fullCalcOnLoad = True

    sheet_name_map: dict[str, str] = {}
    for item in latest_files:
        source_sheet_name = copy_workbook_sheet(workbook, Path(item["latest"]["path"]), str(item["sheet_name"]))
        sheet_name_map[source_sheet_name] = str(item["sheet_name"])

    rewritten_formulas = rewrite_bundle_formulas(workbook, sheet_name_map)
    external_formula_count = count_external_workbook_formulas(workbook)
    if external_formula_count > 0:
        raise RuntimeError(
            f"El consolidado conserva {external_formula_count} formulas con referencias externas despues de reescritura."
        )

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["Sheet"])

    final_output_path = write_workbook_with_retries(workbook, request.output_path.resolve())
    resolved_path, origin, fallback_used = resolve_generated_artifact(
        output_path=final_output_path,
        lines=[f"Excel consolidado generado: {final_output_path}"],
        prefix="Excel consolidado generado:",
    )

    console_lines = [f"Excel consolidado generado: {resolved_path}"]
    console_lines.extend(f"{item['label']}: {item['latest']['path']}" for item in latest_files)

    return ProcessResult(
        success=True,
        output_path=resolved_path,
        label="consolidado-acciones",
        metadata={
            "console": "\n".join(console_lines),
            "output_origin": origin,
            "fallback_used": fallback_used,
            "runtime": "python-native",
            "sheet_count": len(workbook.sheetnames),
            "rewritten_formulas": rewritten_formulas,
            "external_formula_count": external_formula_count,
        },
    )
