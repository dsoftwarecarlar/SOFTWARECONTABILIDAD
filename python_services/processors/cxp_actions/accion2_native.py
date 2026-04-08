from __future__ import annotations

import json
import math
import re
import shutil
import tempfile
import time
import zipfile
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from .accion3_native import remove_external_links_from_package

SHEET_NAME = "RET PROV"
EXPECTED_COLUMNS = [
    "NUM RT",
    "PROVEEDOR",
    "FECHA",
    "FECHA CONT",
    "TIPO",
    "COD",
    "FACT",
    "%",
    "BASE",
    "RETENCION",
]
HEADER_ALIASES = {
    "NUMRT": "NUM RT",
    "NUMERORT": "NUM RT",
    "NUMRET": "NUM RT",
    "NUMRETENCION": "NUM RT",
    "PROVEEDOR": "PROVEEDOR",
    "RAZONSOCIAL": "PROVEEDOR",
    "NOMBREPROVEEDOR": "PROVEEDOR",
    "FECHA": "FECHA",
    "FECHADOC": "FECHA CONT",
    "FECHADOCUMENTO": "FECHA CONT",
    "FECHADOCTO": "FECHA CONT",
    "FECHADOCU": "FECHA CONT",
    "FECHACONT": "FECHA CONT",
    "FECHACONTABLE": "FECHA CONT",
    "TIPO": "TIPO",
    "COD": "COD",
    "CODIGO": "COD",
    "TRANS": "COD",
    "TRANSACCION": "COD",
    "FACT": "FACT",
    "FACTURA": "FACT",
    "NUMFACT": "FACT",
    "PORCENTAJE": "%",
    "%": "%",
    "%RETEN": "%",
    "BASE": "BASE",
    "BASERET": "BASE",
    "BASEIVA": "BASE",
    "VALORRETEN": "RETENCION",
    "VALORRETENCION": "RETENCION",
    "RETENCION": "RETENCION",
}
CRITICAL_TEMPLATE_ENTRIES = [
    "xl/pivotTables/pivotTable1.xml",
    "xl/worksheets/_rels/sheet1.xml.rels",
    "xl/pivotCache/pivotCacheDefinition1.xml",
]
INVALID_XML_CONTROL_CHARS = re.compile(r"[\u0000-\u0008\u000B\u000C\u000E-\u001F]")
QUOTED_SHEET_REFERENCE = re.compile(r"'((?:[^']|'')+)'!")
UNQUOTED_SHEET_REFERENCE = re.compile(r"(?<![A-Z0-9_])([A-Za-z_][A-Za-z0-9_ .-]*)!")


@dataclass(frozen=True)
class NormalizedRow:
    num_rt: int
    proveedor: str
    fecha: datetime
    fecha_cont: datetime
    tipo: str
    cod: int
    fact: int | str
    percent: float
    base: float
    retencion: float


def sanitize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", INVALID_XML_CONTROL_CHARS.sub(" ", "" if value is None else str(value))).strip()


def round2(value: Any) -> float:
    return round(float(value) + 1e-12, 2)


def parse_int_like(value: Any) -> int | None:
    normalized = re.sub(r"[^\d]", "", sanitize_text(value))
    if not normalized:
        return None
    parsed = int(normalized)
    return parsed if math.isfinite(parsed) else None


def parse_decimal_like(value: Any) -> float:
    normalized = re.sub(r"[^\d,.\-]", "", sanitize_text(value))
    if not normalized:
        return 0.0

    has_comma = "," in normalized
    has_dot = "." in normalized
    if has_comma and has_dot:
        if normalized.rfind(".") > normalized.rfind(","):
            normalized = normalized.replace(",", "")
        else:
            normalized = normalized.replace(".", "").replace(",", ".")
    elif has_comma:
        if normalized.count(",") > 1:
            normalized = normalized.replace(",", "")
        else:
            int_part, _, frac_part = normalized.partition(",")
            normalized = f"{int_part}{frac_part}" if len(frac_part) == 3 else f"{int_part}.{frac_part}"
    elif has_dot and normalized.count(".") > 1:
        last_dot = normalized.rfind(".")
        int_part = normalized[:last_dot].replace(".", "")
        frac_part = normalized[last_dot + 1 :]
        normalized = f"{int_part}{frac_part}" if len(frac_part) == 3 else f"{int_part}.{frac_part}"

    try:
        return float(normalized)
    except ValueError:
        return 0.0


def parse_date_flexible(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    clean = sanitize_text(value)
    if not clean:
        return None

    match = re.fullmatch(r"(\d{2})[/-](\d{2})[/-](\d{4})", clean)
    if match:
        day, month, year = map(int, match.groups())
    else:
        match = re.fullmatch(r"(\d{4})[/-](\d{2})[/-](\d{2})", clean)
        if not match:
            return None
        year, month, day = map(int, match.groups())

    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def normalize_header(value: Any) -> str:
    text = sanitize_text(value)
    text = (
        text.replace("Á", "A")
        .replace("É", "E")
        .replace("Í", "I")
        .replace("Ó", "O")
        .replace("Ú", "U")
        .replace("á", "A")
        .replace("é", "E")
        .replace("í", "I")
        .replace("ó", "O")
        .replace("ú", "U")
        .upper()
    )
    return re.sub(r"[^A-Z0-9%]", "", text)


def split_delimited_line(line: str, delimiter: str) -> list[str]:
    if delimiter == "\t":
        return [part.strip() for part in line.split("\t")]

    result: list[str] = []
    current: list[str] = []
    in_quotes = False
    index = 0
    while index < len(line):
        char = line[index]
        next_char = line[index + 1] if index + 1 < len(line) else ""
        if char == '"':
            if in_quotes and next_char == '"':
                current.append('"')
                index += 2
                continue
            in_quotes = not in_quotes
            index += 1
            continue
        if char == delimiter and not in_quotes:
            result.append("".join(current).strip())
            current = []
            index += 1
            continue
        current.append(char)
        index += 1

    result.append("".join(current).strip())
    return result


def map_headers(tokens: list[str]) -> list[str | None]:
    return [HEADER_ALIASES.get(normalize_header(token)) for token in tokens]


def detect_header_config(lines: list[str]) -> dict[str, Any] | None:
    delimiters = ["\t", ";", "|", ","]
    best: dict[str, Any] | None = None
    for index, raw_line in enumerate(lines[:60]):
        if not sanitize_text(raw_line):
            continue
        for delimiter in delimiters:
            tokens = split_delimited_line(raw_line, delimiter)
            if len(tokens) < 5:
                continue
            mapped = map_headers(tokens)
            recognized = sum(1 for item in mapped if item)
            if recognized < 5:
                continue
            score = recognized * 100 + len(tokens)
            if best is None or score > int(best["score"]):
                best = {
                    "score": score,
                    "header_index": index,
                    "delimiter": delimiter,
                    "mapped_headers": mapped,
                }
    return best


def parse_rows_from_header(lines: list[str], config: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    header_index = int(config["header_index"])
    delimiter = str(config["delimiter"])
    mapped_headers = list(config["mapped_headers"])
    for raw_line in lines[header_index + 1 :]:
        if not sanitize_text(raw_line):
            continue
        tokens = split_delimited_line(raw_line, delimiter)
        if all(not sanitize_text(token) for token in tokens):
            continue
        row: dict[str, str] = {}
        for column_index, header in enumerate(mapped_headers):
            if not header:
                continue
            row[header] = tokens[column_index] if column_index < len(tokens) else ""
        if any(sanitize_text(value) for value in row.values()):
            rows.append(row)
    return rows


def parse_rows_by_pattern(lines: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    hard_pattern = re.compile(
        r"^(\d+)\s+(.+?)\s+(\d{2}[/-]\d{2}[/-]\d{4})\s+(\d{2}[/-]\d{2}[/-]\d{4})\s+"
        r"(IVA|RENTA)\s+([A-Z0-9]+)\s+([A-Z0-9\-/]+)\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$",
        re.IGNORECASE,
    )
    for line in lines:
        clean_line = line.strip()
        if not clean_line:
            continue
        match = hard_pattern.fullmatch(clean_line)
        if match:
            rows.append(
                {
                    "NUM RT": match.group(1),
                    "PROVEEDOR": match.group(2),
                    "FECHA": match.group(3),
                    "FECHA CONT": match.group(4),
                    "TIPO": match.group(5),
                    "COD": match.group(6),
                    "FACT": match.group(7),
                    "%": match.group(8),
                    "BASE": match.group(9),
                    "RETENCION": match.group(10),
                }
            )
            continue

        pieces = [part.strip() for part in re.split(r"\s{2,}", clean_line) if part.strip()]
        if len(pieces) >= 10:
            rows.append(
                {
                    "NUM RT": pieces[0],
                    "PROVEEDOR": pieces[1],
                    "FECHA": pieces[2],
                    "FECHA CONT": pieces[3],
                    "TIPO": pieces[4],
                    "COD": pieces[5],
                    "FACT": pieces[6],
                    "%": pieces[7],
                    "BASE": pieces[8],
                    "RETENCION": pieces[9],
                }
            )
    return rows


def parse_rows_from_embedded_ret_report(lines: list[str]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    matched = 0
    for line in lines:
        if "\t" not in line:
            continue
        tokens = [sanitize_text(token) for token in line.split("\t")]
        if len(tokens) < 20:
            continue
        if normalize_header(tokens[1]) != "RETENCION":
            continue
        if normalize_header(tokens[2]) != "FECHA":
            continue
        if normalize_header(tokens[3]) != "FECHADOC":
            continue

        num_rt = tokens[11]
        fecha = tokens[12]
        fecha_cont = tokens[13]
        cod = tokens[14]
        fact = tokens[15]
        percent = tokens[16]
        base = tokens[17]
        ret = tokens[18]
        proveedor = tokens[19]

        if not re.fullmatch(r"\d+", num_rt):
            continue
        if not re.fullmatch(r"\d{2}[/-]\d{2}[/-]\d{4}", fecha):
            continue
        if not re.fullmatch(r"\d{2}[/-]\d{2}[/-]\d{4}", fecha_cont):
            continue
        if not sanitize_text(proveedor):
            continue

        matched += 1
        rows.append(
            {
                "NUM RT": num_rt,
                "PROVEEDOR": proveedor,
                "FECHA": fecha,
                "FECHA CONT": fecha_cont,
                "COD": cod,
                "FACT": fact,
                "%": percent,
                "BASE": base,
                "RETENCION": ret,
                "TIPO": "",
            }
        )

    min_expected = max(20, int(len(lines) * 0.6))
    return rows if matched >= min_expected else []


def load_best_text(buffer: bytes) -> str:
    utf8 = buffer.decode("utf-8", errors="replace").lstrip("\ufeff")
    latin1 = buffer.decode("latin1", errors="replace").lstrip("\ufeff")
    return utf8 if utf8.count("\ufffd") <= latin1.count("\ufffd") else latin1


def extract_rows_from_txt(input_path: Path) -> list[dict[str, str]]:
    text = load_best_text(input_path.read_bytes())
    lines = [line.replace("\u00A0", " ") for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n")]

    raw_rows = parse_rows_from_embedded_ret_report(lines)
    if not raw_rows:
        header_config = detect_header_config(lines)
        if header_config:
            raw_rows = parse_rows_from_header(lines, header_config)
    if not raw_rows:
        raw_rows = parse_rows_by_pattern(lines)
    if not raw_rows:
        raise ValueError("No se detectaron filas validas en el TXT. Verifica encabezados o delimitadores.")
    return raw_rows


def normalize_tipo(value: Any) -> str:
    clean = sanitize_text(value).upper()
    if clean == "IVA":
        return "IVA"
    if clean == "RENTA":
        return "RENTA"
    return clean


def format_date_key(value: Any) -> str:
    parsed = parse_date_flexible(value)
    return parsed.strftime("%d/%m/%Y") if parsed else ""


def normalize_fact_key(value: Any) -> str:
    clean = sanitize_text(value)
    return str(int(clean)) if re.fullmatch(r"\d+", clean) else clean.upper()


def normalize_provider_key(value: Any) -> str:
    return sanitize_text(value).upper()


def build_tipo_match_key(fields: dict[str, Any]) -> str:
    return "|".join(
        [
            "" if fields.get("num_rt") is None else str(fields["num_rt"]),
            format_date_key(fields.get("fecha")),
            format_date_key(fields.get("fecha_cont") or fields.get("fecha")),
            "" if fields.get("cod") is None else str(fields["cod"]),
            normalize_fact_key(fields.get("fact")),
            f"{round2(fields.get('percent', 0)):.4f}",
            f"{round2(fields.get('base', 0)):.2f}",
            f"{round2(fields.get('retencion', 0)):.2f}",
            normalize_provider_key(fields.get("proveedor")),
        ]
    )


def discover_template_candidates(primary_template_path: Path) -> list[Path]:
    candidates: list[Path] = []
    seen: set[Path] = set()
    search_dirs = [
        primary_template_path.parent,
        primary_template_path.parent.parent / "PLANTILLAYARCHIVOS",
    ]

    def add_candidate(path: Path) -> None:
        resolved = path.resolve()
        if resolved in seen or not resolved.is_file():
            return
        if resolved.suffix.lower() not in {".xlsx", ".xlsm"}:
            return
        seen.add(resolved)
        candidates.append(resolved)

    add_candidate(primary_template_path)
    for directory in search_dirs:
        if not directory.is_dir():
            continue
        for pattern in ("*.xlsx", "*.xlsm"):
            for candidate in sorted(directory.glob(pattern)):
                add_candidate(candidate)

    return candidates


def load_template_match_index(template_path: Path) -> dict[str, Any]:
    if not template_path.is_file():
        return {"keys": set(), "row_count": 0}

    workbook = load_workbook(template_path, data_only=False, keep_links=False)
    if SHEET_NAME not in workbook.sheetnames:
        return {"keys": set(), "row_count": 0}

    ws = workbook[SHEET_NAME]
    keys: set[str] = set()
    row_count = 0
    for row_index in range(2, ws.max_row + 1):
        tipo = normalize_tipo(ws.cell(row_index, 5).value)
        if tipo not in {"IVA", "RENTA"}:
            continue

        fields = {
            "num_rt": parse_int_like(ws.cell(row_index, 1).value),
            "proveedor": ws.cell(row_index, 2).value,
            "fecha": ws.cell(row_index, 3).value,
            "fecha_cont": ws.cell(row_index, 4).value,
            "cod": parse_int_like(ws.cell(row_index, 6).value),
            "fact": ws.cell(row_index, 7).value,
            "percent": parse_decimal_like(ws.cell(row_index, 8).value),
            "base": parse_decimal_like(ws.cell(row_index, 9).value),
            "retencion": parse_decimal_like(ws.cell(row_index, 10).value),
        }
        if fields["num_rt"] is None or fields["cod"] is None:
            continue

        keys.add(build_tipo_match_key(fields))
        row_count += 1

    return {"keys": keys, "row_count": row_count}


def resolve_reference_template(primary_template_path: Path, rows: list[NormalizedRow]) -> dict[str, Any]:
    row_count = len(rows)
    primary_result: dict[str, Any] | None = None
    best_result: dict[str, Any] | None = None
    best_score: tuple[int, int, int] | None = None

    input_keys = {
        build_tipo_match_key(
            {
                "num_rt": row.num_rt,
                "proveedor": row.proveedor,
                "fecha": row.fecha,
                "fecha_cont": row.fecha_cont,
                "cod": row.cod,
                "fact": row.fact,
                "percent": row.percent,
                "base": row.base,
                "retencion": row.retencion,
            }
        )
        for row in rows
    }

    for candidate_path in discover_template_candidates(primary_template_path):
        match_index = load_template_match_index(candidate_path)
        matched_rows = sum(1 for key in input_keys if key in match_index["keys"])
        coverage = 0.0 if row_count == 0 else matched_rows / row_count
        candidate_result = {
            "path": candidate_path,
            "matched_rows": matched_rows,
            "coverage": coverage,
            "row_delta": abs(int(match_index["row_count"]) - row_count),
            "is_primary": candidate_path.resolve() == primary_template_path.resolve(),
        }
        if candidate_result["is_primary"]:
            primary_result = candidate_result

        score = (
            int(candidate_result["matched_rows"]),
            -int(candidate_result["row_delta"]),
            1 if candidate_result["is_primary"] else 0,
        )
        if best_result is None or best_score is None or score > best_score:
            best_result = candidate_result
            best_score = score

    selected = primary_result or best_result
    if selected is None:
        return {
            "path": primary_template_path.resolve(),
            "matched_rows": 0,
            "coverage": 0.0,
            "row_delta": row_count,
            "is_primary": True,
            "auto_selected": False,
        }

    if primary_result is None:
        return {**selected, "auto_selected": not selected["is_primary"]}

    should_use_best = (
        best_result is not None
        and not best_result["is_primary"]
        and best_result["matched_rows"] >= max(25, int(row_count * 0.5))
        and best_result["coverage"] >= 0.5
        and best_result["matched_rows"] >= primary_result["matched_rows"] + 20
    )
    return {**(best_result if should_use_best else primary_result), "auto_selected": should_use_best}


def load_template_tipo_hints(template_path: Path) -> dict[str, Any]:
    if not template_path.is_file():
        return {"exact": {}, "by_code_percent": {}}

    workbook = load_workbook(template_path, data_only=False, keep_links=False)
    if SHEET_NAME not in workbook.sheetnames:
        return {"exact": {}, "by_code_percent": {}}

    ws = workbook[SHEET_NAME]
    exact: dict[str, str] = {}
    by_code_percent: dict[str, dict[str, int]] = {}
    for row_index in range(2, ws.max_row + 1):
        tipo = normalize_tipo(ws.cell(row_index, 5).value)
        if tipo not in {"IVA", "RENTA"}:
            continue
        fields = {
            "num_rt": parse_int_like(ws.cell(row_index, 1).value),
            "proveedor": ws.cell(row_index, 2).value,
            "fecha": ws.cell(row_index, 3).value,
            "fecha_cont": ws.cell(row_index, 4).value,
            "cod": parse_int_like(ws.cell(row_index, 6).value),
            "fact": ws.cell(row_index, 7).value,
            "percent": parse_decimal_like(ws.cell(row_index, 8).value),
            "base": parse_decimal_like(ws.cell(row_index, 9).value),
            "retencion": parse_decimal_like(ws.cell(row_index, 10).value),
        }
        if fields["num_rt"] is None or fields["cod"] is None:
            continue

        exact[build_tipo_match_key(fields)] = tipo
        cp_key = f"{fields['cod']}|{round2(fields['percent']):.4f}"
        bucket = by_code_percent.setdefault(cp_key, {"IVA": 0, "RENTA": 0})
        bucket[tipo] += 1

    return {"exact": exact, "by_code_percent": by_code_percent}


def infer_tipo_from_hints(fields: dict[str, Any], hints: dict[str, Any]) -> str:
    exact_key = build_tipo_match_key(fields)
    if exact_key in hints["exact"]:
        return hints["exact"][exact_key]

    rate = round2(fields["percent"])
    cp_key = f"{fields['cod']}|{rate:.4f}"
    cp_hint = hints["by_code_percent"].get(cp_key)
    if cp_hint:
        if cp_hint["IVA"] > cp_hint["RENTA"]:
            return "IVA"
        if cp_hint["RENTA"] > cp_hint["IVA"]:
            return "RENTA"

    if rate == 0:
        # Los comprobantes 0% deben quedar como RENTA; en los libros manuales
        # no forman parte del bloque IVA aunque sigan visibles en el detalle.
        return "RENTA"
    if rate in {20, 30, 70, 100}:
        return "IVA"
    if rate in {1, 1.75, 2, 2.75, 3, 5, 8}:
        return "RENTA"

    return "IVA" if rate >= 20 else "RENTA"


def normalize_parsed_rows(raw_rows: list[dict[str, str]], hints: dict[str, Any]) -> list[NormalizedRow]:
    rows: list[NormalizedRow] = []
    problems: list[str] = []

    for index, source in enumerate(raw_rows, start=1):
        num_rt = parse_int_like(source.get("NUM RT"))
        proveedor = sanitize_text(source.get("PROVEEDOR"))
        fecha = parse_date_flexible(source.get("FECHA"))
        fecha_cont = parse_date_flexible(source.get("FECHA CONT")) or fecha
        cod = parse_int_like(source.get("COD"))
        fact_raw = sanitize_text(source.get("FACT"))
        fact: int | str = int(fact_raw) if re.fullmatch(r"\d+", fact_raw) else fact_raw
        percent = parse_decimal_like(source.get("%"))
        base = parse_decimal_like(source.get("BASE"))
        retencion = parse_decimal_like(source.get("RETENCION"))
        tipo = normalize_tipo(source.get("TIPO"))
        if tipo not in {"IVA", "RENTA"}:
            tipo = infer_tipo_from_hints(
                {
                    "num_rt": num_rt,
                    "proveedor": proveedor,
                    "fecha": fecha,
                    "fecha_cont": fecha_cont,
                    "cod": cod,
                    "fact": fact,
                    "percent": percent,
                    "base": base,
                    "retencion": retencion,
                },
                hints,
            )

        if num_rt is None:
            problems.append(f"Fila {index}: NUM RT invalido.")
        if not proveedor:
            problems.append(f"Fila {index}: PROVEEDOR vacio.")
        if fecha is None:
            problems.append(f"Fila {index}: FECHA invalida.")
        if fecha_cont is None:
            problems.append(f"Fila {index}: FECHA CONT invalida.")
        if tipo not in {"IVA", "RENTA"}:
            problems.append(f"Fila {index}: TIPO invalido ({tipo or 'vacio'}).")
        if cod is None:
            problems.append(f"Fila {index}: COD invalido.")
        if not fact_raw:
            problems.append(f"Fila {index}: FACT vacio.")

        if all(value is not None for value in (num_rt, fecha, fecha_cont, cod)) and fact_raw:
            rows.append(
                NormalizedRow(
                    num_rt=num_rt,
                    proveedor=proveedor,
                    fecha=fecha,
                    fecha_cont=fecha_cont,
                    tipo=tipo,
                    cod=cod,
                    fact=fact,
                    percent=percent,
                    base=base,
                    retencion=retencion,
                )
            )

    if problems:
        preview = " | ".join(problems[:8])
        raise ValueError(f"Validacion de TXT fallida ({len(problems)} problemas). {preview}")

    return rows


def format_percent_label(value: float) -> str:
    if float(value).is_integer():
        return str(int(value))
    text = re.sub(r"(\.\d*?[1-9])0+$", r"\1", f"{value}")
    return re.sub(r"\.0+$", "", text)


def should_include_row_in_summary(row: NormalizedRow) -> bool:
    return row.num_rt != 999999999


def build_summary(rows: list[NormalizedRow]) -> list[dict[str, Any]]:
    type_order = ["IVA", "RENTA"]
    type_map: dict[str, dict[str, Any]] = {}
    total_base_cents = 0
    total_ret_cents = 0

    for row in rows:
        if not should_include_row_in_summary(row):
            continue

        bucket = type_map.setdefault(
            row.tipo,
            {"total_base_cents": 0, "total_ret_cents": 0, "percent_map": {}},
        )
        base_cents = int(round((row.base + 1e-12) * 100))
        ret_cents = int(round((row.retencion + 1e-12) * 100))
        bucket["total_base_cents"] += base_cents
        bucket["total_ret_cents"] += ret_cents
        total_base_cents += base_cents
        total_ret_cents += ret_cents

        percent_bucket = bucket["percent_map"].setdefault(round2(row.percent), {"base_cents": 0, "ret_cents": 0})
        percent_bucket["base_cents"] += base_cents
        percent_bucket["ret_cents"] += ret_cents

    entries: list[dict[str, Any]] = []
    seen: set[str] = set()
    for tipo in type_order:
        if tipo not in type_map:
            continue
        seen.add(tipo)
        bucket = type_map[tipo]
        entries.append(
            {
                "kind": "type",
                "label": tipo,
                "tipo": tipo,
                "base": bucket["total_base_cents"] / 100,
                "ret": bucket["total_ret_cents"] / 100,
                "calc": None,
                "diff": None,
            }
        )
        for percent in sorted(bucket["percent_map"]):
            item = bucket["percent_map"][percent]
            entries.append(
                {
                    "kind": "detail",
                    "label": format_percent_label(percent),
                    "tipo": tipo,
                    "base": item["base_cents"] / 100,
                    "ret": item["ret_cents"] / 100,
                    "calc": item["ret_cents"] / 100,
                    "diff": 0,
                }
            )

    for tipo in sorted(item for item in type_map if item not in seen):
        bucket = type_map[tipo]
        entries.append(
            {
                "kind": "type",
                "label": tipo,
                "tipo": tipo,
                "base": bucket["total_base_cents"] / 100,
                "ret": bucket["total_ret_cents"] / 100,
                "calc": None,
                "diff": None,
            }
        )
        for percent in sorted(bucket["percent_map"]):
            item = bucket["percent_map"][percent]
            entries.append(
                {
                    "kind": "detail",
                    "label": format_percent_label(percent),
                    "tipo": tipo,
                    "base": item["base_cents"] / 100,
                    "ret": item["ret_cents"] / 100,
                    "calc": item["ret_cents"] / 100,
                    "diff": 0,
                }
            )

    entries.append(
        {
            "kind": "total",
            "label": "Total general",
            "tipo": "",
            "base": total_base_cents / 100,
            "ret": total_ret_cents / 100,
            "calc": None,
            "diff": None,
        }
    )
    return entries


def normalize_summary_label(label: Any) -> str:
    return sanitize_text(label).upper()


def extract_formula_sheet_references(formula: str) -> list[str]:
    references: list[str] = []
    for match in QUOTED_SHEET_REFERENCE.finditer(formula):
        references.append(match.group(1).replace("''", "'"))

    unquoted_formula = QUOTED_SHEET_REFERENCE.sub(" ", formula)
    for match in UNQUOTED_SHEET_REFERENCE.finditer(unquoted_formula):
        candidate = sanitize_text(match.group(1))
        if candidate:
            references.append(candidate)

    return references


def formula_has_orphan_reference(formula: str, sheet_names: set[str]) -> bool:
    if not isinstance(formula, str) or not formula.startswith("="):
        return False

    for reference in extract_formula_sheet_references(formula):
        if "[" in reference or "]" in reference:
            return True
        if reference not in sheet_names:
            return True

    return False


def clear_orphan_formula_cells(workbook: Any) -> int:
    sheet_names = set(workbook.sheetnames)
    cleared = 0

    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and formula_has_orphan_reference(cell.value, sheet_names):
                    cell.value = None
                    cleared += 1

    return cleared


def read_template_summary_layout(ws: Worksheet) -> list[dict[str, Any]]:
    layout: list[dict[str, Any]] = []
    current_tipo = ""
    for row_index in range(2, 201):
        label = sanitize_text(ws.cell(row_index, 12).value)
        if not label:
            if layout:
                break
            continue
        normalized = normalize_summary_label(label)
        kind = "detail"
        if normalized == "TOTAL GENERAL":
            kind = "total"
        elif not re.fullmatch(r"[0-9.]+", normalized):
            kind = "type"
            current_tipo = normalized
        layout.append(
            {
                "row": row_index,
                "label": label,
                "kind": kind,
                "tipo": current_tipo if kind == "detail" else normalized if kind == "type" else "",
            }
        )
    return layout


def build_summary_lookup_key(kind: str, label: Any, tipo: Any = "") -> str:
    normalized_kind = sanitize_text(kind).lower()
    normalized_label = normalize_summary_label(label)
    normalized_tipo = normalize_summary_label(tipo)
    if normalized_kind == "detail":
        return f"{normalized_kind}|{normalized_tipo}|{normalized_label}"
    return f"{normalized_kind}|{normalized_label}"


def find_last_fully_styled_row(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> int:
    last = start_row
    for row_index in range(start_row, end_row + 1):
        if all(ws.cell(row_index, col_index).has_style for col_index in range(start_col, end_col + 1)):
            last = row_index
    return last


def clear_range_values(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row_index in range(start_row, end_row + 1):
        for col_index in range(start_col, end_col + 1):
            ws.cell(row_index, col_index).value = None


def copy_row_styles(ws: Worksheet, source_row: int, target_row: int, start_col: int, end_col: int) -> None:
    source_dimension = ws.row_dimensions[source_row]
    target_dimension = ws.row_dimensions[target_row]
    if source_dimension.height is not None:
        target_dimension.height = source_dimension.height
    for col_index in range(start_col, end_col + 1):
        source_cell = ws.cell(source_row, col_index)
        target_cell = ws.cell(target_row, col_index)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)


def write_workbook_with_retries(workbook: Any, preferred_path: Path, max_attempts: int = 20) -> Path:
    preferred_path.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(max_attempts):
        candidate = (
            preferred_path
            if attempt == 0
            else preferred_path.with_name(
                f"{preferred_path.stem}_nuevo{'' if attempt == 1 else f'_{attempt}'}{preferred_path.suffix}"
            )
        )
        try:
            workbook.save(candidate)
            return candidate
        except PermissionError:
            continue
    raise RuntimeError("No se pudo guardar el Excel de salida. Cierra los archivos abiertos y vuelve a intentar.")


def verify_headers(ws: Worksheet) -> None:
    headers = []
    for index in range(1, len(EXPECTED_COLUMNS) + 1):
        raw_value = ws.cell(1, index).value
        normalized = HEADER_ALIASES.get(normalize_header(raw_value))
        headers.append(normalized or sanitize_text(raw_value).upper())
    if headers != EXPECTED_COLUMNS:
        raise RuntimeError("Validacion final: encabezados de Accion 2 alterados.")


def update_xml_entry(raw_bytes: bytes, updater: Any) -> bytes:
    root = ET.fromstring(raw_bytes)
    updater(root)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def find_first_local(root: ET.Element, name: str) -> ET.Element | None:
    for element in root.iter():
        if element.tag.rsplit("}", 1)[-1] == name:
            return element
    return None


def preserve_visual_artifacts(template_path: Path, output_path: Path, data_row_count: int) -> None:
    ref = f"A1:J{max(2, data_row_count + 1)}"
    with zipfile.ZipFile(template_path, "r") as template_zip, zipfile.ZipFile(output_path, "r") as output_zip:
        output_entries = {name: output_zip.read(name) for name in output_zip.namelist()}
        for entry in CRITICAL_TEMPLATE_ENTRIES:
            if entry not in output_entries and entry in template_zip.namelist():
                output_entries[entry] = template_zip.read(entry)

    def patch_pivot_cache(root: ET.Element) -> None:
        worksheet_source = find_first_local(root, "worksheetSource")
        if worksheet_source is not None:
            worksheet_source.set("ref", ref)
            worksheet_source.set("sheet", SHEET_NAME)
        cache_definition = find_first_local(root, "pivotCacheDefinition")
        if cache_definition is not None:
            cache_definition.set("recordCount", str(max(0, data_row_count)))

    def patch_pivot_table(root: ET.Element) -> None:
        _ = find_first_local(root, "pivotTableDefinition")

    if "xl/pivotCache/pivotCacheDefinition1.xml" in output_entries:
        output_entries["xl/pivotCache/pivotCacheDefinition1.xml"] = update_xml_entry(
            output_entries["xl/pivotCache/pivotCacheDefinition1.xml"],
            patch_pivot_cache,
        )
    if "xl/pivotTables/pivotTable1.xml" in output_entries:
        output_entries["xl/pivotTables/pivotTable1.xml"] = update_xml_entry(
            output_entries["xl/pivotTables/pivotTable1.xml"],
            patch_pivot_table,
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_path = Path(temp_file.name)

    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for name, payload in output_entries.items():
                target_zip.writestr(name, payload)
        shutil.move(str(temp_path), str(output_path))
    finally:
        temp_path.unlink(missing_ok=True)


def write_audit_report(audit_path: Path, payload: dict[str, Any]) -> None:
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    audit_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def build_workbook_from_template(template_path: Path, rows: list[NormalizedRow]) -> tuple[Any, list[dict[str, Any]], int]:
    if not template_path.is_file():
        raise FileNotFoundError(f"No se encontro plantilla de Accion 2: {template_path}")

    workbook = load_workbook(template_path, keep_links=False)
    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"La plantilla de Accion 2 no contiene la hoja {SHEET_NAME}.")

    ws = workbook[SHEET_NAME]
    template_row_count = ws.max_row
    max_rows = max(template_row_count, len(rows) + 40, 550)
    last_styled_data_row = find_last_fully_styled_row(ws, 2, template_row_count, 1, 10)

    clear_range_values(ws, 2, max_rows, 1, 11)

    for row_index, row in enumerate(rows, start=2):
        if row_index > last_styled_data_row:
            copy_row_styles(ws, last_styled_data_row, row_index, 1, 11)

        ws.cell(row_index, 1).value = row.num_rt
        ws.cell(row_index, 2).value = row.proveedor
        ws.cell(row_index, 3).value = row.fecha
        ws.cell(row_index, 4).value = row.fecha_cont
        ws.cell(row_index, 5).value = row.tipo
        ws.cell(row_index, 6).value = row.cod
        ws.cell(row_index, 7).value = row.fact
        ws.cell(row_index, 8).value = round2(row.percent)
        ws.cell(row_index, 9).value = round2(row.base)
        ws.cell(row_index, 10).value = round2(row.retencion)
        ws.cell(row_index, 11).value = None

    summary_layout = read_template_summary_layout(ws)
    computed_summary = build_summary(rows)
    summary_lookup = {
        build_summary_lookup_key(item["kind"], item["label"], item.get("tipo", "")): item
        for item in computed_summary
    }
    summary: list[dict[str, Any]] = []
    for slot in summary_layout:
        match = summary_lookup.get(build_summary_lookup_key(slot["kind"], slot["label"], slot.get("tipo", "")))
        base = match["base"] if match else 0
        ret = match["ret"] if match else 0
        ws.cell(slot["row"], 13).value = base
        ws.cell(slot["row"], 14).value = ret
        summary.append(
            {
                "kind": slot["kind"],
                "label": slot["label"],
                "tipo": slot.get("tipo", ""),
                "base": base,
                "ret": ret,
                "calc": match.get("calc") if match else None,
                "diff": match.get("diff") if match else None,
            }
        )

    formula_cells_sanitized = clear_orphan_formula_cells(workbook)
    ws.auto_filter.ref = f"A1:J{max(2, len(rows) + 1)}"
    verify_headers(ws)
    return workbook, summary, formula_cells_sanitized


def run(request: ProcessRequest) -> ProcessResult:
    if len(request.input_paths) != 1:
        raise ValueError("Accion 2 requiere exactamente un archivo TXT de entrada.")
    if request.template_path is None:
        raise ValueError("Accion 2 requiere una plantilla XLSX.")

    input_path = request.input_paths[0].resolve()
    template_path = request.template_path.resolve()
    output_path = request.output_path.resolve()
    started_at = time.perf_counter()

    hints = load_template_tipo_hints(template_path)
    raw_rows = extract_rows_from_txt(input_path)
    rows = normalize_parsed_rows(raw_rows, hints)
    resolved_template = resolve_reference_template(template_path, rows)
    effective_template_path = Path(resolved_template["path"]).resolve()
    if effective_template_path != template_path:
        hints = load_template_tipo_hints(effective_template_path)
        rows = normalize_parsed_rows(raw_rows, hints)
    parse_ms = int((time.perf_counter() - started_at) * 1000)

    build_started = time.perf_counter()
    workbook, summary, formula_cells_sanitized = build_workbook_from_template(effective_template_path, rows)
    workbook.calculation.fullCalcOnLoad = True
    build_ms = int((time.perf_counter() - build_started) * 1000)

    write_started = time.perf_counter()
    final_output_path = write_workbook_with_retries(workbook, output_path)
    preserve_visual_artifacts(effective_template_path, final_output_path, len(rows))
    remove_external_links_from_package(final_output_path)
    write_ms = int((time.perf_counter() - write_started) * 1000)

    audit_path = final_output_path.with_name(f"{final_output_path.stem}_auditoria.json")
    total_ms = int((time.perf_counter() - started_at) * 1000)
    write_audit_report(
        audit_path,
        {
            "fecha_proceso": datetime.now().isoformat(),
            "input_txt": str(input_path),
            "output_xlsx": str(final_output_path),
            "hoja_salida": SHEET_NAME,
            "template_path_solicitada": str(template_path),
            "template_path_utilizada": str(effective_template_path),
            "template_auto_seleccionada": bool(resolved_template["auto_selected"]),
            "template_match_rows": int(resolved_template["matched_rows"]),
            "template_match_coverage": round(float(resolved_template["coverage"]), 6),
            "filas_txt": len(rows),
            "columnas_esperadas": EXPECTED_COLUMNS,
            "resumen_generado_filas": len(summary),
            "formula_cells_sanitized": formula_cells_sanitized,
            "timings_ms": {
                "parse": parse_ms,
                "build": build_ms,
                "write": write_ms,
                "total": total_ms,
            },
            "verificacion_final_ok": True,
        },
    )

    console_lines = [
        f"TXT leido: {input_path}",
        f"Plantilla solicitada: {template_path}",
        f"Plantilla utilizada: {effective_template_path}",
        f"Filas parseadas: {len(rows)}",
        f"Resumen lateral: {len(summary)} filas",
        f"Formulas huerfanas saneadas: {formula_cells_sanitized}",
        f"Rendimiento (ms): parse={parse_ms}, build={build_ms}, write={write_ms}, total={total_ms}",
        f"Excel generado (una sola hoja): {final_output_path}",
        f"Auditoria JSON: {audit_path}",
    ]

    return ProcessResult(
        success=True,
        output_path=final_output_path,
        label="accion2",
        metadata={
            "console": "\n".join(console_lines),
            "output_origin": "default_path",
            "fallback_used": False,
            "audit_path": str(audit_path),
            "rows": len(rows),
            "summary_rows": len(summary),
            "formula_cells_sanitized": formula_cells_sanitized,
            "runtime": "python-native",
        },
    )
