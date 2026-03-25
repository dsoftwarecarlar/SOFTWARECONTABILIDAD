from __future__ import annotations

import json
import math
import re
import shutil
import tempfile
import time
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from .accion3_native import (
    clear_range_values,
    ensure_template_capacity,
    find_first_local,
    parse_decimal_like,
    parse_int_like,
    round2,
    sanitize_text,
    update_xml_entry,
    write_audit_report,
    write_workbook_with_retries,
)

SHEET_NAME = "MAYOR IVA"
EXPECTED_HEADERS = [
    "COD",
    "CUENTA",
    "EXT",
    "FECHA",
    "ORIGEN",
    "ASIENTO",
    "DOCU",
    "DETALLE",
    "DEBE",
    "HABER",
    "SALDO",
]
CRITICAL_TEMPLATE_ENTRIES = [
    "xl/pivotTables/pivotTable1.xml",
    "xl/worksheets/_rels/sheet1.xml.rels",
    "xl/pivotCache/pivotCacheDefinition1.xml",
]
INVALID_XML_CONTROL_CHARS = re.compile(r"[\u0000-\u0008\u000B\u000C\u000E-\u001F]")


def sanitize_detail_text(value: Any) -> str:
    return re.sub(r"\s+", " ", INVALID_XML_CONTROL_CHARS.sub(" ", "" if value is None else str(value)).replace("\t", " ")).strip()


def parse_date_from_report(value: Any) -> datetime | None:
    clean = sanitize_text(value).upper()
    match = re.fullmatch(r"(\d{2})-([A-Z]{3})-(\d{2})", clean)
    if not match:
        return None

    month_map = {
        "JAN": 1,
        "ENE": 1,
        "FEB": 2,
        "MAR": 3,
        "APR": 4,
        "ABR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AUG": 8,
        "AGO": 8,
        "SEP": 9,
        "SET": 9,
        "OCT": 10,
        "NOV": 11,
        "DEC": 12,
        "DIC": 12,
    }
    day = int(match.group(1))
    month = month_map.get(match.group(2))
    year = 2000 + int(match.group(3))
    if month is None:
        return None

    try:
        return datetime(year, month, day)
    except ValueError:
        return None


def normalize_reference(value: Any) -> str:
    return re.sub(r"\s+", "", sanitize_text(value))


def normalize_docu_for_workbook(value: Any) -> int | str | None:
    clean = normalize_reference(value)
    if not clean:
        return None
    if clean.isdigit():
        parsed = int(clean)
        return parsed if math.isfinite(parsed) else clean
    return clean


def to_cents(value: Any) -> int:
    numeric = float(value)
    if not math.isfinite(numeric):
        return 0
    return int(round((numeric + 1e-12) * 100))


def from_cents(value: int) -> float:
    return value / 100


def read_text_file_best_effort(input_path: Path) -> str:
    raw = input_path.read_bytes()
    utf8 = raw.decode("utf-8", errors="replace")
    if "\ufffd" not in utf8:
        return utf8
    return raw.decode("latin-1")


def validate_account_totals(rows: list[dict[str, Any]], account_header_totals: dict[str, dict[str, Any]]) -> dict[str, Any]:
    if not account_header_totals:
        return {"checked_accounts": 0, "mismatches": []}

    actual_by_code: dict[str, dict[str, int]] = {}
    for row in rows:
        item = actual_by_code.setdefault(
            row["COD"],
            {"total_debe_cents": 0, "total_haber_cents": 0, "saldo_final_cents": 0},
        )
        item["total_debe_cents"] += to_cents(row["DEBE"])
        item["total_haber_cents"] += to_cents(row["HABER"])
        item["saldo_final_cents"] = to_cents(row["SALDO"])

    mismatches: list[dict[str, Any]] = []
    for code, expected in account_header_totals.items():
        actual = actual_by_code.get(
            code,
            {"total_debe_cents": 0, "total_haber_cents": 0, "saldo_final_cents": 0},
        )
        expected_debe = to_cents(expected["total_debe"])
        expected_haber = to_cents(expected["total_haber"])
        expected_saldo = to_cents(expected["saldo_final"])
        diff_debe = actual["total_debe_cents"] - expected_debe
        diff_haber = actual["total_haber_cents"] - expected_haber
        diff_saldo = actual["saldo_final_cents"] - expected_saldo

        if abs(diff_debe) > 1 or abs(diff_haber) > 1 or abs(diff_saldo) > 1:
            mismatches.append(
                {
                    "code": code,
                    "name": expected["name"],
                    "expected": {
                        "total_debe": from_cents(expected_debe),
                        "total_haber": from_cents(expected_haber),
                        "saldo_final": from_cents(expected_saldo),
                    },
                    "actual": {
                        "total_debe": from_cents(actual["total_debe_cents"]),
                        "total_haber": from_cents(actual["total_haber_cents"]),
                        "saldo_final": from_cents(actual["saldo_final_cents"]),
                    },
                }
            )
    return {"checked_accounts": len(account_header_totals), "mismatches": mismatches}


def parse_txt_rows(input_txt_path: Path) -> tuple[list[dict[str, Any]], float, float, dict[str, Any]]:
    raw_text = read_text_file_best_effort(input_txt_path)
    lines = [
        line.replace("\ufeff", "").rstrip()
        for line in raw_text.splitlines()
        if sanitize_text(line) != ""
    ]

    rows: list[dict[str, Any]] = []
    account_header_totals: dict[str, dict[str, Any]] = {}
    skipped_date_rows: list[dict[str, Any]] = []
    date_rows_detected = 0
    file_opening_balance: float | None = None
    file_closing_balance: float | None = None
    account_code_detected: str | None = None
    account_name_detected: str | None = None

    for index, line in enumerate(lines, start=1):
        cols = line.split("\t")
        if len(cols) < 30:
            continue

        date_token = sanitize_text(cols[22]).upper()
        has_date_token = bool(re.fullmatch(r"\d{2}-[A-Z]{3}-\d{2}", date_token))
        if has_date_token:
            date_rows_detected += 1

        code = sanitize_text(cols[6])
        account_name = sanitize_text(cols[7])
        fecha = parse_date_from_report(date_token)
        origen = sanitize_text(cols[23]).upper()
        asiento = parse_int_like(cols[24]) or 0
        ext = sanitize_text(cols[21]).upper() or "N"
        docu = normalize_docu_for_workbook(cols[25])
        detalle = sanitize_detail_text(cols[26])
        debe = round2(parse_decimal_like(cols[27]))
        haber = round2(parse_decimal_like(cols[28]))
        saldo = round2(parse_decimal_like(cols[29]))

        if fecha is None:
            if has_date_token:
                skipped_date_rows.append({"line": index, "reason": "invalid_date_value"})
            continue
        if not re.fullmatch(r"\d{2}\.\d{2}\.\d{2}\.\d{2}\.\d{4}", code):
            skipped_date_rows.append({"line": index, "reason": "invalid_account_code"})
            continue
        if not account_name:
            skipped_date_rows.append({"line": index, "reason": "missing_account_name"})
            continue
        if not re.fullmatch(r"[A-Z]{2,6}", origen):
            skipped_date_rows.append({"line": index, "reason": "invalid_origin"})
            continue

        account_header_totals[code] = {
            "name": account_name,
            "saldo_inicial": round2(parse_decimal_like(cols[8])),
            "total_debe": round2(parse_decimal_like(cols[9])),
            "total_haber": round2(parse_decimal_like(cols[10])),
            "saldo_final": round2(parse_decimal_like(cols[11])),
        }

        if file_opening_balance is None:
            file_opening_balance = round2(parse_decimal_like(cols[8]))
        file_closing_balance = round2(parse_decimal_like(cols[11]))

        if account_code_detected is None:
            account_code_detected = code
            account_name_detected = account_name
        elif account_code_detected != code or account_name_detected != account_name:
            raise ValueError(f"Accion 4 requiere un solo mayor IVA por archivo. Mezcla detectada en {input_txt_path.name}.")

        rows.append(
            {
                "COD": code,
                "CUENTA": account_name,
                "EXT": ext,
                "FECHA": fecha,
                "ORIGEN": origen,
                "ASIENTO": asiento,
                "DOCU": docu,
                "DETALLE": detalle,
                "DEBE": debe,
                "HABER": haber,
                "SALDO": saldo,
            }
        )

    if not rows:
        raise ValueError("No se detectaron movimientos validos en el TXT de Accion 4.")
    if skipped_date_rows:
        preview = " | ".join(f"linea {item['line']} ({item['reason']})" for item in skipped_date_rows[:5])
        raise ValueError(
            f"Se detectaron {len(skipped_date_rows)} filas con fecha que no pudieron mapearse en TXT Accion 4. {preview}"
        )

    totals_validation = validate_account_totals(rows, account_header_totals)
    if totals_validation["mismatches"]:
        preview = " | ".join(
            (
                f"{item['code']} ({item['name']}): "
                f"debe TXT={item['expected']['total_debe']} / Excel={item['actual']['total_debe']}, "
                f"haber TXT={item['expected']['total_haber']} / Excel={item['actual']['total_haber']}, "
                f"saldo TXT={item['expected']['saldo_final']} / Excel={item['actual']['saldo_final']}"
            )
            for item in totals_validation["mismatches"][:4]
        )
        raise ValueError(
            f"Validacion contable fallida en TXT Accion 4 ({len(totals_validation['mismatches'])} cuentas). {preview}"
        )
    if len(account_header_totals) != 1:
        raise ValueError(
            f"Accion 4 solo admite un codigo contable por archivo. Se detectaron {len(account_header_totals)} cuentas en {input_txt_path.name}."
        )

    last_raw_balance = round2(rows[-1]["SALDO"])
    closing_balance = round2(file_closing_balance if file_closing_balance is not None else last_raw_balance)
    if abs(round2(last_raw_balance - closing_balance)) > 0.05:
        raise ValueError(f"El saldo final del archivo {input_txt_path.name} no coincide con la ultima fila del TXT.")

    return rows, round2(file_opening_balance or 0), closing_balance, {
        "source_type": "txt",
        "lines_total": len(lines),
        "date_rows_detected": date_rows_detected,
        "movement_rows_extracted": len(rows),
        "skipped_date_rows": len(skipped_date_rows),
        "account_totals_checked": totals_validation["checked_accounts"],
        "account_total_mismatches": len(totals_validation["mismatches"]),
    }


def validate_rows(rows: list[dict[str, Any]]) -> None:
    problems: list[str] = []
    for index, row in enumerate(rows, start=1):
        if not sanitize_text(row["COD"]):
            problems.append(f"Fila {index}: COD vacio.")
        if not sanitize_text(row["CUENTA"]):
            problems.append(f"Fila {index}: CUENTA vacia.")
        if not isinstance(row["FECHA"], datetime):
            problems.append(f"Fila {index}: FECHA invalida.")
        if not sanitize_text(row["ORIGEN"]):
            problems.append(f"Fila {index}: ORIGEN vacio.")
        if not sanitize_text(row["DETALLE"]):
            problems.append(f"Fila {index}: DETALLE vacio.")
    if problems:
        raise ValueError(f"Validacion de filas fallida ({len(problems)} problemas). {' | '.join(problems[:8])}")


def clone_movement_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "COD": row["COD"],
        "CUENTA": row["CUENTA"],
        "EXT": row["EXT"],
        "FECHA": datetime(row["FECHA"].year, row["FECHA"].month, row["FECHA"].day),
        "ORIGEN": row["ORIGEN"],
        "ASIENTO": int(row["ASIENTO"]),
        "DOCU": row["DOCU"],
        "DETALLE": row["DETALLE"],
        "DEBE": round2(row["DEBE"]),
        "HABER": round2(row["HABER"]),
        "SALDO": round2(row["SALDO"]),
    }


def origin_priority(row: dict[str, Any]) -> int:
    origin = sanitize_text(row["ORIGEN"]).upper()
    priorities = {
        "CP": 1,
        "INVEN": 2,
        "INVSE": 3,
        "SUBCO": 4,
        "REPTO": 5,
        "COMIS": 5,
        "VENSE": 6,
    }
    return priorities.get(origin, 4)


def build_action4_output_plan(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    normalized_rows = [clone_movement_row(row) for row in rows]
    agcm_count = 0
    while agcm_count < len(normalized_rows) and sanitize_text(normalized_rows[agcm_count]["ORIGEN"]).upper() == "AGCM":
        agcm_count += 1

    saldo_offset = round2(-normalized_rows[agcm_count - 1]["SALDO"]) if agcm_count > 0 else 0.0

    adjusted_rows = [
        {
            **clone_movement_row(row),
            "__source_index": index,
            "__priority": origin_priority(row),
            "SALDO": round2(row["SALDO"] + saldo_offset),
        }
        for index, row in enumerate(normalized_rows[agcm_count:])
    ]
    ordered_rows = sorted(adjusted_rows, key=lambda row: (int(row["__priority"]), int(row["__source_index"])))
    movement_rows = [{k: v for k, v in row.items() if not k.startswith("__")} for row in ordered_rows]

    first_block = [row for row in ordered_rows if int(row["__priority"]) <= 4]
    second_block = [row for row in ordered_rows if int(row["__priority"]) == 5]
    third_block = [row for row in ordered_rows if int(row["__priority"]) >= 6]

    row_plan: list[dict[str, Any]] = []
    current_excel_row = 2

    for row in first_block:
        row_plan.append({"type": "movement", "row": {k: v for k, v in row.items() if not k.startswith("__")}})
        current_excel_row += 1

    if first_block and (second_block or third_block):
        first_block_end = current_excel_row - 1
        row_plan.append({"type": "subtotal", "from_row": 2, "to_row": first_block_end})
        current_excel_row += 1
        row_plan.append({"type": "subtotal_balance", "mode": "debe_minus_haber"})
        current_excel_row += 1
        row_plan.extend({"type": "blank"} for _ in range(9))
        current_excel_row += 9

    second_block_start = current_excel_row
    for row in second_block:
        row_plan.append({"type": "movement", "row": {k: v for k, v in row.items() if not k.startswith("__")}})
        current_excel_row += 1

    if second_block and third_block:
        second_block_end = current_excel_row - 1
        row_plan.append({"type": "subtotal", "from_row": second_block_start, "to_row": second_block_end})
        current_excel_row += 1
        row_plan.append({"type": "subtotal_balance", "mode": "haber_minus_debe"})
        current_excel_row += 1
        row_plan.append({"type": "blank"})
        current_excel_row += 1

    for row in third_block:
        row_plan.append({"type": "movement", "row": {k: v for k, v in row.items() if not k.startswith("__")}})

    return movement_rows, row_plan, {
        "agcm_rows_omitidas": agcm_count,
        "comis_rows_movidas_al_final": len(second_block),
        "saldo_offset_aplicado": saldo_offset,
        "subtotal_rows_insertadas": sum(1 for item in row_plan if item["type"] == "subtotal_balance"),
        "blank_rows_insertadas": sum(1 for item in row_plan if item["type"] == "blank"),
    }


def read_template_summary_labels(ws: Worksheet) -> list[dict[str, Any]]:
    labels: list[dict[str, Any]] = []
    for row_index in range(2, 121):
        label = sanitize_text(ws.cell(row_index, 13).value)
        if not label:
            if labels:
                break
            continue
        labels.append({"row": row_index, "label": label})
        if label.upper() == "TOTAL GENERAL":
            break
    return labels


def build_summary(rows: list[dict[str, Any]], template_labels: list[dict[str, Any]]) -> list[dict[str, Any]]:
    bucket: dict[str, dict[str, float]] = {}
    total_debe = 0.0
    total_haber = 0.0

    for row in rows:
        label = sanitize_text(row["CUENTA"])
        item = bucket.setdefault(label, {"debe": 0.0, "haber": 0.0})
        item["debe"] += float(row["DEBE"])
        item["haber"] += float(row["HABER"])
        total_debe += float(row["DEBE"])
        total_haber += float(row["HABER"])

    result: list[dict[str, Any]] = []
    for label_item in template_labels:
        if label_item["label"].upper() == "TOTAL GENERAL":
            result.append({"row": label_item["row"], "debe": round2(total_debe), "haber": round2(total_haber)})
            continue
        values = bucket.get(label_item["label"], {"debe": 0.0, "haber": 0.0})
        result.append({"row": label_item["row"], "debe": round2(values["debe"]), "haber": round2(values["haber"])})
    return result


def build_workbook_from_template(
    template_path: Path,
    row_plan: list[dict[str, Any]],
    movement_rows: list[dict[str, Any]],
) -> tuple[Any, list[dict[str, Any]]]:
    workbook = load_workbook(template_path, keep_links=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"La plantilla de Accion 4 no contiene la hoja {SHEET_NAME}.")

    ws = workbook[SHEET_NAME]
    max_rows = max(ws.max_row, len(row_plan) + 1)
    ensure_template_capacity(ws, max_rows)
    clear_range_values(ws, 2, max_rows, 1, 11)
    clear_range_values(ws, 2, max_rows, 14, 15)

    for row_index, plan_item in enumerate(row_plan, start=2):
        if plan_item["type"] == "movement":
            row = plan_item["row"]
            ws.cell(row_index, 1).value = row["COD"]
            ws.cell(row_index, 2).value = row["CUENTA"]
            ws.cell(row_index, 3).value = row["EXT"]
            ws.cell(row_index, 4).value = row["FECHA"]
            ws.cell(row_index, 5).value = row["ORIGEN"]
            ws.cell(row_index, 6).value = row["ASIENTO"]
            ws.cell(row_index, 7).value = row["DOCU"]
            ws.cell(row_index, 8).value = row["DETALLE"]
            ws.cell(row_index, 9).value = round2(row["DEBE"])
            ws.cell(row_index, 10).value = round2(row["HABER"])
            ws.cell(row_index, 11).value = round2(row["SALDO"])
        elif plan_item["type"] == "subtotal":
            ws.cell(row_index, 9).value = f"=SUM(I{plan_item['from_row']}:I{plan_item['to_row']})"
            ws.cell(row_index, 10).value = f"=SUM(J{plan_item['from_row']}:J{plan_item['to_row']})"
        elif plan_item["type"] == "subtotal_balance":
            if plan_item["mode"] == "haber_minus_debe":
                ws.cell(row_index, 11).value = f"=+J{row_index - 1}-I{row_index - 1}"
            else:
                ws.cell(row_index, 11).value = f"=+I{row_index - 1}-J{row_index - 1}"

    summary = build_summary(movement_rows, read_template_summary_labels(ws))
    for item in summary:
        ws.cell(item["row"], 14).value = item["debe"]
        ws.cell(item["row"], 15).value = item["haber"]

    ws.auto_filter.ref = f"A1:K{max(2, len(row_plan) + 1)}"
    return workbook, summary


def preserve_visual_artifacts(template_path: Path, output_path: Path, data_row_count: int) -> None:
    ref = f"A1:K{max(2, data_row_count + 1)}"
    with zipfile.ZipFile(template_path, "r") as template_zip, zipfile.ZipFile(output_path, "r") as output_zip:
        output_entries = {name: output_zip.read(name) for name in output_zip.namelist()}
        for entry in CRITICAL_TEMPLATE_ENTRIES:
            if entry not in output_entries and entry in template_zip.namelist():
                output_entries[entry] = template_zip.read(entry)

    def patch_pivot_cache(root: ET.Element) -> None:
        worksheet_source = find_first_local(root, "worksheetSource")
        if worksheet_source is not None:
            worksheet_source.set("ref", ref)
            worksheet_source.set("sheet", SHEET_NAME)
        cache_definition = find_first_local(root, "pivotCacheDefinition")
        if cache_definition is not None:
            cache_definition.set("recordCount", str(max(0, data_row_count)))

    if "xl/pivotCache/pivotCacheDefinition1.xml" in output_entries:
        output_entries["xl/pivotCache/pivotCacheDefinition1.xml"] = update_xml_entry(
            output_entries["xl/pivotCache/pivotCacheDefinition1.xml"],
            patch_pivot_cache,
        )

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_path = Path(temp_file.name)

    try:
        with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as target_zip:
            for name, payload in output_entries.items():
                target_zip.writestr(name, payload)
        shutil.move(str(temp_path), str(output_path))
    finally:
        temp_path.unlink(missing_ok=True)


def styles_match(left_cell: Any, right_cell: Any) -> bool:
    return left_cell._style == right_cell._style


def verify_output(output_path: Path, template_path: Path, row_plan: list[dict[str, Any]]) -> None:
    with zipfile.ZipFile(output_path, "r") as output_zip, zipfile.ZipFile(template_path, "r") as template_zip:
        for entry in CRITICAL_TEMPLATE_ENTRIES:
            if entry in template_zip.namelist() and entry not in output_zip.namelist():
                raise RuntimeError(f"Validacion final: falta artefacto visual de plantilla ({entry}).")

    workbook = load_workbook(output_path, keep_links=True)
    if workbook.sheetnames != [SHEET_NAME]:
        raise RuntimeError(f"Validacion final: el archivo debe tener una sola hoja llamada {SHEET_NAME}.")

    ws = workbook[SHEET_NAME]
    headers = [sanitize_text(ws.cell(1, index).value).upper() for index in range(1, len(EXPECTED_HEADERS) + 1)]
    if headers != EXPECTED_HEADERS:
        raise RuntimeError("Validacion final: encabezados incorrectos en Accion 4.")

    for row_index, plan_item in enumerate(row_plan, start=2):
        if plan_item["type"] != "movement":
            continue
        value = ws.cell(row_index, 4).value
        if not isinstance(value, (datetime, date)):
            raise RuntimeError(f"Validacion final: FECHA invalida en fila {row_index}.")

    template_workbook = load_workbook(template_path, keep_links=True)
    template_ws = template_workbook[SHEET_NAME]
    for col_index in range(1, 17):
        if not styles_match(template_ws.cell(1, col_index), ws.cell(1, col_index)):
            raise RuntimeError(f"Validacion final: estilo de encabezado alterado en fila 1, columna {col_index}.")

    last_movement_row = 1
    for index in range(len(row_plan) - 1, -1, -1):
        if row_plan[index]["type"] == "movement":
            last_movement_row = index + 2
            break
    if last_movement_row >= 2:
        expected_data_style_row = min(last_movement_row, template_ws.max_row)
        for col_index in range(1, 12):
            if not styles_match(template_ws.cell(expected_data_style_row, col_index), ws.cell(last_movement_row, col_index)):
                raise RuntimeError(f"Validacion final: estilo de datos alterado en {last_movement_row}:{col_index}.")

    for item in read_template_summary_labels(template_ws):
        for col_index in range(13, 17):
            if not styles_match(template_ws.cell(item["row"], col_index), ws.cell(item["row"], col_index)):
                raise RuntimeError(f"Validacion final: estilo lateral alterado en {item['row']}:{col_index}.")


def run(request: ProcessRequest) -> ProcessResult:
    input_paths = [path.resolve() for path in request.input_paths]
    if len(input_paths) != 1:
        raise ValueError("Accion 4 requiere exactamente un archivo TXT de entrada.")
    if input_paths[0].suffix.lower() != ".txt":
        raise ValueError("Solo se permiten archivos TXT para Accion 4.")
    if request.template_path is None:
        raise ValueError("Accion 4 requiere una plantilla XLSX.")

    input_path = input_paths[0]
    template_path = request.template_path.resolve()
    output_path = request.output_path.resolve()
    started_at = time.perf_counter()

    rows, opening_balance, closing_balance, diagnostics = parse_txt_rows(input_path)
    parse_ms = int((time.perf_counter() - started_at) * 1000)
    validate_rows(rows)
    validate_ms = int((time.perf_counter() - started_at) * 1000) - parse_ms
    movement_rows, row_plan, transform = build_action4_output_plan(rows)

    build_started = time.perf_counter()
    workbook, summary = build_workbook_from_template(template_path, row_plan, movement_rows)
    build_ms = int((time.perf_counter() - build_started) * 1000)

    write_started = time.perf_counter()
    final_output_path = write_workbook_with_retries(workbook, output_path)
    preserve_visual_artifacts(template_path, final_output_path, len(row_plan))
    verify_output(final_output_path, template_path, row_plan)
    write_ms = int((time.perf_counter() - write_started) * 1000)
    total_ms = int((time.perf_counter() - started_at) * 1000)

    audit_path = final_output_path.with_name(f"{final_output_path.stem}_auditoria.json")
    write_audit_report(
        audit_path,
        {
            "fecha_proceso": datetime.now().isoformat(),
            "input_source": str(input_path),
            "input_sources": [str(input_path)],
            "input_pdf": None,
            "input_tipo": diagnostics["source_type"],
            "archivos_origen": [{"path": str(input_path), "name": input_path.name, "type": "txt"}],
            "total_archivos_origen": 1,
            "orden_consolidacion": "orden_de_carga",
            "output_xlsx": str(final_output_path),
            "hoja_salida": SHEET_NAME,
            "movimientos_extraidos": len(movement_rows),
            "filas_salida_totales": len(row_plan),
            "filas_fecha_detectadas": diagnostics["date_rows_detected"],
            "filas_fecha_omitidas": diagnostics["skipped_date_rows"],
            "cuentas_validadas_fuente": diagnostics["account_totals_checked"],
            "cuentas_descuadradas_fuente": diagnostics["account_total_mismatches"],
            "saldo_inicial_consolidado": opening_balance,
            "saldo_final_consolidado": closing_balance,
            "cuentas_validadas_pdf": diagnostics["account_totals_checked"],
            "cuentas_descuadradas_pdf": diagnostics["account_total_mismatches"],
            "resumen_filtrado": len(summary),
            "transformacion_manual": transform,
            "timings_ms": {
                "parse": parse_ms,
                "validate": validate_ms,
                "build": build_ms,
                "write": write_ms,
                "total": total_ms,
            },
            "verificacion_final_ok": True,
        },
    )

    console_lines = [
        f"Archivo leido: {input_path}",
        f"Movimientos extraidos: {len(movement_rows)}",
        f"Filas totales en salida: {len(row_plan)}",
        f"Resumen lateral: {len(summary)} filas",
        f"Rendimiento (ms): parse={parse_ms}, validate={validate_ms}, build={build_ms}, write={write_ms}, total={total_ms}",
        f"Excel generado (una sola hoja): {final_output_path}",
        f"Auditoria JSON: {audit_path}",
    ]

    return ProcessResult(
        success=True,
        output_path=final_output_path,
        label="accion4",
        metadata={
            "console": "\n".join(console_lines),
            "output_origin": "default_path",
            "fallback_used": False,
            "runtime": "python-native",
            "audit_path": str(audit_path),
            "rows": len(movement_rows),
            "source_count": 1,
        },
    )
