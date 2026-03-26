from __future__ import annotations

import warnings
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook_quiet(path: Path | str, **kwargs: Any):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style.*",
            category=UserWarning,
        )
        return load_workbook(path, **kwargs)


def copy_row_dimensions(
    source_worksheet: Worksheet,
    target_worksheet: Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    source_dimension = source_worksheet.row_dimensions[source_row]
    target_dimension = target_worksheet.row_dimensions[target_row]
    for attribute in ("height", "hidden", "outlineLevel", "collapsed", "ht"):
        if hasattr(source_dimension, attribute):
            setattr(target_dimension, attribute, getattr(source_dimension, attribute))


def copy_cell_style(source_cell: Any, target_cell: Any) -> None:
    if isinstance(source_cell, MergedCell) or isinstance(target_cell, MergedCell):
        return
    if getattr(source_cell, "has_style", False):
        target_cell._style = copy(source_cell._style)


def copy_row_style(
    source_worksheet: Worksheet,
    target_worksheet: Worksheet,
    source_row: int,
    target_row: int,
    last_column: int,
) -> None:
    copy_row_dimensions(source_worksheet, target_worksheet, source_row, target_row)
    for column in range(1, last_column + 1):
        source_cell = source_worksheet.cell(row=source_row, column=column)
        target_cell = target_worksheet.cell(row=target_row, column=column)
        copy_cell_style(source_cell, target_cell)


def get_row_merge_definitions(
    worksheet: Worksheet,
    row_number: int,
    last_column: int | None = None,
    start_column: int = 1,
) -> list[tuple[int, int]]:
    definitions: list[tuple[int, int]] = []
    max_column = None if last_column is None else int(last_column)
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row != row_number or merged_range.max_row != row_number:
            continue
        if merged_range.max_col < start_column:
            continue
        if max_column is not None and merged_range.min_col > max_column:
            continue
        definitions.append((int(merged_range.min_col), int(merged_range.max_col)))
    definitions.sort()
    return definitions


def clear_merges_in_row_range(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    last_column: int | None = None,
    start_column: int = 1,
) -> None:
    if start_row > end_row:
        return

    max_column = None if last_column is None else int(last_column)
    to_unmerge: list[str] = []
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.min_row < start_row or merged_range.max_row > end_row:
            continue
        if merged_range.max_col < start_column:
            continue
        if max_column is not None and merged_range.min_col > max_column:
            continue
        to_unmerge.append(str(merged_range))

    for range_ref in to_unmerge:
        worksheet.unmerge_cells(range_ref)


def apply_row_merge_definitions(
    worksheet: Worksheet,
    row_number: int,
    merge_definitions: list[tuple[int, int]],
) -> None:
    clear_merges_in_row_range(worksheet, row_number, row_number)
    for left, right in merge_definitions:
        worksheet.merge_cells(
            start_row=row_number,
            start_column=int(left),
            end_row=row_number,
            end_column=int(right),
        )


def row_has_merged_cells(
    worksheet: Worksheet,
    row_number: int,
    last_column: int,
    start_column: int = 1,
) -> bool:
    for column in range(start_column, last_column + 1):
        if isinstance(worksheet.cell(row=row_number, column=column), MergedCell):
            return True
    return False


def row_has_visible_value(
    worksheet: Worksheet,
    row_number: int,
    last_column: int,
    start_column: int = 1,
) -> bool:
    for column in range(start_column, last_column + 1):
        cell = worksheet.cell(row=row_number, column=column)
        if isinstance(cell, MergedCell):
            continue
        value = cell.value
        if value not in (None, ""):
            return True
    return False


def find_style_source_row(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    last_column: int,
) -> int:
    for row_number in range(end_row, start_row - 1, -1):
        if row_has_merged_cells(worksheet, row_number, last_column):
            continue
        if row_has_visible_value(worksheet, row_number, last_column):
            return row_number
    return max(start_row, end_row)


def unmerge_ranges_in_band(
    worksheet: Worksheet,
    start_row: int,
    end_row: int,
    last_column: int,
) -> set[int]:
    affected_rows: set[int] = set()
    to_unmerge: list[str] = []
    for merged_range in list(worksheet.merged_cells.ranges):
        if merged_range.max_col < 1 or merged_range.min_col > last_column:
            continue
        if merged_range.max_row < start_row or merged_range.min_row > end_row:
            continue
        to_unmerge.append(str(merged_range))
        for row_number in range(
            max(start_row, merged_range.min_row),
            min(end_row, merged_range.max_row) + 1,
        ):
            affected_rows.add(row_number)

    for range_ref in to_unmerge:
        worksheet.unmerge_cells(range_ref)

    return affected_rows


def clear_row_values(worksheet: Worksheet, row_number: int, last_column: int) -> None:
    for column in range(1, last_column + 1):
        cell = worksheet.cell(row=row_number, column=column)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None


def clear_rows(worksheet: Worksheet, start_row: int, end_row: int, last_column: int) -> None:
    if start_row > end_row:
        return
    for row_number in range(start_row, end_row + 1):
        clear_row_values(worksheet, row_number, last_column)


def prepare_marker_row(
    worksheet: Worksheet,
    template_row: int,
    target_row: int,
    last_column: int,
    label_column: int,
    label_text: str = "MAYOR",
) -> int:
    copy_row_style(worksheet, worksheet, template_row, target_row, last_column)
    clear_row_values(worksheet, target_row, last_column)
    label_cell = worksheet.cell(row=target_row, column=label_column)
    if not isinstance(label_cell, MergedCell):
        label_cell.value = label_text
        label_cell.data_type = "s"
    return target_row


def write_literal_string(cell: Any, value: str) -> None:
    cell.value = value
    cell.data_type = "s"
    if hasattr(cell, "quotePrefix"):
        cell.quotePrefix = value.startswith(("=", "+", "-", "@"))
