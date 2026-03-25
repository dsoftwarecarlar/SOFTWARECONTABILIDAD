from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from ..cxp_actions.accion3_native import write_workbook_with_retries
from .nc_stage import load_source_workbook as load_nc_source_workbook
from .rep_stage import load_source_workbook as load_rep_source_workbook
from .workbook_tools import copy_row_style, load_workbook_quiet

INVALID_XML_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

SOURCE_CONFIGS = (
    {
        "key": "tyt",
        "label": "MATRIZ",
        "rep_field": "excel_tyt",
        "nc_field": "excel_nc_tyt",
        "my_sheet": "MY REP TYT",
        "optional_nc": False,
    },
    {
        "key": "peug",
        "label": "PEUGEOT",
        "rep_field": "excel_peug",
        "nc_field": "excel_nc_peug",
        "my_sheet": "MY REP PEUG",
        "optional_nc": False,
    },
    {
        "key": "chgn",
        "label": "CHANGAN",
        "rep_field": "excel_chgn",
        "nc_field": "excel_nc_chgn",
        "my_sheet": "MY REP CHGN",
        "optional_nc": True,
    },
    {
        "key": "szk",
        "label": "SUZUKI",
        "rep_field": "excel_szk",
        "nc_field": "excel_nc_szk",
        "my_sheet": "MY REP SZK",
        "optional_nc": False,
    },
)

MY_LAYOUTS: dict[str, dict[str, Any]] = {
    "tyt": {
        "my_sheet_name": "MY REP TYT",
        "detail_column": 7,
        "seat_column": 6,
        "date_column": 4,
        "debit_column": 8,
        "credit_column": 9,
        "saldo_column": 10,
        "sections": (
            {"name": "sales", "start_row": 2, "end_row": 37, "amount_column": 9, "opposite_column": 8},
            {"name": "discount", "start_row": 42, "end_row": 52, "amount_column": 8, "opposite_column": 9},
        ),
    },
    "peug": {
        "my_sheet_name": "MY REP PEUG",
        "detail_column": 7,
        "seat_column": 6,
        "date_column": 4,
        "debit_column": 8,
        "credit_column": 9,
        "saldo_column": 10,
        "sections": (
            {"name": "sales", "start_row": 2, "end_row": 18, "amount_column": 9, "opposite_column": 8},
            {"name": "discount", "start_row": 21, "end_row": 24, "amount_column": 8, "opposite_column": 9},
        ),
    },
    "chgn": {
        "my_sheet_name": "MY REP CHGN",
        "detail_column": 8,
        "seat_column": 6,
        "date_column": 4,
        "debit_column": 9,
        "credit_column": 10,
        "saldo_column": 11,
        "sections": (
            {"name": "sales", "start_row": 2, "end_row": 7, "amount_column": 10, "opposite_column": 9},
            {"name": "discount", "start_row": 12, "end_row": 14, "amount_column": 9, "opposite_column": 10},
        ),
    },
    "szk": {
        "my_sheet_name": "MY REP SZK",
        "detail_column": 7,
        "seat_column": 6,
        "date_column": 4,
        "debit_column": 8,
        "credit_column": 9,
        "saldo_column": 10,
        "sections": (
            {"name": "sales", "start_row": 2, "end_row": 30, "amount_column": 9, "opposite_column": 8},
            {"name": "discount", "start_row": 34, "end_row": 49, "amount_column": 8, "opposite_column": 9},
        ),
    },
}

MY_ACCOUNT_STATIC_METADATA = {
    "04.01.01.01.0001": {"name": "VTAS RPTOS TOYOTA - CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.01.0003": {"name": "VTAS RPTOS TOYOTA - CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.01.0005": {"name": "DESC VTAS RPTOS TOYOTA CONTADO  CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.01.0007": {"name": "DESC VTAS RPTOS TOYOTA CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.01.0009": {"name": "DEVOL VTAS RPTOS TOYOTA CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.01.0011": {"name": "DEVOL VTAS RPTOS TOYOTA CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.02.0001": {"name": "VTAS RPTOS CHANGAN - CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.02.0003": {"name": "VTAS RPTOS CHANGAN - CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.02.0005": {"name": "DESC VTAS RPTOS CHANGAN CONTADO  CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.02.0007": {"name": "DESC VTAS RPTOS CHANGAN CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.02.0009": {"name": "DEVOL VTAS RPTOS CHANGAN CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.02.0011": {"name": "DEVOL VTAS RPTOS CHANGAN CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.03.0001": {"name": "VTAS RPTOS PEUGEOT - CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.03.0003": {"name": "VTAS RPTOS PEUGEOT - CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.03.0005": {"name": "DESC VTAS RPTOS PEUGEOT CONTADO  CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.03.0007": {"name": "DESC VTAS RPTOS PEUGEOT CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.03.0009": {"name": "DEVOL VTAS RPTOS PEUGEOT CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.03.0011": {"name": "DEVOL VTAS RPTOS PEUGEOT CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.04.0001": {"name": "VTAS RPTOS SUZUKI - CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.04.0003": {"name": "VTAS RPTOS SUZUKI - CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.04.0005": {"name": "DESC VTAS RPTOS SUZUKI CONTADO  CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.04.0007": {"name": "DESC VTAS RPTOS SUZUKI CREDITO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.04.0009": {"name": "DEVOL VTAS RPTOS SUZUKI CONTADO CON IVA", "marker": "N", "type": "REPTO"},
    "04.01.01.04.0011": {"name": "DEVOL VTAS RPTOS SUZUKI CREDITO CON IVA", "marker": "N", "type": "REPTO"},
}


def literal_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return INVALID_XML_CONTROL_CHARS.sub(" ", str(value)).strip()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", literal_text(value)).strip()


def round_amount(value: Any, decimals: int = 2) -> float:
    return round(float(value or 0) + 1e-12, decimals)


def parse_date_text(value: Any) -> datetime | None:
    text = normalize_text(value)
    if text == "":
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return datetime(parsed.year, parsed.month, parsed.day)
        except ValueError:
            continue
    return None


def convert_date_value(raw_value: Any, fallback_text: Any = "") -> datetime | None:
    if isinstance(raw_value, datetime):
        return datetime(raw_value.year, raw_value.month, raw_value.day)
    if isinstance(raw_value, date):
        return datetime(raw_value.year, raw_value.month, raw_value.day)
    return parse_date_text(fallback_text if fallback_text not in (None, "") else raw_value)


def worksheet_cell_text(worksheet: Worksheet, row: int, column: int) -> str:
    cell = worksheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return ""
    return literal_text(cell.value)


def worksheet_cell_number(worksheet: Worksheet, row: int, column: int) -> float:
    cell = worksheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return 0.0
    value = cell.value
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    if text == "":
        return 0.0
    normalized = re.sub(r"[^\d,.\-]", "", text)
    if normalized == "":
        return 0.0
    if "," in normalized and "." in normalized:
        if normalized.rfind(".") > normalized.rfind(","):
            normalized = normalized.replace(",", "")
        else:
            normalized = normalized.replace(".", "").replace(",", ".")
    elif "," in normalized:
        if normalized.count(",") > 1:
            normalized = normalized.replace(",", "")
        else:
            integer, _, fraction = normalized.partition(",")
            normalized = f"{integer}.{fraction}"
    elif normalized.count(".") > 1:
        last_dot = normalized.rfind(".")
        integer = normalized[:last_dot].replace(".", "")
        fraction = normalized[last_dot + 1 :]
        normalized = f"{integer}.{fraction}"
    try:
        return float(normalized)
    except ValueError:
        return 0.0


def normalize_doc_number(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def natural_key(value: Any) -> tuple[tuple[int, Any], ...]:
    parts = re.split(r"(\d+)", normalize_text(value).upper())
    key: list[tuple[int, Any]] = []
    for part in parts:
        if part == "":
            continue
        if part.isdigit():
            key.append((0, int(part)))
        else:
            key.append((1, part))
    return tuple(key)


def date_sort_key(value: datetime | None) -> tuple[int, datetime]:
    if isinstance(value, datetime):
        return (0, value)
    return (1, datetime.max)


def seat_sort_key(value: Any) -> tuple[int, Any]:
    text = normalize_text(value)
    return (0, int(text)) if text.isdigit() else (1, natural_key(text))


def compare_posting_key(entry: dict[str, Any]) -> tuple[Any, ...]:
    return (
        date_sort_key(entry["date_value"]),
        seat_sort_key(entry["seat"]),
        natural_key(entry["detail"]),
        natural_key(entry["account"]),
        natural_key(entry["side"]),
    )


def add_grouped_amount(
    groups: dict[str, dict[str, Any]],
    group_key: str,
    amount: Any,
    date_value: datetime | None,
    date_text: str,
    seat: str,
    detail: str,
) -> None:
    rounded_amount = round_amount(amount, 6)
    if abs(rounded_amount) < 0.0000001:
        return
    if group_key not in groups:
        groups[group_key] = {
            "amount": 0.0,
            "date_value": date_value,
            "date_text": normalize_text(date_text),
            "seat": normalize_text(seat),
            "detail": normalize_text(detail),
        }
    entry = groups[group_key]
    entry["amount"] = round_amount(entry["amount"] + rounded_amount, 6)
    if isinstance(date_value, datetime):
        current = entry.get("date_value")
        if not isinstance(current, datetime) or date_value < current:
            entry["date_value"] = date_value
            entry["date_text"] = normalize_text(date_text)
    if normalize_text(entry.get("date_text", "")) == "" and normalize_text(date_text) != "":
        entry["date_text"] = normalize_text(date_text)


def get_rep_detail_name(key: str, agency: Any) -> str:
    normalized_agency = normalize_text(agency)
    if key == "tyt":
        return "MOD. REPUESTOS REP01"
    if key == "peug":
        return "MOD. REPUESTOS REP06"
    if key == "chgn":
        return "MOD. REPUESTOS REP05"
    if key == "szk":
        return "MOD. REPUESTOS REP08" if normalized_agency == "08" else "MOD. REPUESTOS REP07"
    raise ValueError(f"Unsupported key: {key}")


def get_posting_account(key: str, category: str, form: Any) -> str:
    normalized_form = normalize_text(form).upper()
    mappings = {
        "tyt": {
            "sales": {"CONTADO": "04.01.01.01.0001", "CREDITO": "04.01.01.01.0003"},
            "discount": {"CONTADO": "04.01.01.01.0005", "CREDITO": "04.01.01.01.0007"},
        },
        "peug": {
            "sales": {"CONTADO": "04.01.01.03.0001", "CREDITO": "04.01.01.03.0003"},
            "discount": {"CONTADO": "04.01.01.03.0005", "CREDITO": "04.01.01.03.0007"},
        },
        "chgn": {
            "sales": {"CONTADO": "04.01.01.02.0001", "CREDITO": "04.01.01.02.0003"},
            "discount": {"CONTADO": "04.01.01.02.0005", "CREDITO": "04.01.01.02.0007"},
        },
        "szk": {
            "sales": {"CONTADO": "04.01.01.04.0001", "CREDITO": "04.01.01.04.0003"},
            "discount": {"CONTADO": "04.01.01.04.0005", "CREDITO": "04.01.01.04.0007"},
        },
    }
    return str(mappings.get(key, {}).get(category, {}).get(normalized_form, ""))


def get_devol_account(key: str, form: Any) -> str:
    normalized_form = normalize_text(form).upper()
    mappings = {
        "tyt": {"CONTADO": "04.01.01.01.0009", "CREDITO": "04.01.01.01.0011"},
        "peug": {"CONTADO": "04.01.01.03.0009", "CREDITO": "04.01.01.03.0011"},
        "chgn": {"CONTADO": "04.01.01.02.0009", "CREDITO": "04.01.01.02.0011"},
        "szk": {"CONTADO": "04.01.01.04.0009", "CREDITO": "04.01.01.04.0011"},
    }
    return str(mappings.get(key, {}).get(normalized_form, ""))


def infer_form_from_description(description: Any) -> str:
    text = normalize_text(description).upper()
    if "CREDITO" in text:
        return "CREDITO"
    if "CONTADO" in text:
        return "CONTADO"
    return ""


def build_rep_posting_groups(source_data: Any, key: str) -> dict[str, dict[str, dict[str, Any]]]:
    sales: dict[str, dict[str, Any]] = {}
    discount: dict[str, dict[str, Any]] = {}
    vat: dict[str, dict[str, Any]] = {}
    worksheet = source_data.worksheet
    for row in range(11, source_data.total_row):
        seat = worksheet_cell_text(worksheet, row, 38)
        if seat == "" or normalize_text(seat).upper() == "ASIENTO":
            continue
        form = worksheet_cell_text(worksheet, row, 40).upper()
        if form not in {"CONTADO", "CREDITO"}:
            continue
        agency = worksheet_cell_text(worksheet, row, 39)
        detail = get_rep_detail_name(key, agency)
        raw_date = worksheet.cell(row=row, column=3).value
        date_text = worksheet_cell_text(worksheet, row, 3)
        date_value = convert_date_value(raw_date, date_text)
        sales_amount = worksheet_cell_number(worksheet, row, 18)
        discount_amount = worksheet_cell_number(worksheet, row, 20)
        vat_amount = worksheet_cell_number(worksheet, row, 26)
        sales_account = get_posting_account(key, "sales", form)
        if sales_account != "":
            add_grouped_amount(sales, f"{sales_account}|{seat}|{detail}", sales_amount, date_value, date_text, seat, detail)
        discount_account = get_posting_account(key, "discount", form)
        if discount_account != "":
            add_grouped_amount(
                discount,
                f"{discount_account}|{seat}|{detail}",
                discount_amount,
                date_value,
                date_text,
                seat,
                detail,
            )
        add_grouped_amount(vat, f"{seat}|{detail}", vat_amount, date_value, date_text, seat, detail)
    return {"sales": sales, "discount": discount, "vat": vat}


def build_nc_rows_from_source(source_data: Any | None, key: str) -> list[dict[str, Any]]:
    if source_data is None or getattr(source_data, "worksheet", None) is None:
        return []
    worksheet = source_data.worksheet
    rows: list[dict[str, Any]] = []
    for row in range(8, source_data.total_row):
        nc_number = normalize_doc_number(worksheet_cell_text(worksheet, row, 9))
        original_doc = normalize_doc_number(worksheet_cell_text(worksheet, row, 10))
        description = worksheet_cell_text(worksheet, row, 11)
        sub_total = worksheet_cell_number(worksheet, row, 23)
        total = worksheet_cell_number(worksheet, row, 36)
        if nc_number == "" and original_doc == "" and description == "" and abs(sub_total) < 0.0000001 and abs(total) < 0.0000001:
            continue
        nc_date_text = worksheet_cell_text(worksheet, row, 4)
        fact_date_text = worksheet_cell_text(worksheet, row, 14)
        agency = worksheet_cell_text(worksheet, row, 5)
        bodega = worksheet_cell_text(worksheet, row, 3)
        rows.append(
            {
                "nc_date_text": nc_date_text,
                "nc_date_value": convert_date_value(worksheet.cell(row=row, column=4).value, nc_date_text),
                "fact_date_text": fact_date_text,
                "fact_date_value": convert_date_value(worksheet.cell(row=row, column=14).value, fact_date_text),
                "agency": agency,
                "bodega": bodega,
                "seat": worksheet_cell_text(worksheet, row, 41),
                "description": description,
                "sub_total": sub_total,
                "discount": worksheet_cell_number(worksheet, row, 24),
                "vat": worksheet_cell_number(worksheet, row, 30),
                "form": infer_form_from_description(description),
                "detail": get_rep_detail_name(key, agency or bodega),
            }
        )
    return rows


def build_nc_groups_from_rows(rows: list[dict[str, Any]], key: str) -> dict[str, dict[str, dict[str, Any]]]:
    discount_credit: dict[str, dict[str, Any]] = {}
    devol: dict[str, dict[str, Any]] = {}
    vat: dict[str, dict[str, Any]] = {}
    for row in rows:
        form = row["form"] or infer_form_from_description(row["description"])
        seat = normalize_text(row["seat"])
        detail = normalize_text(row["detail"] or get_rep_detail_name(key, row["agency"] or row["bodega"]))
        date_value = row["nc_date_value"] or row["fact_date_value"]
        date_text = row["nc_date_text"] or row["fact_date_text"]
        discount_account = get_posting_account(key, "discount", form)
        if discount_account != "" and seat != "" and detail != "":
            add_grouped_amount(
                discount_credit,
                f"{discount_account}|{seat}|{detail}",
                row["discount"],
                date_value,
                date_text,
                seat,
                detail,
            )
        devol_account = get_devol_account(key, form)
        if devol_account != "" and seat != "" and detail != "":
            add_grouped_amount(
                devol,
                f"{devol_account}|{seat}|{detail}",
                row["sub_total"],
                date_value,
                date_text,
                seat,
                detail,
            )
        if seat != "" and detail != "":
            add_grouped_amount(vat, f"{seat}|{detail}", row["vat"], date_value, date_text, seat, detail)
    return {"discount_credit": discount_credit, "devol": devol, "vat": vat}


def build_posting_entries(group_map: dict[str, dict[str, Any]], side: str) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    for group_key, group in (group_map or {}).items():
        parts = group_key.split("|", 2)
        if len(parts) < 3:
            continue
        account, seat, detail = parts
        amount = round_amount(group.get("amount", 0), 2)
        if account == "" or detail == "" or abs(amount) < 0.0000001:
            continue
        entries.append(
            {
                "account": account,
                "seat": seat or "",
                "detail": detail or "",
                "date_value": group.get("date_value"),
                "date_text": group.get("date_text", ""),
                "amount": amount,
                "side": side,
            }
        )
    return entries


def get_section_last_column(layout: dict[str, Any]) -> int:
    return max(
        int(layout["date_column"]),
        int(layout["seat_column"]),
        int(layout["detail_column"]),
        int(layout["debit_column"]),
        int(layout["credit_column"]),
        int(layout["saldo_column"]),
    )


def get_template_account_blocks(template_sheet: Worksheet, section: dict[str, Any]) -> list[dict[str, Any]]:
    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in range(section["start_row"], section["end_row"] + 1):
        account = worksheet_cell_text(template_sheet, row, 1)
        if account == "":
            if current is not None:
                blocks.append(current)
                current = None
            continue
        if current is None or current["account"] != account:
            if current is not None:
                blocks.append(current)
            current = {"account": account, "start_row": row, "end_row": row}
            continue
        current["end_row"] = row
    if current is not None:
        blocks.append(current)
    return blocks


def get_section_opening_balances(template_values_sheet: Worksheet, layout: dict[str, Any], section: dict[str, Any]) -> dict[str, float]:
    balances: dict[str, float] = {}
    for block in get_template_account_blocks(template_values_sheet, section):
        debit = worksheet_cell_number(template_values_sheet, block["start_row"], layout["debit_column"])
        credit = worksheet_cell_number(template_values_sheet, block["start_row"], layout["credit_column"])
        saldo = worksheet_cell_number(template_values_sheet, block["start_row"], layout["saldo_column"])
        balances[block["account"]] = round_amount(saldo - debit + credit, 2)
    return balances


def get_section_static_columns(layout: dict[str, Any]) -> list[int]:
    dynamic_columns = {
        int(layout["date_column"]),
        int(layout["seat_column"]),
        int(layout["detail_column"]),
        int(layout["debit_column"]),
        int(layout["credit_column"]),
        int(layout["saldo_column"]),
    }
    return [column for column in range(1, get_section_last_column(layout) + 1) if column not in dynamic_columns]


def get_section_static_templates(template_sheet: Worksheet, layout: dict[str, Any], section: dict[str, Any]) -> dict[str, dict[int, Any]]:
    templates: dict[str, dict[int, Any]] = {}
    for block in get_template_account_blocks(template_sheet, section):
        templates[block["account"]] = {}
        for column in get_section_static_columns(layout):
            templates[block["account"]][column] = copy(template_sheet.cell(row=block["start_row"], column=column).value)
    return templates


def get_account_order(template_sheet: Worksheet, section: dict[str, Any]) -> dict[str, int]:
    return {block["account"]: index for index, block in enumerate(get_template_account_blocks(template_sheet, section))}


def get_my_account_static_fallback(layout: dict[str, Any], account: str) -> dict[int, Any]:
    metadata = MY_ACCOUNT_STATIC_METADATA.get(normalize_text(account))
    if metadata is None:
        return {}
    static_columns = set(get_section_static_columns(layout))
    values: dict[int, Any] = {}
    if 2 in static_columns:
        values[2] = metadata["name"]
    if 3 in static_columns:
        values[3] = metadata.get("marker", "N")
    if 5 in static_columns:
        values[5] = metadata.get("type", "REPTO")
    return values


def get_devol_section_range(template_sheet: Worksheet, layout: dict[str, Any]) -> dict[str, Any] | None:
    last_section_end = max(section["end_row"] for section in layout["sections"])
    start_row: int | None = None
    end_row: int | None = None
    for row in range(last_section_end + 1, max(template_sheet.max_row, last_section_end + 1) + 1):
        account = worksheet_cell_text(template_sheet, row, 1)
        name = worksheet_cell_text(template_sheet, row, 2).upper()
        if account == "" or "DEVOL" not in name:
            continue
        if start_row is None:
            start_row = row
        end_row = row
    if start_row is None or end_row is None:
        return None
    return {
        "name": "devol",
        "start_row": start_row,
        "end_row": end_row,
        "amount_column": int(layout["debit_column"]),
        "opposite_column": int(layout["credit_column"]),
    }


def clear_row_values(worksheet: Worksheet, row_number: int, last_column: int) -> None:
    for column in range(1, last_column + 1):
        cell = worksheet.cell(row=row_number, column=column)
        if isinstance(cell, MergedCell):
            continue
        cell.value = None


def write_static_values(worksheet: Worksheet, row_number: int, values: dict[int, Any]) -> None:
    for column, value in values.items():
        cell = worksheet.cell(row=row_number, column=int(column))
        if isinstance(cell, MergedCell):
            continue
        cell.value = copy(value)


def write_date_cell(worksheet: Worksheet, row_number: int, column_number: int, date_value: datetime | None, date_text: str) -> None:
    cell = worksheet.cell(row=row_number, column=column_number)
    if isinstance(cell, MergedCell):
        return
    cell.value = date_value if isinstance(date_value, datetime) else (date_text or None)


def sum_column(worksheet: Worksheet, start_row: int, end_row: int, column_number: int) -> float:
    total = 0.0
    for row in range(start_row, end_row + 1):
        total += worksheet_cell_number(worksheet, row, column_number)
    return round_amount(total, 2)


def apply_section(
    output_sheet: Worksheet,
    template_sheet: Worksheet,
    template_values_sheet: Worksheet,
    layout: dict[str, Any],
    template_section: dict[str, Any],
    output_section: dict[str, Any],
    source_groups: dict[str, dict[str, dict[str, Any]]],
) -> dict[str, int]:
    last_column = get_section_last_column(layout)
    debit_groups = source_groups.get("debit", {})
    credit_groups = source_groups.get("credit", {})
    static_templates = get_section_static_templates(template_sheet, layout, template_section)
    opening_balances = get_section_opening_balances(template_values_sheet, layout, template_section)
    account_order = get_account_order(template_sheet, template_section)
    entries = build_posting_entries(debit_groups, "debit") + build_posting_entries(credit_groups, "credit")
    entries.sort(key=lambda entry: (account_order.get(entry["account"], 999999), compare_posting_key(entry)))
    capacity = output_section["end_row"] - output_section["start_row"] + 1
    added_rows = max(0, len(entries) - capacity)
    if added_rows > 0:
        insert_row = output_section["end_row"] + 1
        output_sheet.insert_rows(insert_row, amount=added_rows)
        style_source_row = max(output_section["start_row"], insert_row - 1)
        for offset in range(added_rows):
            row_number = insert_row + offset
            copy_row_style(output_sheet, output_sheet, style_source_row, row_number, last_column)
            clear_row_values(output_sheet, row_number, last_column)
        output_section = {
            **output_section,
            "end_row": output_section["end_row"] + added_rows,
        }

    for row in range(output_section["start_row"], output_section["end_row"] + 1):
        clear_row_values(output_sheet, row, last_column)

    current_account = ""
    running_balance = 0.0
    for index, entry in enumerate(entries):
        row_number = output_section["start_row"] + index
        static_values = {**get_my_account_static_fallback(layout, entry["account"]), **static_templates.get(entry["account"], {})}
        write_static_values(output_sheet, row_number, static_values)
        output_sheet.cell(row=row_number, column=1).value = entry["account"]
        write_date_cell(output_sheet, row_number, int(layout["date_column"]), entry["date_value"], entry["date_text"])
        output_sheet.cell(row=row_number, column=int(layout["seat_column"])).value = entry["seat"] or None
        output_sheet.cell(row=row_number, column=int(layout["detail_column"])).value = entry["detail"] or None
        output_sheet.cell(row=row_number, column=int(layout["debit_column"])).value = entry["amount"] if entry["side"] == "debit" else 0
        output_sheet.cell(row=row_number, column=int(layout["credit_column"])).value = entry["amount"] if entry["side"] == "credit" else 0
        if entry["account"] != current_account:
            current_account = entry["account"]
            running_balance = opening_balances.get(entry["account"], 0.0)
        debit = entry["amount"] if entry["side"] == "debit" else 0.0
        credit = entry["amount"] if entry["side"] == "credit" else 0.0
        running_balance = round_amount(running_balance + debit - credit, 2)
        output_sheet.cell(row=row_number, column=int(layout["saldo_column"])).value = running_balance

    total_row = output_section["end_row"] + 1
    clear_row_values(output_sheet, total_row, last_column)
    output_sheet.cell(row=total_row, column=int(template_section["amount_column"])).value = sum_column(
        output_sheet,
        output_section["start_row"],
        output_section["end_row"],
        int(template_section["amount_column"]),
    )
    output_sheet.cell(row=total_row, column=int(template_section["opposite_column"])).value = sum_column(
        output_sheet,
        output_section["start_row"],
        output_section["end_row"],
        int(template_section["opposite_column"]),
    )
    return {
        "rows": len(entries),
        "total_row": total_row,
        "added_rows": added_rows,
        "start_row": int(output_section["start_row"]),
        "end_row": int(output_section["end_row"]),
    }


def shift_section(section: dict[str, Any], row_offset: int) -> dict[str, Any]:
    shifted = dict(section)
    shifted["start_row"] = int(section["start_row"]) + row_offset
    shifted["end_row"] = int(section["end_row"]) + row_offset
    return shifted


def resolve_saved_input(saved_inputs: dict[str, Any], field: str) -> Path | None:
    payload = saved_inputs.get(field)
    candidate = payload.get("path") if isinstance(payload, dict) else payload
    if not candidate:
        return None
    path = Path(str(candidate)).resolve()
    return path if path.is_file() else None


def load_optional_nc_source(path: Path | None, label: str, optional: bool) -> Any | None:
    if path is None:
        return None
    try:
        return load_nc_source_workbook(path, label)
    except Exception:
        if optional:
            return None
        raise


def run(request: ProcessRequest) -> ProcessResult:
    template_path = request.template_path.resolve() if request.template_path else None
    if template_path is None or not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    saved_inputs = request.options.get("saved_inputs", {})
    if not isinstance(saved_inputs, dict):
        raise ValueError("saved_inputs must be a mapping.")

    workbook = load_workbook_quiet(template_path, data_only=False, keep_links=True)
    template_reference = load_workbook_quiet(template_path, data_only=False, keep_links=True)
    template_values = load_workbook_quiet(template_path, data_only=True, keep_links=True)

    summary: list[dict[str, Any]] = []
    metadata_sections: dict[str, Any] = {}
    for config in SOURCE_CONFIGS:
        key = str(config["key"])
        layout = MY_LAYOUTS[key]
        rep_path = resolve_saved_input(saved_inputs, str(config["rep_field"]))
        if rep_path is None:
            raise FileNotFoundError(f"Input not found for {config['rep_field']}")
        rep_source = load_rep_source_workbook(rep_path, str(config["label"]))
        nc_path = resolve_saved_input(saved_inputs, str(config["nc_field"]))
        nc_source = load_optional_nc_source(nc_path, str(config["label"]), bool(config["optional_nc"]))

        rep_groups = build_rep_posting_groups(rep_source, key)
        nc_groups = build_nc_groups_from_rows(build_nc_rows_from_source(nc_source, key), key)

        output_sheet = workbook[str(config["my_sheet"])]
        template_sheet = template_reference[str(config["my_sheet"])]
        template_values_sheet = template_values[str(config["my_sheet"])]
        row_shift = 0
        sales_template_section = dict(layout["sections"][0])
        sales_output_section = shift_section(sales_template_section, row_shift)

        sales_result = apply_section(
            output_sheet,
            template_sheet,
            template_values_sheet,
            layout,
            sales_template_section,
            sales_output_section,
            {"debit": {}, "credit": rep_groups["sales"]},
        )
        row_shift += sales_result["added_rows"]
        discount_template_section = dict(layout["sections"][1])
        discount_output_section = shift_section(discount_template_section, row_shift)
        discount_result = apply_section(
            output_sheet,
            template_sheet,
            template_values_sheet,
            layout,
            discount_template_section,
            discount_output_section,
            {"debit": rep_groups["discount"], "credit": nc_groups["discount_credit"]},
        )
        row_shift += discount_result["added_rows"]

        devol_result = None
        devol_section = get_devol_section_range(template_sheet, layout)
        if devol_section is not None:
            devol_output_section = shift_section(devol_section, row_shift)
            devol_result = apply_section(
                output_sheet,
                template_sheet,
                template_values_sheet,
                layout,
                devol_section,
                devol_output_section,
                {"debit": nc_groups["devol"], "credit": {}},
            )
            row_shift += devol_result["added_rows"]

        metadata_sections[key] = {
            "my_sheet": str(config["my_sheet"]),
            "sales_rows": sales_result["rows"],
            "discount_rows": discount_result["rows"],
            "devol_rows": 0 if devol_result is None else devol_result["rows"],
            "sales_total_row": sales_result["total_row"],
            "discount_total_row": discount_result["total_row"],
            "devol_total_row": None if devol_result is None else devol_result["total_row"],
        }
        summary.append(
            {
                "label": str(config["label"]),
                "sales_rows": sales_result["rows"],
                "discount_rows": discount_result["rows"],
                "devol_rows": 0 if devol_result is None else devol_result["rows"],
            }
        )

    output_path = request.output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_workbook_with_retries(workbook, output_path)
    return ProcessResult(
        success=True,
        output_path=output_path,
        label="repuestos_tytserv_my_stage",
        metadata={
            "runtime": "python-native-my-stage",
            "summary": summary,
            "sections": metadata_sections,
        },
    )
