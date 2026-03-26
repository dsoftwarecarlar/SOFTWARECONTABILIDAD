from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.views import Selection
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from ..cxp_actions.accion3_native import write_workbook_with_retries
from . import mayor_iva_stage, my_stage, nc_stage, rep_stage
from .workbook_tools import load_workbook_quiet

NC_MAX_COLUMN = 42
REP_MAX_COLUMN = 41
NC_SHEET_MAP = {
    "tyt": "NC REP TYT",
    "peug": "NC REP PEUG",
    "chgn": "NC REP CHGN",
    "szk": "NC REP SZK",
}
REP_MAYOR_SPECS = {
    "tyt": {"sheet": "REP TYT", "my_sheet": "MY REP TYT", "sales_column": 9, "discount_column": 8, "discount_target_column": 18},
    "peug": {"sheet": "REP PEUGT", "my_sheet": "MY REP PEUG", "sales_column": 9, "discount_column": 8, "discount_target_column": 18},
    "chgn": {"sheet": "REP CHGN", "my_sheet": "MY REP CHGN", "sales_column": 10, "discount_column": 9, "discount_target_column": 19},
    "szk": {"sheet": "REP SZK", "my_sheet": "MY REP SZK", "sales_column": 9, "discount_column": 8, "discount_target_column": 18},
}
OPEN_VIEW_SPECS = {
    "REP TYT": {"active_cell": "A11", "top_left_cell": "A9", "tab_selected": True},
    "REP PEUGT": {"active_cell": "A11", "top_left_cell": "A9"},
    "REP CHGN": {"active_cell": "A11", "top_left_cell": "A9"},
    "REP SZK": {"active_cell": "A11", "top_left_cell": "A9"},
    "NC REP TYT": {"active_cell": "A8", "top_left_cell": "A5"},
    "NC REP PEUG": {"active_cell": "A8", "top_left_cell": "A5"},
    "NC REP SZK": {"active_cell": "A8", "top_left_cell": "A5"},
    "MY REP TYT": {"active_cell": "A1", "top_left_cell": "A1"},
    "MY REP PEUG": {"active_cell": "A1", "top_left_cell": "A1"},
    "MY REP CHGN": {"active_cell": "A1", "top_left_cell": "A1"},
    "MY REP SZK": {"active_cell": "A1", "top_left_cell": "A1"},
    "MAYOR IVA": {"active_cell": "A299", "top_left_cell": "A295"},
}


def run(request: ProcessRequest) -> ProcessResult:
    template_path = request.template_path.resolve() if request.template_path else None
    if template_path is None or not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    saved_inputs = request.options.get("saved_inputs", {})
    if not isinstance(saved_inputs, dict):
        raise ValueError("saved_inputs must be a mapping.")

    output_path = request.output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    staging_root = output_path.parent / "__staging_repuestos_tytserv"
    staging_root.mkdir(parents=True, exist_ok=True)
    temp_root = staging_root / f"run_{uuid4().hex}"
    temp_root.mkdir(parents=True, exist_ok=True)

    try:
        rep_result = run_stage(rep_stage.run, request, template_path, temp_root / "rep_stage.xlsx")
        nc_result = run_stage(nc_stage.run, request, rep_result.output_path, temp_root / "nc_stage.xlsx")
        my_result = run_stage(my_stage.run, request, nc_result.output_path, temp_root / "my_stage.xlsx")
        mayor_result = run_stage(mayor_iva_stage.run, request, my_result.output_path, temp_root / "mayor_iva_stage.xlsx")

        workbook = load_workbook_quiet(mayor_result.output_path, data_only=False, keep_links=True)
        template_reference = load_workbook_quiet(template_path, data_only=False, keep_links=True)
        my_sections = my_result.metadata.get("sections")
        update_nc_mayor_rows(workbook, template_reference, my_sections if isinstance(my_sections, dict) else None)
        update_rep_mayor_rows(workbook, my_sections if isinstance(my_sections, dict) else None)
        apply_open_presentation(workbook)
        final_output = write_workbook_with_retries(workbook, output_path)
        try:
            integrity_checks = validate_final_output(final_output, rep_result, nc_result, my_result, mayor_result)
        except Exception:
            final_output.unlink(missing_ok=True)
            raise
    finally:
        shutil.rmtree(temp_root, ignore_errors=True)

    return ProcessResult(
        success=True,
        output_path=final_output,
        label="repuestos_tytserv",
        metadata={
            "runtime": "python-native",
            "summary": build_summary(rep_result),
            "console": build_console(rep_result, nc_result, my_result, mayor_result),
            "output_origin": "default_path",
            "fallback_used": False,
            "stages": {
                "rep": rep_result.metadata,
                "nc": nc_result.metadata,
                "my": my_result.metadata,
                "mayor_iva": mayor_result.metadata,
            },
            "integrity_checks": integrity_checks,
        },
    )


def run_stage(
    processor: Any,
    request: ProcessRequest,
    template_path: Path,
    output_path: Path,
) -> ProcessResult:
    stage_request = ProcessRequest(
        input_paths=request.input_paths,
        output_path=output_path.resolve(),
        template_path=template_path.resolve(),
        options=request.options,
    )
    return processor(stage_request)


def build_summary(rep_result: ProcessResult) -> list[dict[str, int | str]]:
    summary = rep_result.metadata.get("summary", [])
    return summary if isinstance(summary, list) else []


def build_console(*results: ProcessResult) -> str:
    lines: list[str] = []
    for result in results:
        runtime = str(result.metadata.get("runtime", result.label)).strip()
        label = result.label.strip()
        if runtime != "":
            lines.append(f"INFO|runtime|{label}={runtime}")
    lines.append("INFO|mayor_iva|updated=1")
    return "\n".join(lines)


def worksheet_value(worksheet: Worksheet, row: int, column: int) -> Any:
    cell = worksheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return None
    return cell.value


def worksheet_text(worksheet: Worksheet, row: int, column: int) -> str:
    value = worksheet_value(worksheet, row, column)
    return string_value(value)


def worksheet_number(worksheet: Worksheet, row: int, column: int) -> float:
    cell = worksheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return 0.0
    value = cell.value
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text == "":
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def round_amount(value: Any, decimals: int = 2) -> float:
    return round(float(value or 0) + 1e-12, decimals)


def string_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = string_value(value).replace(",", "")
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def values_match(expected: Any, actual: Any, *, force_text: bool = False) -> bool:
    if force_text:
        return string_value(expected) == string_value(actual)

    expected_number = parse_number(expected)
    actual_number = parse_number(actual)
    if expected_number is not None and actual_number is not None:
        tolerance = max(1e-9, max(abs(expected_number), abs(actual_number)) * 1e-12)
        return abs(expected_number - actual_number) <= tolerance

    return string_value(expected) == string_value(actual)


def find_row_containing(
    worksheet: Worksheet,
    needle: str,
    start_row: int = 1,
    end_row: int | None = None,
    last_column: int = REP_MAX_COLUMN,
) -> int | None:
    expected = str(needle).strip().upper()
    max_row = end_row if end_row is not None else max(worksheet.max_row, start_row)
    for row in range(start_row, max_row + 1):
        for column in range(1, last_column + 1):
            if expected in worksheet_text(worksheet, row, column).upper():
                return row
    return None


def apply_open_presentation(workbook: Any) -> None:
    if "REP TYT" in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index("REP TYT")

    for worksheet in workbook.worksheets:
        sheet_name = worksheet.title
        spec = OPEN_VIEW_SPECS.get(sheet_name, {})
        active_cell = str(spec.get("active_cell", "A1"))
        top_left_cell = str(spec.get("top_left_cell", "A1"))
        worksheet.sheet_view.tabSelected = bool(spec.get("tab_selected", False))
        worksheet.sheet_view.topLeftCell = top_left_cell
        if worksheet.sheet_view.selection:
            selection = worksheet.sheet_view.selection[0]
            selection.activeCell = active_cell
            selection.sqref = active_cell
        else:
            worksheet.sheet_view.selection = [Selection(activeCell=active_cell, sqref=active_cell)]


def first_mismatch(
    expected_sheet: Worksheet,
    actual_sheet: Worksheet,
    *,
    end_row: int,
    last_column: int,
    text_columns: set[int] | None = None,
) -> tuple[int, int, str, str] | None:
    forced_text = text_columns or set()
    for row in range(1, end_row + 1):
        for column in range(1, last_column + 1):
            expected_value = worksheet_value(expected_sheet, row, column)
            actual_value = worksheet_value(actual_sheet, row, column)
            if values_match(expected_value, actual_value, force_text=column in forced_text):
                continue
            return row, column, string_value(expected_value), string_value(actual_value)
    return None


def merge_signature(
    worksheet: Worksheet,
    *,
    start_row: int = 1,
    end_row: int | None = None,
    last_column: int | None = None,
) -> tuple[tuple[int, int, int, int], ...]:
    max_row = end_row if end_row is not None else worksheet.max_row
    max_column = last_column if last_column is not None else worksheet.max_column
    signature: list[tuple[int, int, int, int]] = []
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.max_row < start_row or merged_range.min_row > max_row:
            continue
        if merged_range.max_col < 1 or merged_range.min_col > max_column:
            continue
        signature.append(
            (
                int(merged_range.min_row),
                int(merged_range.max_row),
                int(merged_range.min_col),
                int(merged_range.max_col),
            )
        )
    signature.sort()
    return tuple(signature)


def validate_final_output(
    final_output: Path,
    rep_result: ProcessResult,
    nc_result: ProcessResult,
    my_result: ProcessResult,
    mayor_result: ProcessResult,
) -> dict[str, str]:
    final_workbook = load_workbook_quiet(final_output, data_only=False, keep_links=True)
    rep_workbook = load_workbook_quiet(rep_result.output_path, data_only=False, keep_links=True)
    nc_workbook = load_workbook_quiet(nc_result.output_path, data_only=False, keep_links=True)
    my_workbook = load_workbook_quiet(my_result.output_path, data_only=False, keep_links=True)
    mayor_workbook = load_workbook_quiet(mayor_result.output_path, data_only=False, keep_links=True)

    rep_metadata = rep_result.metadata.get("rep_stage", {})
    if not isinstance(rep_metadata, dict):
        raise RuntimeError("La etapa REP no devolvio metadata verificable.")

    for config in rep_stage.SHEET_CONFIGS:
        key = str(config["key"])
        sheet_name = str(config["target_sheet"])
        expected_sheet = rep_workbook[sheet_name]
        actual_sheet = final_workbook[sheet_name]
        sheet_meta = rep_metadata.get(key, {})
        if not isinstance(sheet_meta, dict):
            raise RuntimeError(f"La etapa REP no devolvio metadata para {sheet_name}.")
        end_row = int(sheet_meta.get("source_total_row", 0))
        mismatch = first_mismatch(
            expected_sheet,
            actual_sheet,
            end_row=end_row,
            last_column=REP_MAX_COLUMN,
            text_columns=set(rep_stage.REP_TEXT_COLUMNS),
        )
        if mismatch is not None:
            row, column, expected_value, actual_value = mismatch
            raise RuntimeError(
                f"La salida final altero {sheet_name} en fila {row} columna {column}. "
                f"Esperado='{expected_value}' actual='{actual_value}'."
            )
        if merge_signature(expected_sheet, last_column=REP_MAX_COLUMN) != merge_signature(actual_sheet, last_column=REP_MAX_COLUMN):
            raise RuntimeError(f"La salida final altero la estructura fusionada de {sheet_name}.")

    nc_metadata = nc_result.metadata.get("nc_stage", {})
    if not isinstance(nc_metadata, dict):
        raise RuntimeError("La etapa NC no devolvio metadata verificable.")

    for config in nc_stage.SHEET_CONFIGS:
        key = str(config["key"])
        sheet_name = str(config["target_sheet"])
        expected_sheet = nc_workbook[sheet_name]
        actual_sheet = final_workbook[sheet_name]
        sheet_meta = nc_metadata.get(key, {})
        if not isinstance(sheet_meta, dict):
            raise RuntimeError(f"La etapa NC no devolvio metadata para {sheet_name}.")
        end_row = int(sheet_meta.get("source_total_row", 0))
        mismatch = first_mismatch(
            expected_sheet,
            actual_sheet,
            end_row=end_row,
            last_column=NC_MAX_COLUMN,
        )
        if mismatch is not None:
            row, column, expected_value, actual_value = mismatch
            raise RuntimeError(
                f"La salida final altero {sheet_name} en fila {row} columna {column}. "
                f"Esperado='{expected_value}' actual='{actual_value}'."
            )
        if merge_signature(expected_sheet, last_column=NC_MAX_COLUMN) != merge_signature(actual_sheet, last_column=NC_MAX_COLUMN):
            raise RuntimeError(f"La salida final altero la estructura fusionada de {sheet_name}.")

    for config in my_stage.SOURCE_CONFIGS:
        sheet_name = str(config["my_sheet"])
        expected_sheet = my_workbook[sheet_name]
        actual_sheet = final_workbook[sheet_name]
        mismatch = first_mismatch(
            expected_sheet,
            actual_sheet,
            end_row=max(expected_sheet.max_row, actual_sheet.max_row),
            last_column=max(expected_sheet.max_column, actual_sheet.max_column),
        )
        if mismatch is not None:
            row, column, expected_value, actual_value = mismatch
            raise RuntimeError(
                f"La salida final altero {sheet_name} en fila {row} columna {column}. "
                f"Esperado='{expected_value}' actual='{actual_value}'."
            )

    expected_mayor_sheet = mayor_workbook["MAYOR IVA"]
    actual_mayor_sheet = final_workbook["MAYOR IVA"]
    mayor_mismatch = first_mismatch(
        expected_mayor_sheet,
        actual_mayor_sheet,
        end_row=max(expected_mayor_sheet.max_row, actual_mayor_sheet.max_row),
        last_column=max(expected_mayor_sheet.max_column, actual_mayor_sheet.max_column),
    )
    if mayor_mismatch is not None:
        row, column, expected_value, actual_value = mayor_mismatch
        raise RuntimeError(
            f"La salida final altero MAYOR IVA en fila {row} columna {column}. "
            f"Esperado='{expected_value}' actual='{actual_value}'."
        )

    for config in rep_stage.SHEET_CONFIGS:
        key = str(config["key"])
        sheet_name = str(config["target_sheet"])
        sheet_meta = rep_metadata.get(key, {})
        if not isinstance(sheet_meta, dict):
            continue
        target_sheet = final_workbook[sheet_name]
        source_total_row = int(sheet_meta.get("source_total_row", 0))
        for row in range(rep_stage.REP_DETAIL_START_ROW, source_total_row):
            document = worksheet_text(target_sheet, row, 5)
            if document == "":
                continue
            height = target_sheet.row_dimensions[row].height
            if height is not None and height < 10:
                raise RuntimeError(
                    f"La salida final dejo {sheet_name} con fila visualmente colapsada en {row}. "
                    f"Altura actual={height}."
                )

    return {
        "rep": "ok",
        "nc": "ok",
        "my": "ok",
        "mayor_iva": "ok",
        "presentation": "ok",
    }


def my_total_rows(template_reference: Any, my_sections: dict[str, Any] | None = None) -> dict[str, dict[str, int]]:
    totals: dict[str, dict[str, int]] = {}
    for config in my_stage.SOURCE_CONFIGS:
        key = str(config["key"])
        layout = my_stage.MY_LAYOUTS[key]
        template_sheet = template_reference[str(config["my_sheet"])]
        section_meta = my_sections.get(key, {}) if isinstance(my_sections, dict) else {}
        rows = {
            "sales": int(section_meta.get("sales_total_row", int(layout["sections"][0]["end_row"]) + 1)),
            "discount": int(section_meta.get("discount_total_row", int(layout["sections"][1]["end_row"]) + 1)),
        }
        devol_total_row = section_meta.get("devol_total_row") if isinstance(section_meta, dict) else None
        if devol_total_row is not None:
            rows["devol"] = int(devol_total_row)
        else:
            devol_section = my_stage.get_devol_section_range(template_sheet, layout)
            if devol_section is not None:
                rows["devol"] = int(devol_section["end_row"]) + 1
        totals[key] = rows
    return totals


def update_nc_mayor_rows(workbook: Any, template_reference: Any, my_sections: dict[str, Any] | None = None) -> None:
    total_rows_by_key = my_total_rows(template_reference, my_sections)
    for config in my_stage.SOURCE_CONFIGS:
        key = str(config["key"])
        sheet_name = NC_SHEET_MAP.get(key, "")
        if sheet_name == "" or sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        mayor_row = find_row_containing(
            worksheet,
            "MAYOR",
            1,
            max(200, worksheet.max_row + 20),
            NC_MAX_COLUMN,
        )
        if mayor_row is None:
            continue

        layout = my_stage.MY_LAYOUTS[key]
        my_sheet_name = str(layout["my_sheet_name"])
        if my_sheet_name not in workbook.sheetnames:
            continue

        my_worksheet = workbook[my_sheet_name]
        totals = total_rows_by_key.get(key, {})
        devol_row = int(totals.get("devol", 0))
        discount_row = int(totals.get("discount", 0))
        devol_total = worksheet_number(my_worksheet, devol_row, int(layout["debit_column"])) if devol_row > 0 else 0.0
        discount_total = worksheet_number(my_worksheet, discount_row, int(layout["credit_column"])) if discount_row > 0 else 0.0

        worksheet.cell(row=mayor_row, column=22).value = "MAYOR"
        worksheet.cell(row=mayor_row, column=23).value = round_amount(devol_total, 2)
        worksheet.cell(row=mayor_row, column=24).value = round_amount(discount_total, 2)


def update_rep_mayor_rows(workbook: Any, my_sections: dict[str, Any] | None = None) -> None:
    for key, spec in REP_MAYOR_SPECS.items():
        sheet_name = str(spec["sheet"])
        my_sheet_name = str(spec["my_sheet"])
        if sheet_name not in workbook.sheetnames or my_sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        my_worksheet = workbook[my_sheet_name]
        layout = my_stage.MY_LAYOUTS[key]
        section_meta = my_sections.get(key, {}) if isinstance(my_sections, dict) else {}
        total_row = find_row_containing(worksheet, "TOTAL GENERAL", 1, max(200, worksheet.max_row + 20), REP_MAX_COLUMN)
        mayor_row = find_row_containing(worksheet, "MAYOR", 1, max(200, worksheet.max_row + 20), REP_MAX_COLUMN)
        if total_row is None or mayor_row is None:
            continue

        sales_total_row = int(section_meta.get("sales_total_row", int(layout["sections"][0]["end_row"]) + 1))
        discount_total_row = int(section_meta.get("discount_total_row", int(layout["sections"][1]["end_row"]) + 1))
        worksheet.cell(row=mayor_row, column=16).value = round_amount(
            worksheet_number(my_worksheet, sales_total_row, int(spec["sales_column"])),
            2,
        )
        worksheet.cell(row=mayor_row, column=int(spec["discount_target_column"])).value = round_amount(
            worksheet_number(my_worksheet, discount_total_row, int(spec["discount_column"])),
            2,
        )

        if key == "peug" and "NC REP PEUG" in workbook.sheetnames:
            nc_sheet = workbook["NC REP PEUG"]
            nc_total_row = find_row_containing(
                nc_sheet,
                "TOTAL GENERAL",
                1,
                max(200, nc_sheet.max_row + 20),
                NC_MAX_COLUMN,
            )
            nc_vat = worksheet_number(nc_sheet, nc_total_row, 30) if nc_total_row is not None else 0.0
            total_base = worksheet_number(worksheet, total_row, 24)
            worksheet.cell(row=mayor_row, column=25).value = round_amount(total_base - nc_vat, 2)
            worksheet.cell(row=mayor_row, column=26).value = round_amount(total_base, 2)
