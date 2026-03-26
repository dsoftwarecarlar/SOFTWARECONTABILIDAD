from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from ..cxp_actions.accion3_native import write_workbook_with_retries
from .workbook_tools import (
    apply_row_merge_definitions,
    clear_merges_in_row_range,
    clear_rows,
    copy_row_dimensions,
    copy_row_style,
    get_row_merge_definitions,
    load_workbook_quiet,
    prepare_marker_row,
    write_literal_string,
)

REP_MAX_COLUMN = 41
REP_DETAIL_START_ROW = 11
REP_PAYLOAD_COLUMNS = tuple(range(1, REP_MAX_COLUMN + 1))
REP_TEXT_COLUMNS = {
    1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16,
    37, 38, 39, 40, 41,
}
INVALID_XML_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

SHEET_CONFIGS = (
    {"key": "tyt", "field": "excel_tyt", "label": "MATRIZ", "target_sheet": "REP TYT"},
    {"key": "peug", "field": "excel_peug", "label": "PEUGEOT", "target_sheet": "REP PEUGT"},
    {"key": "chgn", "field": "excel_chgn", "label": "CHANGAN", "target_sheet": "REP CHGN"},
    {"key": "szk", "field": "excel_szk", "label": "SUZUKI", "target_sheet": "REP SZK"},
)

SOURCE_SHEET_NAME = "RepLibroVentasGeneral"
HEADER_CHECKS = (
    (9, 5, "# DOC"),
    (9, 9, "RUC"),
    (9, 10, "CLIENTE"),
    (9, 11, "CLIENTE"),
    (9, 16, "ITEM"),
    (9, 18, "SUBTOT"),
)


@dataclass(frozen=True)
class SourceWorkbook:
    path: Path
    label: str
    worksheet: Worksheet
    total_row: int
    last_row: int
    payload_signature: tuple[int, str]


def literal_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    return INVALID_XML_CONTROL_CHARS.sub(" ", str(value)).strip()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", literal_text(value)).strip()


def worksheet_text(worksheet: Worksheet, row: int, column: int) -> str:
    cell = worksheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return ""

    return literal_text(cell.value)


def find_row_containing(
    worksheet: Worksheet,
    needle: str,
    start_row: int = 1,
    end_row: int | None = None,
    last_column: int = REP_MAX_COLUMN,
) -> int | None:
    expected = normalize_text(needle).upper()
    max_row = end_row if end_row is not None else max(worksheet.max_row, start_row)
    for row in range(start_row, max_row + 1):
        for column in range(1, last_column + 1):
            if expected in worksheet_text(worksheet, row, column).upper():
                return row

    return None


def source_cell_value(cell: Any, column: int) -> Any:
    if isinstance(cell, MergedCell) or cell is None:
        return None

    value = cell.value
    if value is None:
        return None

    if column in REP_TEXT_COLUMNS:
        text = literal_text(value)
        return text if text != "" else None

    return value


def build_payload_signature(worksheet: Worksheet, total_row: int) -> tuple[int, str]:
    import hashlib

    rows: list[str] = []
    for row in range(REP_DETAIL_START_ROW, total_row):
        document = worksheet_text(worksheet, row, 5)
        if document == "" or document.upper().startswith("ANULAD"):
            continue

        values = [worksheet_text(worksheet, row, column) for column in REP_PAYLOAD_COLUMNS]
        rows.append("|".join(values))

    digest = hashlib.sha256("\n".join(rows).encode("utf-8")).hexdigest()
    return len(rows), digest


def row_has_rep_payload_content(worksheet: Worksheet, row_number: int) -> bool:
    for column in REP_PAYLOAD_COLUMNS:
        if worksheet_text(worksheet, row_number, int(column)) != "":
            return True
    return False


def find_rep_footer_start_row(worksheet: Worksheet, total_row: int) -> int:
    if total_row <= REP_DETAIL_START_ROW:
        return total_row

    footer_start_row = total_row
    for row in range(total_row - 1, REP_DETAIL_START_ROW - 1, -1):
        if row_has_rep_payload_content(worksheet, row):
            break
        footer_start_row = row

    return footer_start_row


def find_rep_style_row(worksheet: Worksheet, footer_start_row: int) -> int:
    for row in range(footer_start_row - 1, REP_DETAIL_START_ROW - 1, -1):
        if row_has_rep_payload_content(worksheet, row):
            return row
    return REP_DETAIL_START_ROW


def find_last_populated_source_detail_row(source: SourceWorkbook) -> int:
    for row in range(source.total_row - 1, REP_DETAIL_START_ROW - 1, -1):
        if row_has_rep_payload_content(source.worksheet, row):
            return row
    return REP_DETAIL_START_ROW - 1


def assert_source_headers(worksheet: Worksheet, label: str) -> None:
    for row, column, needle in HEADER_CHECKS:
        actual = worksheet_text(worksheet, row, column).upper()
        expected = normalize_text(needle).upper()
        if expected not in actual:
            raise RuntimeError(
                f"La hoja fuente {label} no coincide con la estructura esperada en fila {row} columna {column}. "
                f"Esperado contiene '{needle}' y llego '{actual}'."
            )


def load_source_workbook(path: Path, label: str) -> SourceWorkbook:
    workbook = load_workbook_quiet(path, data_only=False, keep_links=True)
    if SOURCE_SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(
            f"El archivo fuente {label} debe contener la hoja '{SOURCE_SHEET_NAME}'. "
            "No se aceptan plantillas ni salidas ya generadas."
        )

    worksheet = workbook[SOURCE_SHEET_NAME]
    assert_source_headers(worksheet, label)
    total_row = find_row_containing(worksheet, "TOTAL GENERAL", 1, max(worksheet.max_row, 200), REP_MAX_COLUMN)
    if total_row is None:
        raise RuntimeError(f"No se encontro TOTAL GENERAL en la fuente {label}.")

    return SourceWorkbook(
        path=path,
        label=label,
        worksheet=worksheet,
        total_row=total_row,
        last_row=worksheet.max_row,
        payload_signature=build_payload_signature(worksheet, total_row),
    )


def ensure_rep_sheet_capacity(
    target: Worksheet,
    template_total_row: int,
    template_mayor_row: int,
    required_last_detail_row: int,
) -> dict[str, int]:
    if required_last_detail_row < REP_DETAIL_START_ROW:
        return {
            "total_row": template_total_row,
            "mayor_row": template_mayor_row,
            "format_mayor_row": template_mayor_row,
            "template_total_row": template_total_row,
        }

    footer_start_row = find_rep_footer_start_row(target, template_total_row)
    spacer_count = max(0, template_total_row - footer_start_row)
    desired_total_row = max(
        REP_DETAIL_START_ROW + spacer_count + 1,
        required_last_detail_row + spacer_count + 1,
    )
    extra_rows = max(0, desired_total_row - template_total_row)
    style_row = find_rep_style_row(target, footer_start_row)
    detail_merges = get_row_merge_definitions(target, style_row, REP_MAX_COLUMN)
    spacer_merges = get_row_merge_definitions(target, footer_start_row, REP_MAX_COLUMN)
    total_merges = get_row_merge_definitions(target, template_total_row, REP_MAX_COLUMN)

    if extra_rows > 0:
        target.insert_rows(footer_start_row, amount=extra_rows)

    template_spacer_row = footer_start_row + extra_rows
    shifted_template_total_row = template_total_row + extra_rows
    format_mayor_row = template_mayor_row + extra_rows if template_mayor_row > 0 else template_mayor_row
    desired_spacer_row = desired_total_row - spacer_count

    clear_merges_in_row_range(
        target,
        min(desired_spacer_row, template_spacer_row),
        max(shifted_template_total_row, format_mayor_row or shifted_template_total_row),
        REP_MAX_COLUMN,
    )

    for row in range(footer_start_row, footer_start_row + extra_rows):
        copy_row_style(target, target, style_row, row, REP_MAX_COLUMN)
        apply_row_merge_definitions(target, row, detail_merges)

    copy_row_style(target, target, template_spacer_row, desired_spacer_row, REP_MAX_COLUMN)
    apply_row_merge_definitions(target, desired_spacer_row, spacer_merges)

    copy_row_style(target, target, shifted_template_total_row, desired_total_row, REP_MAX_COLUMN)
    apply_row_merge_definitions(target, desired_total_row, total_merges)

    return {
        "total_row": desired_total_row,
        "mayor_row": desired_total_row + 1,
        "format_mayor_row": format_mayor_row,
        "template_total_row": shifted_template_total_row,
    }


def copy_source_to_rep_sheet(source: SourceWorkbook, target: Worksheet) -> tuple[int, int, int]:
    template_total_row = find_row_containing(
        target,
        "TOTAL GENERAL",
        1,
        max(target.max_row, 200),
        REP_MAX_COLUMN,
    )
    template_mayor_row = find_row_containing(
        target,
        "MAYOR",
        1,
        max(target.max_row, 200),
        REP_MAX_COLUMN,
    )

    if template_total_row is None or template_mayor_row is None:
        raise RuntimeError(f"La hoja {target.title} no tiene filas base TOTAL GENERAL/MAYOR.")

    required_last_detail_row = find_last_populated_source_detail_row(source)
    capacity = ensure_rep_sheet_capacity(
        target,
        template_total_row,
        template_mayor_row,
        required_last_detail_row,
    )
    shifted_total_row = int(capacity["total_row"])
    shifted_mayor_row = int(capacity["mayor_row"])
    format_mayor_row = int(capacity["format_mayor_row"])
    old_last_row = max(target.max_row, shifted_mayor_row, 200)

    for row in range(1, source.total_row + 1):
        copy_row_dimensions(source.worksheet, target, row, row)

        for column in range(1, REP_MAX_COLUMN + 1):
            cell = target.cell(row=row, column=column)
            if isinstance(cell, MergedCell):
                continue

            source_cell = source.worksheet.cell(row=row, column=column)
            value = source_cell_value(source_cell, column)
            if column in REP_TEXT_COLUMNS:
                if isinstance(value, str):
                    write_literal_string(cell, value)
                else:
                    cell.value = value
            else:
                cell.value = value

    actual_mayor_row = prepare_marker_row(
        target,
        format_mayor_row,
        source.total_row + 1,
        REP_MAX_COLUMN,
        14,
    )
    clear_rows(
        target,
        actual_mayor_row + 1,
        max(old_last_row, source.last_row + 10),
        REP_MAX_COLUMN,
    )

    return template_total_row, template_mayor_row, actual_mayor_row


def run(request: ProcessRequest) -> ProcessResult:
    template_path = request.template_path.resolve() if request.template_path else None
    if template_path is None or not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    saved_inputs = request.options.get("saved_inputs", {})
    if not isinstance(saved_inputs, dict):
        raise ValueError("saved_inputs must be a mapping.")

    workbook = load_workbook_quiet(template_path, data_only=False, keep_links=True)
    summary: list[dict[str, int | str]] = []
    rep_metadata: dict[str, Any] = {}

    for config in SHEET_CONFIGS:
        input_payload = saved_inputs.get(config["field"])
        if not isinstance(input_payload, dict):
            raise RuntimeError(f"No se encontro input guardado para {config['field']}.")

        source_path = Path(str(input_payload.get("path", ""))).resolve()
        if not source_path.is_file():
            raise FileNotFoundError(f"Input not found for {config['field']}: {source_path}")

        source = load_source_workbook(source_path, str(config["label"]))
        if config["target_sheet"] not in workbook.sheetnames:
            raise RuntimeError(f"No existe la hoja requerida en plantilla: {config['target_sheet']}")

        target = workbook[str(config["target_sheet"])]
        total_row, mayor_row, actual_mayor_row = copy_source_to_rep_sheet(source, target)
        target_signature = build_payload_signature(target, source.total_row)
        if target_signature != source.payload_signature:
            raise RuntimeError(
                f"La hoja {target.title} no conserva los datos del archivo subido. "
                f"Filas fuente={source.payload_signature[0]}, filas salida={target_signature[0]}."
            )

        summary.append({"label": str(config["label"]), "rows": source.payload_signature[0]})
        rep_metadata[str(config["key"])] = {
            "target_sheet": str(config["target_sheet"]),
            "rows": source.payload_signature[0],
            "source_total_row": source.total_row,
            "template_total_row": total_row,
            "template_mayor_row": mayor_row,
            "actual_mayor_row": actual_mayor_row,
            "payload_hash": source.payload_signature[1],
        }

    output_path = write_workbook_with_retries(workbook, request.output_path.resolve())
    return ProcessResult(
        success=True,
        output_path=output_path,
        label="repuestos_tytserv_rep_stage",
        metadata={
            "runtime": "python-native-rep-stage",
            "summary": summary,
            "rep_stage": rep_metadata,
            "console": "Etapa REP generada desde Python nativo.",
            "output_origin": "default_path",
            "fallback_used": False,
        },
    )
