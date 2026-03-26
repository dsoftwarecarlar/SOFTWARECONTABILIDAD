from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from ..cxp_actions.accion3_native import write_workbook_with_retries
from .workbook_tools import (
    clear_rows,
    clear_row_values,
    copy_row_style,
    find_style_source_row,
    load_workbook_quiet,
    prepare_marker_row,
    unmerge_ranges_in_band,
    write_literal_string,
)

NC_MAX_COLUMN = 42
NC_DETAIL_START_ROW = 8
INVALID_XML_CONTROL_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
SOURCE_SHEET_NAME = "RepLibroDevolucionesGeneral"

SHEET_CONFIGS = (
    {"key": "tyt", "field": "excel_nc_tyt", "label": "MATRIZ", "target_sheet": "NC REP TYT"},
    {"key": "peug", "field": "excel_nc_peug", "label": "PEUGEOT", "target_sheet": "NC REP PEUG"},
    {"key": "szk", "field": "excel_nc_szk", "label": "SUZUKI", "target_sheet": "NC REP SZK"},
)

HEADER_CHECKS = (
    (7, 9, "N/C No."),
    (7, 10, "DEV. A FAC."),
    (7, 11, "DESCRIPCION"),
    (7, 14, "FECHA FACT"),
    (7, 23, "SUBTOT"),
    (7, 36, "TOT. NC"),
    (7, 41, "ASIENTO"),
)


@dataclass(frozen=True)
class NcSourceWorkbook:
    path: Path
    label: str
    worksheet: Worksheet | None
    total_row: int
    detail_signature: tuple[int, str]
    is_empty: bool = False


def literal_text(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    return INVALID_XML_CONTROL_CHARS.sub(" ", str(value)).strip()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", literal_text(value)).strip()


def worksheet_text(worksheet: Worksheet, row: int, column: int) -> str:
    cell = worksheet.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return ""

    return literal_text(cell.value)


def find_row_containing(
    worksheet: Worksheet,
    needle: str,
    start_row: int = 1,
    end_row: int | None = None,
    last_column: int = NC_MAX_COLUMN,
) -> int | None:
    expected = normalize_text(needle).upper()
    max_row = end_row if end_row is not None else max(worksheet.max_row, start_row)
    for row in range(start_row, max_row + 1):
        for column in range(1, last_column + 1):
            if expected in worksheet_text(worksheet, row, column).upper():
                return row

    return None


def assert_source_headers(worksheet: Worksheet, label: str) -> None:
    for row, column, needle in HEADER_CHECKS:
        actual = worksheet_text(worksheet, row, column).upper()
        expected = normalize_text(needle).upper()
        if expected not in actual:
            raise RuntimeError(
                f"La hoja fuente {label} no coincide con la estructura esperada en fila {row} columna {column}. "
                f"Esperado contiene '{needle}' y llego '{actual}'."
            )


def source_cell_value(cell: Any) -> Any:
    if isinstance(cell, MergedCell) or cell is None:
        return None

    value = cell.value
    if value is None:
        return None

    if isinstance(value, str):
        text = literal_text(value)
        return text if text != "" else None

    return value


def build_detail_signature(worksheet: Worksheet, total_row: int) -> tuple[int, str]:
    import hashlib

    rows: list[str] = []
    for row in range(NC_DETAIL_START_ROW, total_row + 1):
        values = [worksheet_text(worksheet, row, column) for column in range(1, NC_MAX_COLUMN + 1)]
        if all(value == "" for value in values):
            continue
        rows.append("|".join(values))

    digest = hashlib.sha256("\n".join(rows).encode("utf-8")).hexdigest()
    return len(rows), digest


def empty_detail_signature() -> tuple[int, str]:
    import hashlib

    return 0, hashlib.sha256(b"").hexdigest()


def apply_empty_source_to_nc_sheet(
    target: Worksheet,
    template_total_row: int,
    template_mayor_row: int,
    detail_style_row: int,
    template_values_sheet: Worksheet,
) -> tuple[int, int, int]:
    old_last_row = max(target.max_row, template_mayor_row, 200)
    affected_rows = unmerge_ranges_in_band(
        target,
        NC_DETAIL_START_ROW,
        template_mayor_row,
        NC_MAX_COLUMN,
    )

    for row in range(NC_DETAIL_START_ROW, template_total_row):
        if row != detail_style_row or row in affected_rows:
            copy_row_style(target, target, detail_style_row, row, NC_MAX_COLUMN)
        clear_row_values(target, row, NC_MAX_COLUMN)

    for column in range(1, NC_MAX_COLUMN + 1):
        cell = target.cell(row=template_total_row, column=column)
        if isinstance(cell, MergedCell):
            continue

        template_value = template_values_sheet.cell(row=template_total_row, column=column).value
        live_value = cell.value
        text = normalize_text(live_value)
        if "TOTAL GENERAL" in text.upper():
            write_literal_string(cell, literal_text(live_value or "TOTAL GENERAL:"))
            continue

        if isinstance(template_value, (int, float)) or isinstance(live_value, (int, float)):
            cell.value = 0
            continue

        cell.value = None

    clear_rows(
        target,
        template_mayor_row + 1,
        max(old_last_row, template_total_row + 10),
        NC_MAX_COLUMN,
    )
    return template_total_row, template_mayor_row, template_mayor_row


def load_source_workbook(path: Path, label: str) -> NcSourceWorkbook:
    workbook = load_workbook_quiet(path, data_only=False, keep_links=True)
    if SOURCE_SHEET_NAME not in workbook.sheetnames:
        if workbook.sheetnames == []:
            return NcSourceWorkbook(
                path=path,
                label=label,
                worksheet=None,
                total_row=NC_DETAIL_START_ROW - 1,
                detail_signature=empty_detail_signature(),
                is_empty=True,
            )
        raise RuntimeError(
            f"El archivo fuente {label} debe contener la hoja '{SOURCE_SHEET_NAME}'. "
            "No se aceptan plantillas ni salidas ya generadas."
        )

    worksheet = workbook[SOURCE_SHEET_NAME]
    assert_source_headers(worksheet, label)
    total_row = find_row_containing(worksheet, "TOTAL GENERAL", 1, max(worksheet.max_row, 200), NC_MAX_COLUMN)
    if total_row is None:
        raise RuntimeError(f"No se encontro TOTAL GENERAL en la fuente {label}.")

    return NcSourceWorkbook(
        path=path,
        label=label,
        worksheet=worksheet,
        total_row=total_row,
        detail_signature=build_detail_signature(worksheet, total_row),
    )


def copy_template_value_row(source: Worksheet, target: Worksheet, source_row: int, target_row: int) -> None:
    for column in range(1, NC_MAX_COLUMN + 1):
        target_cell = target.cell(row=target_row, column=column)
        if isinstance(target_cell, MergedCell):
            continue

        source_cell = source.cell(row=source_row, column=column)
        value = source_cell.value
        if isinstance(value, str):
            write_literal_string(target_cell, value)
        else:
            target_cell.value = value

    for column in (23, 24):
        target_cell = target.cell(row=target_row, column=column)
        if not isinstance(target_cell, MergedCell) and target_cell.value in (None, ""):
            target_cell.value = 0


def ensure_nc_sheet_capacity(
    target: Worksheet,
    template_total_row: int,
    template_mayor_row: int,
    required_total_row: int,
) -> tuple[int, int, int]:
    added_rows = max(0, required_total_row - template_total_row)
    if added_rows > 0:
        target.insert_rows(template_total_row, amount=added_rows)

    return template_total_row + added_rows, template_mayor_row + added_rows, added_rows


def copy_source_to_nc_sheet_with_template(
    source: NcSourceWorkbook,
    target: Worksheet,
    template_values_sheet: Worksheet,
) -> tuple[int, int, int]:
    template_total_row = find_row_containing(
        target,
        "TOTAL GENERAL",
        1,
        max(target.max_row, 200),
        NC_MAX_COLUMN,
    )
    template_mayor_row = find_row_containing(
        target,
        "MAYOR",
        1,
        max(target.max_row, 200),
        NC_MAX_COLUMN,
    )

    if template_total_row is None or template_mayor_row is None:
        raise RuntimeError(f"La hoja {target.title} no tiene filas base TOTAL GENERAL/MAYOR.")

    detail_style_row = find_style_source_row(
        target,
        NC_DETAIL_START_ROW,
        max(NC_DETAIL_START_ROW, template_total_row - 1),
        NC_MAX_COLUMN,
    )

    if source.is_empty:
        return apply_empty_source_to_nc_sheet(
            target,
            template_total_row,
            template_mayor_row,
            detail_style_row,
            template_values_sheet,
        )

    shifted_total_row, shifted_mayor_row, added_rows = ensure_nc_sheet_capacity(
        target,
        template_total_row,
        template_mayor_row,
        source.total_row,
    )
    affected_rows = unmerge_ranges_in_band(
        target,
        NC_DETAIL_START_ROW,
        source.total_row,
        NC_MAX_COLUMN,
    )
    old_last_row = max(target.max_row, shifted_mayor_row, 200)

    for row in range(1, source.total_row + 1):
        if row >= NC_DETAIL_START_ROW:
            if row < source.total_row and (row >= template_total_row or row in affected_rows):
                copy_row_style(target, target, detail_style_row, row, NC_MAX_COLUMN)
            elif row == source.total_row and (row != shifted_total_row or row in affected_rows):
                copy_row_style(target, target, shifted_total_row, row, NC_MAX_COLUMN)

        for column in range(1, NC_MAX_COLUMN + 1):
            cell = target.cell(row=row, column=column)
            if isinstance(cell, MergedCell):
                continue

            source_cell = source.worksheet.cell(row=row, column=column)
            value = source_cell_value(source_cell)
            if isinstance(value, str):
                write_literal_string(cell, value)
            else:
                cell.value = value

    actual_mayor_row = prepare_marker_row(
        target,
        shifted_mayor_row,
        source.total_row + 1,
        NC_MAX_COLUMN,
        22,
    )
    clear_rows(
        target,
        actual_mayor_row + 1,
        max(old_last_row, source.total_row + added_rows + 10),
        NC_MAX_COLUMN,
    )

    return template_total_row, template_mayor_row, actual_mayor_row


def run(request: ProcessRequest) -> ProcessResult:
    template_path = request.template_path.resolve() if request.template_path else None
    if template_path is None or not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    saved_inputs = request.options.get("saved_inputs", {})
    if not isinstance(saved_inputs, dict):
        raise ValueError("saved_inputs must be a mapping.")

    workbook = load_workbook_quiet(template_path, data_only=False, keep_links=True)
    template_values = load_workbook_quiet(template_path, data_only=True, keep_links=True)
    summary: list[dict[str, int | str]] = []
    nc_metadata: dict[str, Any] = {}

    for config in SHEET_CONFIGS:
        input_payload = saved_inputs.get(config["field"])
        if not isinstance(input_payload, dict):
            raise RuntimeError(f"No se encontro input guardado para {config['field']}.")

        source_path = Path(str(input_payload.get("path", ""))).resolve()
        if not source_path.is_file():
            raise FileNotFoundError(f"Input not found for {config['field']}: {source_path}")

        source = load_source_workbook(source_path, str(config["label"]))
        if config["target_sheet"] not in workbook.sheetnames:
            raise RuntimeError(f"No existe la hoja requerida en plantilla: {config['target_sheet']}")

        target = workbook[str(config["target_sheet"])]
        template_values_sheet = template_values[str(config["target_sheet"])]
        total_row, mayor_row, actual_mayor_row = copy_source_to_nc_sheet_with_template(source, target, template_values_sheet)
        copy_template_value_row(template_values_sheet, target, mayor_row, actual_mayor_row)
        target_signature = build_detail_signature(target, source.total_row)
        if target_signature != source.detail_signature:
            raise RuntimeError(
                f"La hoja {target.title} no conserva el detalle NC del archivo subido. "
                f"Filas fuente={source.detail_signature[0]}, filas salida={target_signature[0]}."
            )

        summary.append({"label": str(config["label"]), "rows": source.detail_signature[0]})
        nc_metadata[str(config["key"])] = {
            "target_sheet": str(config["target_sheet"]),
            "rows": source.detail_signature[0],
            "source_total_row": source.total_row,
            "template_total_row": total_row,
            "template_mayor_row": mayor_row,
            "actual_mayor_row": actual_mayor_row,
            "detail_hash": source.detail_signature[1],
            "empty_source": source.is_empty,
        }

    output_path = write_workbook_with_retries(workbook, request.output_path.resolve())
    return ProcessResult(
        success=True,
        output_path=output_path,
        label="repuestos_tytserv_nc_stage",
        metadata={
            "runtime": "python-native-nc-stage",
            "summary": summary,
            "nc_stage": nc_metadata,
            "console": "Etapa NC generada desde Python nativo.",
            "output_origin": "default_path",
            "fallback_used": False,
        },
    )
