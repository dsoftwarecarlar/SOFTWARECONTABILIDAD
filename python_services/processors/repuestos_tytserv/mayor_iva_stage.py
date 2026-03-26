from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.worksheet import Worksheet

from ..contracts import ProcessRequest, ProcessResult
from ..cxp_actions.accion3_native import write_workbook_with_retries
from .my_stage import (
    SOURCE_CONFIGS,
    build_nc_groups_from_rows,
    build_nc_rows_from_source,
    build_rep_posting_groups,
    clear_row_values,
    compare_posting_key,
    date_sort_key,
    load_optional_nc_source,
    load_rep_source_workbook,
    natural_key,
    resolve_saved_input,
    round_amount,
    seat_sort_key,
    sum_column,
    write_date_cell,
)
from .workbook_tools import copy_row_style, load_workbook_quiet

DETAIL_START_ROW = 299
DETAIL_END_ROW = 366
LAST_COLUMN = 10
SUMMARY_ROW_START = 367


def get_family_order(detail: str) -> int:
    normalized = detail.strip().upper()
    if normalized == "MOD. REPUESTOS REP01":
        return 0
    if normalized in {"MOD. REPUESTOS REP07", "MOD. REPUESTOS REP08"}:
        return 1
    if normalized == "MOD. REPUESTOS REP06":
        return 2
    if normalized == "MOD. REPUESTOS REP05":
        return 3
    return 99


def get_detail_order(detail: str) -> int:
    normalized = detail.strip().upper()
    if normalized == "MOD. REPUESTOS REP07":
        return 0
    if normalized == "MOD. REPUESTOS REP08":
        return 1
    return 0


def compare_mayor_iva_key(entry: dict[str, Any]) -> tuple[Any, ...]:
    return (
        get_family_order(entry["detail"]),
        date_sort_key(entry["date_value"]),
        seat_sort_key(entry["seat"]),
        get_detail_order(entry["detail"]),
        natural_key(entry["side"]),
    )


def build_mayor_iva_entries(
    rep_groups_by_key: dict[str, dict[str, dict[str, dict[str, Any]]]],
    nc_groups_by_key: dict[str, dict[str, dict[str, dict[str, Any]]]],
) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []

    for rep_groups in rep_groups_by_key.values():
        for group in (rep_groups.get("vat") or {}).values():
            amount = round_amount(group.get("amount", 0), 2)
            if abs(amount) < 0.0000001:
                continue
            entries.append(
                {
                    "side": "credit",
                    "amount": amount,
                    "date_value": group.get("date_value"),
                    "date_text": group.get("date_text", ""),
                    "seat": group.get("seat", ""),
                    "detail": group.get("detail", ""),
                }
            )

    for nc_groups in nc_groups_by_key.values():
        for group in (nc_groups.get("vat") or {}).values():
            amount = round_amount(group.get("amount", 0), 2)
            if abs(amount) < 0.0000001:
                continue
            entries.append(
                {
                    "side": "debit",
                    "amount": amount,
                    "date_value": group.get("date_value"),
                    "date_text": group.get("date_text", ""),
                    "seat": group.get("seat", ""),
                    "detail": group.get("detail", ""),
                }
            )

    return [
        entry
        for entry in sorted(entries, key=compare_mayor_iva_key)
        if str(entry["seat"]).strip() != "" and str(entry["detail"]).strip() != ""
    ]


def get_opening_balance(template_values_sheet: Worksheet) -> float:
    saldo = template_values_sheet.cell(row=DETAIL_START_ROW, column=10).value or 0
    debit = template_values_sheet.cell(row=DETAIL_START_ROW, column=8).value or 0
    credit = template_values_sheet.cell(row=DETAIL_START_ROW, column=9).value or 0
    return round_amount(float(saldo) - float(debit) + float(credit), 2)


def clear_dynamic_columns(worksheet: Worksheet, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        for column in (4, 6, 7, 8, 9, 10):
            worksheet.cell(row=row, column=column).value = None if column in (4, 6, 7, 10) else 0


def copy_row_values(worksheet: Worksheet, source_row: int, target_row: int) -> None:
    for column in range(1, LAST_COLUMN + 1):
        source_cell = worksheet.cell(row=source_row, column=column)
        target_cell = worksheet.cell(row=target_row, column=column)
        if isinstance(source_cell, MergedCell) or isinstance(target_cell, MergedCell):
            continue
        target_cell.value = copy(source_cell.value)


def ensure_mayor_iva_capacity(worksheet: Worksheet, required_rows: int) -> dict[str, int]:
    base_capacity = DETAIL_END_ROW - DETAIL_START_ROW + 1
    extra_rows = max(0, required_rows - base_capacity)
    if extra_rows > 0:
        insert_row = DETAIL_END_ROW + 1
        worksheet.insert_rows(insert_row, amount=extra_rows)
        for row in range(insert_row, insert_row + extra_rows):
            copy_row_style(worksheet, worksheet, DETAIL_END_ROW, row, LAST_COLUMN)
            copy_row_values(worksheet, DETAIL_END_ROW, row)

    return {
        "detail_end_row": DETAIL_END_ROW + extra_rows,
        "summary_row_start": SUMMARY_ROW_START + extra_rows,
    }


def update_summary_rows(
    worksheet: Worksheet,
    *,
    detail_end_row: int = DETAIL_END_ROW,
    summary_start_row: int = SUMMARY_ROW_START,
) -> None:
    debit_all = sum_column(worksheet, 2, detail_end_row, 8)
    credit_all = sum_column(worksheet, 2, detail_end_row, 9)
    debit_window = sum_column(worksheet, 283, detail_end_row, 8)
    credit_window = sum_column(worksheet, 283, detail_end_row, 9)

    clear_row_values(worksheet, summary_start_row, LAST_COLUMN)
    clear_row_values(worksheet, summary_start_row + 1, LAST_COLUMN)
    clear_row_values(worksheet, summary_start_row + 2, LAST_COLUMN)
    clear_row_values(worksheet, summary_start_row + 3, LAST_COLUMN)

    worksheet.cell(row=summary_start_row, column=8).value = debit_all
    worksheet.cell(row=summary_start_row, column=9).value = credit_all
    worksheet.cell(row=summary_start_row + 1, column=9).value = round_amount(credit_all - debit_all, 2)
    worksheet.cell(row=summary_start_row + 2, column=8).value = debit_window
    worksheet.cell(row=summary_start_row + 2, column=9).value = credit_window
    worksheet.cell(row=summary_start_row + 3, column=9).value = round_amount(credit_window - debit_window, 2)


def run(request: ProcessRequest) -> ProcessResult:
    template_path = request.template_path.resolve() if request.template_path else None
    if template_path is None or not template_path.is_file():
        raise FileNotFoundError(f"Template not found: {template_path}")

    saved_inputs = request.options.get("saved_inputs", {})
    if not isinstance(saved_inputs, dict):
        raise ValueError("saved_inputs must be a mapping.")

    workbook = load_workbook_quiet(template_path, data_only=False, keep_links=True)
    template_values = load_workbook_quiet(template_path, data_only=True, keep_links=True)

    rep_groups_by_key: dict[str, dict[str, dict[str, dict[str, Any]]]] = {}
    nc_groups_by_key: dict[str, dict[str, dict[str, dict[str, Any]]]] = {}

    for config in SOURCE_CONFIGS:
        key = str(config["key"])
        rep_path = resolve_saved_input(saved_inputs, str(config["rep_field"]))
        if rep_path is None:
            raise FileNotFoundError(f"Input not found for {config['rep_field']}")
        rep_source = load_rep_source_workbook(rep_path, str(config["label"]))
        nc_path = resolve_saved_input(saved_inputs, str(config["nc_field"]))
        nc_source = load_optional_nc_source(nc_path, str(config["label"]), bool(config["optional_nc"]))

        rep_groups_by_key[key] = build_rep_posting_groups(rep_source, key)
        nc_groups_by_key[key] = build_nc_groups_from_rows(build_nc_rows_from_source(nc_source, key), key)

    worksheet = workbook["MAYOR IVA"]
    template_values_sheet = template_values["MAYOR IVA"]
    entries = build_mayor_iva_entries(rep_groups_by_key, nc_groups_by_key)
    layout = ensure_mayor_iva_capacity(worksheet, len(entries))
    detail_end_row = int(layout["detail_end_row"])
    summary_row_start = int(layout["summary_row_start"])

    clear_dynamic_columns(worksheet, DETAIL_START_ROW, detail_end_row)
    opening_balance = get_opening_balance(template_values_sheet)
    running_balance = opening_balance

    for index, entry in enumerate(entries):
        row_number = DETAIL_START_ROW + index
        write_date_cell(worksheet, row_number, 4, entry["date_value"], str(entry["date_text"]))
        worksheet.cell(row=row_number, column=6).value = entry["seat"]
        worksheet.cell(row=row_number, column=7).value = entry["detail"]
        worksheet.cell(row=row_number, column=8).value = entry["amount"] if entry["side"] == "debit" else 0
        worksheet.cell(row=row_number, column=9).value = entry["amount"] if entry["side"] == "credit" else 0
        running_balance = round_amount(
            running_balance
            + (entry["amount"] if entry["side"] == "debit" else 0)
            - (entry["amount"] if entry["side"] == "credit" else 0),
            2,
        )
        worksheet.cell(row=row_number, column=10).value = running_balance

    for row in range(DETAIL_START_ROW + len(entries), detail_end_row + 1):
        worksheet.cell(row=row, column=10).value = None

    update_summary_rows(
        worksheet,
        detail_end_row=detail_end_row,
        summary_start_row=summary_row_start,
    )

    output_path = request.output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_workbook_with_retries(workbook, output_path)
    return ProcessResult(
        success=True,
        output_path=output_path,
        label="repuestos_tytserv_mayor_iva_stage",
        metadata={
            "runtime": "python-native-mayor-iva-stage",
            "rows": len(entries),
            "detail_end_row": detail_end_row,
            "summary_row_start": summary_row_start,
        },
    )
